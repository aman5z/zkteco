"""
ZKTeco Attendance Report Generator
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  TODAY mode   → punches from devices, employee list from .mdb
  HISTORY mode → everything from .mdb, filtered by date range

Report format: Absent employees only
  Columns: Name | Emp Code | Department | Date
  Sorted by: Department

Requirements:
    pip install pyzk pyodbc openpyxl pandas

Usage:
    python attendance_tool.py                                          # today
    python attendance_tool.py history backup.mdb 01/03/2026 12/03/2026
"""

import sys
import os
import warnings
warnings.filterwarnings("ignore")
from datetime import datetime, date
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ═══════════════════════════════════════════════════════════════════════════════
#  CONFIG
# ═══════════════════════════════════════════════════════════════════════════════

DEVICE_IPS     = ["10.20.141.21", "10.20.141.22", "10.20.141.23", "10.20.141.24"]
DEVICE_PORT    = 4370
DEVICE_TIMEOUT = 10

SCRIPT_DIR     = os.path.dirname(os.path.abspath(__file__))
MDB_PATH       = os.path.join(SCRIPT_DIR, "your_backup.mdb")
OUTPUT_TODAY   = os.path.join(SCRIPT_DIR, "absent_today.xlsx")
OUTPUT_HISTORY = os.path.join(SCRIPT_DIR, "absent_history.xlsx")

# Only include these departments in reports (case-insensitive)
INCLUDE_DEPARTMENTS = ["admin", "teaching", "support", "cleaning staff", "transport"]

# Confirmed table/column names from your .mdb
CHECKINOUT_TABLE = "CHECKINOUT"
USERINFO_TABLE   = "USERINFO"
DEPT_TABLE       = "DEPARTMENTS"
COL_USER_ID      = "USERID"
COL_CHECKTIME    = "CHECKTIME"
COL_EMP_NAME     = "Name"
COL_BADGE        = "Badgenumber"
COL_DEPT_ID      = "DEFAULTDEPTID"
COL_ATT_FLAG     = "ATT"
COL_DEPT_NAME    = "DEPTNAME"
COL_DEPT_ID_PK   = "DEPTID"

# ═══════════════════════════════════════════════════════════════════════════════


CLR_HEADER_BG = "1F4E79"
CLR_HEADER_FG = "FFFFFF"
CLR_ALT_ROW   = "DCE6F1"


def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def header_cell(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=True, color=CLR_HEADER_FG, size=10)
    c.fill      = PatternFill("solid", start_color=CLR_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = thin_border()
    return c


def data_cell(ws, row, col, value, bg=None, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", size=10)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = thin_border()
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    return c


# ═══════════════════════════════════════════════════════════════════════════════
#  DATABASE
# ═══════════════════════════════════════════════════════════════════════════════

def connect_mdb(mdb_path):
    try:
        import pyodbc
        return pyodbc.connect(
            r"Driver={Microsoft Access Driver (*.mdb, *.accdb)};"
            f"Dbq={mdb_path};"
        )
    except ImportError:
        print("[ERROR] pyodbc not installed. Run: pip install pyodbc")
        sys.exit(1)
    except Exception as e:
        print(f"[ERROR] Cannot open database: {e}")
        print("Install driver: https://www.microsoft.com/en-us/download/details.aspx?id=54920")
        sys.exit(1)


def load_employees(conn):
    """Load active employees (ATT=1) with department names."""
    emp_df = pd.read_sql(
        f"SELECT [{COL_USER_ID}], [{COL_EMP_NAME}], [{COL_BADGE}], "
        f"[{COL_DEPT_ID}], [{COL_ATT_FLAG}] FROM [{USERINFO_TABLE}]",
        conn
    )
    dept_df = pd.read_sql(
        f"SELECT [{COL_DEPT_ID_PK}], [{COL_DEPT_NAME}] FROM [{DEPT_TABLE}]",
        conn
    )

    # Active employees only
    emp_df = emp_df[emp_df[COL_ATT_FLAG] == 1].copy()
    emp_df[COL_USER_ID]  = emp_df[COL_USER_ID].astype(str).str.strip()
    emp_df[COL_BADGE]    = emp_df[COL_BADGE].astype(str).str.strip()
    emp_df[COL_DEPT_ID]  = emp_df[COL_DEPT_ID].astype(str).str.strip()
    dept_df[COL_DEPT_ID_PK] = dept_df[COL_DEPT_ID_PK].astype(str).str.strip()

    # Merge department names
    emp_df = emp_df.merge(
        dept_df, left_on=COL_DEPT_ID, right_on=COL_DEPT_ID_PK, how="left"
    )
    emp_df[COL_DEPT_NAME] = emp_df[COL_DEPT_NAME].fillna("Unknown")

    # Filter to included departments only
    emp_df = emp_df[
        emp_df[COL_DEPT_NAME].str.strip().str.lower().isin(
            [d.lower() for d in INCLUDE_DEPARTMENTS]
        )
    ].copy()

    print(f"Active employees loaded: {len(emp_df)} (filtered to selected departments)")
    return emp_df


# ═══════════════════════════════════════════════════════════════════════════════
#  ABSENT CALCULATION
# ═══════════════════════════════════════════════════════════════════════════════

def get_absent_records(emp_df, present_ids_by_date):
    """
    present_ids_by_date: dict of {date: set of user_id strings}
    Returns a DataFrame of absent records sorted by date then department.
    """
    rows = []
    for d, present_ids in sorted(present_ids_by_date.items()):
        for _, emp in emp_df.iterrows():
            uid = emp[COL_USER_ID]
            if uid not in present_ids:
                rows.append({
                    "Date":       d,
                    "Name":       emp[COL_EMP_NAME],
                    "Emp Code":   emp[COL_BADGE],
                    "Department": emp[COL_DEPT_NAME],
                })

    if not rows:
        return pd.DataFrame(columns=["Date", "Name", "Emp Code", "Department"])

    absent_df = pd.DataFrame(rows)
    absent_df = absent_df.sort_values(["Date", "Department", "Name"]).reset_index(drop=True)
    return absent_df


# ═══════════════════════════════════════════════════════════════════════════════
#  EXCEL WRITER
# ═══════════════════════════════════════════════════════════════════════════════

def write_absent_excel(absent_df, output_path, title_suffix=""):
    wb = Workbook()
    wb.remove(wb.active)

    dates = sorted(absent_df["Date"].unique()) if not absent_df.empty else []

    # ── Summary sheet ──────────────────────────────────────────────────────────
    sum_ws = wb.create_sheet("Summary", 0)
    sum_ws.sheet_view.showGridLines = False

    sum_ws.merge_cells("A1:D1")
    c = sum_ws.cell(row=1, column=1, value=f"Absent Report — {title_suffix}")
    c.font      = Font(name="Arial", bold=True, size=14, color=CLR_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    sum_ws.row_dimensions[1].height = 30

    sum_ws.cell(row=2, column=1,
        value=f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}").font = \
        Font(name="Arial", size=9, italic=True, color="888888")

    row = 4
    header_cell(sum_ws, row, 1, "Date")
    header_cell(sum_ws, row, 2, "Total Absent")
    sum_ws.row_dimensions[row].height = 20
    row += 1

    for i, d in enumerate(dates):
        bg    = CLR_ALT_ROW if i % 2 == 0 else "FFFFFF"
        count = len(absent_df[absent_df["Date"] == d])
        data_cell(sum_ws, row, 1, d.strftime("%d %b %Y") if hasattr(d, "strftime") else str(d), bg=bg, align="center")
        data_cell(sum_ws, row, 2, count, bg=bg, align="center")
        row += 1

    sum_ws.column_dimensions["A"].width = 18
    sum_ws.column_dimensions["B"].width = 16

    # ── One sheet per date ─────────────────────────────────────────────────────
    for d in dates:
        day_absent = absent_df[absent_df["Date"] == d].reset_index(drop=True)
        sheet_name = d.strftime("%d-%b-%Y") if hasattr(d, "strftime") else str(d)
        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:D1")
        label = d.strftime("%A, %d %B %Y") if hasattr(d, "strftime") else str(d)
        c = ws.cell(row=1, column=1, value=f"Absent Employees — {label}")
        c.font      = Font(name="Arial", bold=True, size=13, color=CLR_HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        ws.cell(row=2, column=1, value=f"Total Absent: {len(day_absent)}").font = \
            Font(name="Arial", size=10, bold=True, color="9C0006")
        ws.row_dimensions[2].height = 18

        row = 4
        header_cell(ws, row, 1, "Name")
        header_cell(ws, row, 2, "Emp Code")
        header_cell(ws, row, 3, "Department")
        header_cell(ws, row, 4, "Date")
        ws.row_dimensions[row].height = 18
        row += 1

        prev_dept = None
        for i, rec in day_absent.iterrows():
            dept = rec["Department"]
            # Light separator row when department changes
            if dept != prev_dept and prev_dept is not None:
                for col in range(1, 5):
                    c = ws.cell(row=row, column=col, value="")
                    c.fill   = PatternFill("solid", start_color="F2F2F2")
                    c.border = thin_border()
                ws.row_dimensions[row].height = 6
                row += 1
            prev_dept = dept

            bg = CLR_ALT_ROW if i % 2 == 0 else "FFFFFF"
            data_cell(ws, row, 1, rec["Name"],       bg=bg)
            data_cell(ws, row, 2, rec["Emp Code"],   bg=bg, align="center")
            data_cell(ws, row, 3, rec["Department"], bg=bg)
            data_cell(ws, row, 4,
                d.strftime("%d/%m/%Y") if hasattr(d, "strftime") else str(d),
                bg=bg, align="center")
            row += 1

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 24
        ws.column_dimensions["D"].width = 14

    wb.save(output_path)


# ═══════════════════════════════════════════════════════════════════════════════
#  TODAY MODE
# ═══════════════════════════════════════════════════════════════════════════════

def pull_punches_from_devices():
    try:
        from zk import ZK
    except ImportError:
        print("[ERROR] pyzk not installed. Run: pip install pyzk")
        sys.exit(1)

    today       = date.today()
    present_ids = set()

    for ip in DEVICE_IPS:
        zk   = ZK(ip, port=DEVICE_PORT, timeout=DEVICE_TIMEOUT, verbose=False)
        conn = None
        try:
            print(f"  Connecting to {ip}:{DEVICE_PORT} ...", end=" ", flush=True)
            conn = zk.connect()
            conn.disable_device()
            records = conn.get_attendance()
            day_hits = 0
            for r in records:
                if r.timestamp.date() == today:
                    present_ids.add(str(r.user_id).strip())
                    day_hits += 1
            conn.enable_device()
            print(f"OK  ({day_hits} punches today)")
        except Exception as e:
            print(f"FAILED — {e}")
        finally:
            if conn:
                try: conn.disconnect()
                except: pass

    return present_ids


def run_today():
    today = date.today()
    print(f"\n{'═'*55}")
    print(f"  TODAY'S ABSENT REPORT  —  {today.strftime('%d %B %Y')}")
    print(f"{'═'*55}")

    # Employee list from .mdb
    if not os.path.exists(MDB_PATH):
        print(f"[ERROR] .mdb file not found: {MDB_PATH}")
        print("Place your backup .mdb in the same folder as this script.")
        sys.exit(1)

    print(f"\nLoading employees from .mdb ...")
    conn   = connect_mdb(MDB_PATH)
    emp_df = load_employees(conn)
    conn.close()

    # Punches from devices
    print(f"\nPulling today's punches from {len(DEVICE_IPS)} device(s) ...\n")
    present_ids = pull_punches_from_devices()
    print(f"\nPresent (punched): {len(present_ids)}")

    absent_df = get_absent_records(emp_df, {today: present_ids})
    print(f"Absent           : {len(absent_df)}")

    write_absent_excel(absent_df, OUTPUT_TODAY,
                       title_suffix=today.strftime("%d %B %Y"))
    print(f"\n✓ Report saved → {OUTPUT_TODAY}")


# ═══════════════════════════════════════════════════════════════════════════════
#  HISTORY MODE
# ═══════════════════════════════════════════════════════════════════════════════

def run_history(mdb_path, date_from, date_to):
    print(f"\n{'═'*55}")
    print(f"  ABSENT HISTORY REPORT")
    print(f"  Range: {date_from.strftime('%d %b %Y')} → {date_to.strftime('%d %b %Y')}")
    print(f"{'═'*55}\n")

    conn   = connect_mdb(mdb_path)
    emp_df = load_employees(conn)

    # Load punch records for date range only
    from_str = date_from.strftime("%Y-%m-%d")
    to_str   = date_to.strftime("%Y-%m-%d")

    punch_df = pd.read_sql(
        f"SELECT [{COL_USER_ID}], [{COL_CHECKTIME}] FROM [{CHECKINOUT_TABLE}] "
        f"WHERE [{COL_CHECKTIME}] >= #{from_str}# AND [{COL_CHECKTIME}] <= #{to_str} 23:59:59#",
        conn
    )
    conn.close()

    punch_df[COL_CHECKTIME] = pd.to_datetime(punch_df[COL_CHECKTIME], errors="coerce")
    punch_df = punch_df.dropna(subset=[COL_CHECKTIME])
    punch_df["_date"] = punch_df[COL_CHECKTIME].dt.date
    punch_df["_uid"]  = punch_df[COL_USER_ID].astype(str).str.strip()

    print(f"Punch records in range: {len(punch_df)}")

    # Build present_ids_by_date
    present_by_date = {}
    all_dates = pd.date_range(date_from, date_to).date
    for d in all_dates:
        day_df = punch_df[punch_df["_date"] == d]
        present_by_date[d] = set(day_df["_uid"].unique())

    absent_df = get_absent_records(emp_df, present_by_date)
    print(f"Total absent records: {len(absent_df)}")

    range_label = f"{date_from.strftime('%d %b')} – {date_to.strftime('%d %b %Y')}"
    write_absent_excel(absent_df, OUTPUT_HISTORY, title_suffix=range_label)
    print(f"\n✓ Report saved → {OUTPUT_HISTORY}")


# ═══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    mode = sys.argv[1].lower() if len(sys.argv) > 1 else "today"

    if mode == "today":
        run_today()
    elif mode == "history":
        mdb = sys.argv[2] if len(sys.argv) > 2 else MDB_PATH
        if not os.path.exists(mdb):
            print(f"[ERROR] File not found: {mdb}")
            sys.exit(1)
        if len(sys.argv) < 5:
            print("[ERROR] Please provide start and end dates.")
            print("Usage: python attendance_tool.py history backup.mdb DD/MM/YYYY DD/MM/YYYY")
            sys.exit(1)
        try:
            date_from = datetime.strptime(sys.argv[3], "%d/%m/%Y").date()
            date_to   = datetime.strptime(sys.argv[4], "%d/%m/%Y").date()
        except ValueError:
            print("[ERROR] Invalid date format. Use DD/MM/YYYY")
            sys.exit(1)
        run_history(mdb, date_from, date_to)
    else:
        print("Usage:")
        print("  python attendance_tool.py")
        print("  python attendance_tool.py history backup.mdb 01/03/2026 12/03/2026")
        sys.exit(1)
