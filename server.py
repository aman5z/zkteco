# -*- coding: utf-8 -*-
"""
ZKTeco Attendance Dashboard -- Flask Backend  v2.2
Merged: server.py + db_manager.py
Run: python server.py  ->  open http://localhost:5000

New in v2:
  - Login system (admin + employee logins)
  - Role-based access control (admin configures per-user permissions)
  - Device rename (friendly names stored, IP always shown)
  - Employee calendar API (punch times per day, source selectable)
  - Light/dark theme support (stored server-side per user)
  - Departments collapsible in employees tab
  - Bug fixes: MDB date query, URL typos, duplicate sys import
"""

import os
import sys
import html
import warnings
import threading
import time
import io
import json
import hashlib
import binascii
import shutil
import socket
import queue
import configparser
import sqlite3
from datetime import datetime, date, timedelta
import calendar
from concurrent.futures import ThreadPoolExecutor, as_completed
from functools import wraps
warnings.filterwarnings("ignore")


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

from flask import Flask, jsonify, request, send_from_directory, Response, send_file, session
# ==============================================================================
#  DATABASE LAYER  (merged from db_manager.py)
# ==============================================================================
DB_PATH = os.path.join(SCRIPT_DIR, 'attendance.db')


DB_SCHEMA = """
CREATE TABLE IF NOT EXISTS employees (
    badge       TEXT PRIMARY KEY,
    name        TEXT NOT NULL DEFAULT '',
    dept        TEXT NOT NULL DEFAULT '',
    active      INTEGER NOT NULL DEFAULT 1,
    updated_at  TEXT
);
CREATE TABLE IF NOT EXISTS punches (
    badge       TEXT NOT NULL,
    punch_time  TEXT NOT NULL,
    device_ip   TEXT NOT NULL DEFAULT '',
    PRIMARY KEY (badge, punch_time, device_ip)
);
CREATE TABLE IF NOT EXISTS unknown_users (
    device_ip   TEXT NOT NULL,
    uid         TEXT NOT NULL,
    seen_at     TEXT NOT NULL,
    resolved    INTEGER NOT NULL DEFAULT 0,
    PRIMARY KEY (device_ip, uid)
);
CREATE TABLE IF NOT EXISTS settings (
    key   TEXT PRIMARY KEY,
    value TEXT
);
CREATE TABLE IF NOT EXISTS audit_log (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    ts          TEXT NOT NULL,
    username    TEXT NOT NULL DEFAULT '',
    action      TEXT NOT NULL,
    detail      TEXT DEFAULT '',
    ip_addr     TEXT DEFAULT ''
);
CREATE TABLE IF NOT EXISTS shift_times (
    dept        TEXT PRIMARY KEY,
    shift_start TEXT NOT NULL DEFAULT '07:30',
    shift_end   TEXT NOT NULL DEFAULT '15:00',
    grace_mins  INTEGER NOT NULL DEFAULT 15
);
CREATE TABLE IF NOT EXISTS dept_workdays (
    dept        TEXT PRIMARY KEY,
    workdays    TEXT NOT NULL DEFAULT '[0,1,2,3,6]'
);
CREATE TABLE IF NOT EXISTS emp_workdays (
    badge       TEXT PRIMARY KEY,
    workdays    TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS holidays (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    date        TEXT NOT NULL,
    date_end    TEXT NOT NULL,
    label       TEXT NOT NULL DEFAULT 'Holiday',
    scope       TEXT NOT NULL DEFAULT 'all',
    dept        TEXT NOT NULL DEFAULT '',
    employees   TEXT NOT NULL DEFAULT '',
    created_at  TEXT
);
CREATE INDEX IF NOT EXISTS idx_holidays_date ON holidays(date);
CREATE INDEX IF NOT EXISTS idx_punches_badge ON punches(badge);
CREATE INDEX IF NOT EXISTS idx_punches_time  ON punches(punch_time);
CREATE INDEX IF NOT EXISTS idx_audit_ts      ON audit_log(ts);

CREATE TABLE IF NOT EXISTS announcements (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    message     TEXT NOT NULL,
    active      INTEGER NOT NULL DEFAULT 1,
    created_at  TEXT NOT NULL,
    created_by  TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS messages (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    sender      TEXT NOT NULL,
    receiver    TEXT NOT NULL,
    message     TEXT NOT NULL,
    timestamp   TEXT NOT NULL,
    is_read     INTEGER NOT NULL DEFAULT 0
);

CREATE TABLE IF NOT EXISTS notes (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    user_badge  TEXT NOT NULL,
    text        TEXT NOT NULL,
    note_type   TEXT NOT NULL DEFAULT 'note',
    timestamp   TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS tickets (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    title           TEXT NOT NULL,
    desc            TEXT NOT NULL,
    priority        TEXT NOT NULL DEFAULT 'Medium',
    status          TEXT NOT NULL DEFAULT 'Open',
    requester_badge TEXT NOT NULL,
    assigned_to     TEXT,
    due_time        TEXT,
    created_at      TEXT NOT NULL
);
"""

def get_db():
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    return conn

def init_db():
    conn = get_db()
    conn.executescript(DB_SCHEMA)
    try: conn.execute("ALTER TABLE audit_log ADD COLUMN detail_json TEXT DEFAULT '{}'")
    except Exception: pass
    try: conn.execute("ALTER TABLE audit_log ADD COLUMN user_agent TEXT DEFAULT ''")
    except Exception: pass
    conn.commit()
    conn.close()
    print("[DB] attendance.db ready: {0}".format(DB_PATH))
    return DB_PATH

# -- Employees -----------------------------------------------------------------
def upsert_employees(emp_list):
    """emp_list: list of (badge, name, dept, active)"""
    conn = get_db()
    now  = datetime.now().isoformat()
    conn.executemany(
        "INSERT OR REPLACE INTO employees (badge, name, dept, active, updated_at) VALUES (?,?,?,?,?)",
        [(e[0], e[1], e[2], e[3], now) for e in emp_list]
    )
    conn.commit()
    n = conn.execute("SELECT COUNT(*) FROM employees").fetchone()[0]
    conn.close()
    return n

def get_employees(active_only=False):
    conn = get_db()
    if active_only:
        rows = conn.execute("SELECT badge, name, dept FROM employees WHERE active=1 ORDER BY dept, name").fetchall()
    else:
        rows = conn.execute("SELECT badge, name, dept, active FROM employees ORDER BY dept, name").fetchall()
    conn.close()
    return [dict(r) for r in rows]

def get_employee(badge):
    conn = get_db()
    row  = conn.execute("SELECT badge, name, dept FROM employees WHERE badge=?", (badge,)).fetchone()
    conn.close()
    return dict(row) if row else None

# -- Punches -------------------------------------------------------------------
def store_punches(punch_list):
    """punch_list: list of (badge, punch_time_isostr, device_ip)"""
    if not punch_list: return 0
    conn = get_db()
    conn.executemany(
        "INSERT OR IGNORE INTO punches (badge, punch_time, device_ip) VALUES (?,?,?)",
        punch_list
    )
    inserted = conn.total_changes
    conn.commit()
    conn.close()
    return inserted

def get_punch_count():
    conn = get_db()
    n    = conn.execute("SELECT COUNT(*) FROM punches").fetchone()[0]
    conn.close()
    return n

def get_punches_for_day(day_date):
    """Returns set of badges that punched on given date."""
    start = day_date.strftime("%Y-%m-%d") + " 00:00:00"
    end   = day_date.strftime("%Y-%m-%d") + " 23:59:59"
    conn  = get_db()
    rows  = conn.execute(
        "SELECT DISTINCT badge FROM punches WHERE punch_time >= ? AND punch_time <= ?",
        (start, end)
    ).fetchall()
    conn.close()
    return set(r[0] for r in rows)

def get_punch_records_for_employee(badge, date_from, date_to):
    """Returns list of {punch_time, device_ip} for an employee in date range."""
    start = date_from.strftime("%Y-%m-%d") + " 00:00:00"
    end   = date_to.strftime("%Y-%m-%d")   + " 23:59:59"
    conn  = get_db()
    rows  = conn.execute(
        "SELECT punch_time, device_ip FROM punches WHERE badge=? AND punch_time >= ? AND punch_time <= ? ORDER BY punch_time",
        (badge, start, end)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]

# -- Unknown users (new device enrollments) ------------------------------------
def record_unknown_user(device_ip, uid):
    conn = get_db()
    conn.execute(
        "INSERT OR IGNORE INTO unknown_users (device_ip, uid, seen_at) VALUES (?,?,?)",
        (device_ip, uid, datetime.now().isoformat())
    )
    conn.commit()
    conn.close()

def get_unknown_users():
    conn = get_db()
    rows = conn.execute(
        "SELECT device_ip, uid, seen_at FROM unknown_users WHERE resolved=0 ORDER BY seen_at DESC"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def resolve_unknown_user(device_ip, uid):
    conn = get_db()
    conn.execute("UPDATE unknown_users SET resolved=1 WHERE device_ip=? AND uid=?", (device_ip, uid))
    conn.commit()
    conn.close()

# -- Settings ------------------------------------------------------------------
def get_setting(key, default=None):
    try:
        conn = get_db()
        row  = conn.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
        conn.close()
        return row[0] if row else default
    except Exception:
        return default

def set_setting(key, value):
    conn = get_db()
    conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?,?)", (key, str(value)))
    conn.commit()
    conn.close()
# -- DB stats ------------------------------------------------------------------
def get_db_stats():
    conn = get_db()
    emps   = conn.execute("SELECT COUNT(*) FROM employees").fetchone()[0]
    active = conn.execute("SELECT COUNT(*) FROM employees WHERE active=1").fetchone()[0]
    punch  = conn.execute("SELECT COUNT(*) FROM punches").fetchone()[0]
    first  = conn.execute("SELECT MIN(punch_time) FROM punches").fetchone()[0]
    last   = conn.execute("SELECT MAX(punch_time) FROM punches").fetchone()[0]
    unk    = conn.execute("SELECT COUNT(*) FROM unknown_users WHERE resolved=0").fetchone()[0]
    conn.close()
    size   = os.path.getsize(DB_PATH) if os.path.exists(DB_PATH) else 0
    return {
        "db_path":       DB_PATH,
        "size_mb":       round(size / 1048576, 2),
        "employees":     emps,
        "active":        active,
        "punch_records": punch,
        "first_punch":   first,
        "last_punch":    last,
        "unknown_users": unk,
    }

# -- MDB import ----------------------------------------------------------------
def import_employees_from_csv(csv_path, exclude_depts=None):
    """Import employees from the employees_export.csv into SQLite."""
    import csv
    if not os.path.exists(csv_path):
        raise FileNotFoundError("CSV not found: {0}".format(csv_path))
    exclude = set(d.upper() for d in (exclude_depts or []))
    emp_list = []
    with open(csv_path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            badge = str(row.get("Badgenumber", "")).strip()
            name  = str(row.get("Name", "")).strip()
            dept  = str(row.get("DEPTNAME", "")).strip()
            if not badge or badge == "nan" or not name or name == "nan":
                continue
            active = 0 if dept.upper() in exclude else 1
            emp_list.append((badge, name, dept, active))
    n = upsert_employees(emp_list)
    print("[DB] Imported {0} employees from CSV".format(n))
    return n

def import_punches_from_mdb(mdb_path, uid_to_badge, checkinout_table="CHECKINOUT",
                             col_uid="USERID", col_time="CHECKTIME"):
    """Read historical punches from MDB CHECKINOUT and store to SQLite."""
    import pyodbc
    path = os.path.abspath(mdb_path)
    conn = pyodbc.connect(
        "Driver={{Microsoft Access Driver (*.mdb, *.accdb)}};"
        "Dbq={0};Uid=Admin;Pwd=;".format(path)
    )
    cursor = conn.cursor()
    cursor.execute("SELECT [{0}],[{1}] FROM [{2}]".format(col_uid, col_time, checkinout_table))
    punch_list = []
    unknown    = set()
    for row in cursor.fetchall():
        uid        = str(row[0]).strip()
        badge      = uid_to_badge.get(uid)
        punch_time = row[1]
        if not badge:
            unknown.add(uid)
            continue
        if hasattr(punch_time, "strftime"):
            ts = punch_time.strftime("%Y-%m-%d %H:%M:%S")
        else:
            ts = str(punch_time)
        punch_list.append((badge, ts, "mdb-import"))
    conn.close()
    inserted = store_punches(punch_list)
    print("[DB] MDB import: {0} punches stored, {1} unknown UIDs skipped".format(
        inserted, len(unknown)))
    return inserted, len(unknown)

# -- Audit Log -----------------------------------------------------------------
def write_audit(username, action, detail='', ip_addr=None, detail_json='{}', user_agent=None):
    try:
        from flask import request, has_request_context
        if has_request_context():
            if ip_addr is None or ip_addr == '': ip_addr = request.remote_addr or ''
            if user_agent is None or user_agent == '': user_agent = request.environ.get('HTTP_USER_AGENT', '')[:200]
        else:
            ip_addr = ip_addr or ''
            user_agent = user_agent or ''
            
        conn = get_db()
        conn.execute(
            "INSERT INTO audit_log (ts, username, action, detail, ip_addr, detail_json, user_agent) VALUES (?,?,?,?,?,?,?)",
            (datetime.now().isoformat(), username, action, str(detail)[:500], ip_addr, detail_json, user_agent)
        )
        conn.commit(); conn.close()
    except Exception as e: print("[Audit] Error:", e)

def get_audit_log(limit=200):
    conn = get_db()
    rows = conn.execute(
        "SELECT id, ts, username, action, detail, ip_addr, detail_json, user_agent FROM audit_log ORDER BY id DESC LIMIT ?",
        (limit,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]

# -- Shift times ---------------------------------------------------------------
DEFAULT_SHIFTS = {
    "ADMIN":          ("07:30","15:00",15),
    "SUPPORT":        ("07:30","15:00",15),
    "TEACHING":       ("07:00","14:30",10),
    "CONDUCTOR":      ("06:30","14:30",15),
    "DRIVER":         ("06:30","14:30",15),
    "CLEANING STAFF": ("06:00","14:00",15),
}

def get_shift_times():
    """Return dict of dept -> {start, end, grace_mins}. Inserts defaults if empty."""
    conn = get_db()
    rows = conn.execute("SELECT dept, shift_start, shift_end, grace_mins FROM shift_times").fetchall()
    conn.close()
    result = {}
    for r in rows:
        result[r["dept"].upper()] = {"start": r["shift_start"], "end": r["shift_end"], "grace": r["grace_mins"]}
    # Fill missing depts with defaults
    for dept, (st, en, gr) in DEFAULT_SHIFTS.items():
        if dept not in result:
            result[dept] = {"start": st, "end": en, "grace": gr}
    return result

def save_shift_times(shifts):
    """shifts: list of {dept, start, end, grace_mins}"""
    conn = get_db()
    for s in shifts:
        conn.execute(
            "INSERT OR REPLACE INTO shift_times (dept, shift_start, shift_end, grace_mins) VALUES (?,?,?,?)",
            (s["dept"].upper(), s["start"], s["end"], int(s.get("grace", 15)))
        )
    conn.commit(); conn.close()

# -- Employee active toggle ----------------------------------------------------
def set_employee_active(badge, active):
    conn = get_db()
    conn.execute("UPDATE employees SET active=?, updated_at=? WHERE badge=?",
                 (1 if active else 0, datetime.now().isoformat(), badge))
    conn.commit(); conn.close()

# -- Workday configuration -------------------------------------------------------
def get_dept_workdays():
    """Return {dept_upper: [weekday_ints]} for all configured departments."""
    conn = get_db()
    rows = conn.execute("SELECT dept, workdays FROM dept_workdays").fetchall()
    conn.close()
    result = {}
    for r in rows:
        try:
            import json as _j
            result[r["dept"].upper()] = _j.loads(r["workdays"])
        except Exception:
            pass
    return result

def save_dept_workday(dept, workdays_list):
    import json as _j
    conn = get_db()
    conn.execute(
        "INSERT OR REPLACE INTO dept_workdays (dept, workdays) VALUES (?,?)",
        (dept.upper(), _j.dumps(workdays_list))
    )
    conn.commit(); conn.close()

def get_emp_workday(badge):
    """Return custom workday list for a specific employee, or None if using dept default."""
    conn = get_db()
    row = conn.execute("SELECT workdays FROM emp_workdays WHERE badge=?", (badge,)).fetchone()
    conn.close()
    if row:
        try:
            import json as _j
            return _j.loads(row["workdays"])
        except Exception:
            pass
    return None

def save_emp_workday(badge, workdays_list):
    import json as _j
    conn = get_db()
    if workdays_list is None:
        conn.execute("DELETE FROM emp_workdays WHERE badge=?", (badge,))
    else:
        conn.execute(
            "INSERT OR REPLACE INTO emp_workdays (badge, workdays) VALUES (?,?)",
            (badge, _j.dumps(workdays_list))
        )
    conn.commit(); conn.close()

def get_all_emp_workdays():
    """Return all employee-specific workday overrides {badge: [weekday_ints]}."""
    import json as _j
    conn = get_db()
    rows = conn.execute("SELECT badge, workdays FROM emp_workdays").fetchall()
    conn.close()
    result = {}
    for r in rows:
        try:
            result[r["badge"]] = _j.loads(r["workdays"])
        except Exception:
            pass
    return result

# -- Holidays ------------------------------------------------------------------
def get_holidays(year=None):
    conn = get_db()
    if year:
        rows = conn.execute(
            "SELECT * FROM holidays WHERE date LIKE ? ORDER BY date",
            (str(year)+'%',)
        ).fetchall()
    else:
        rows = conn.execute("SELECT * FROM holidays ORDER BY date").fetchall()
    conn.close()
    return [dict(r) for r in rows]

def add_holiday(date, date_end, label, scope='all', dept='', employees=''):
    conn = get_db()
    conn.execute(
        "INSERT INTO holidays (date, date_end, label, scope, dept, employees, created_at)"
        " VALUES (?,?,?,?,?,?,?)",
        (date, date_end or date, label, scope, dept, employees, datetime.now().isoformat())
    )
    conn.commit()
    hid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.close()
    return hid

def delete_holiday(hid):
    conn = get_db()
    conn.execute("DELETE FROM holidays WHERE id=?", (hid,))
    conn.commit()
    conn.close()

def is_holiday(day_date, badge='', dept=''):
    """Return (True, label) if day_date is a holiday for this employee, else (False, '')."""
    conn = get_db()
    date_str = day_date.strftime("%Y-%m-%d")
    rows = conn.execute(
        "SELECT * FROM holidays WHERE date <= ? AND date_end >= ?",
        (date_str, date_str)
    ).fetchall()
    conn.close()
    for r in rows:
        scope = r["scope"]
        if scope == 'all':
            return True, r["label"]
        if scope == 'dept' and r["dept"] and dept:
            if r["dept"].upper() == dept.upper():
                return True, r["label"]
        if scope == 'employee' and r["employees"] and badge:
            codes = [c.strip() for c in r["employees"].split(',')]
            if badge in codes:
                return True, r["label"]
    return False, ''

# ── Compatibility shim: db_manager.xxx calls work without separate module ──────
class _DBMgrShim:
    DB_PATH                        = DB_PATH
    get_db                         = staticmethod(get_db)
    init_db                        = staticmethod(init_db)
    upsert_employees               = staticmethod(upsert_employees)
    get_employees                  = staticmethod(get_employees)
    get_employee                   = staticmethod(get_employee)
    store_punches                  = staticmethod(store_punches)
    get_punch_count                = staticmethod(get_punch_count)
    get_punches_for_day            = staticmethod(get_punches_for_day)
    get_punch_records_for_employee = staticmethod(get_punch_records_for_employee)
    record_unknown_user            = staticmethod(record_unknown_user)
    get_unknown_users              = staticmethod(get_unknown_users)
    resolve_unknown_user           = staticmethod(resolve_unknown_user)
    get_setting                    = staticmethod(get_setting)
    set_setting                    = staticmethod(set_setting)
    get_db_stats                   = staticmethod(get_db_stats)
    import_employees_from_csv      = staticmethod(import_employees_from_csv)
    import_punches_from_mdb        = staticmethod(import_punches_from_mdb)
    write_audit                    = staticmethod(write_audit)
    get_audit_log                  = staticmethod(get_audit_log)
    get_shift_times                = staticmethod(get_shift_times)
    save_shift_times               = staticmethod(save_shift_times)
    set_employee_active            = staticmethod(set_employee_active)
    get_dept_workdays              = staticmethod(get_dept_workdays)
    save_dept_workday              = staticmethod(save_dept_workday)
    get_emp_workday                = staticmethod(get_emp_workday)
    save_emp_workday               = staticmethod(save_emp_workday)
    get_all_emp_workdays           = staticmethod(get_all_emp_workdays)
    get_holidays                   = staticmethod(get_holidays)
    add_holiday                    = staticmethod(add_holiday)
    delete_holiday                 = staticmethod(delete_holiday)
    is_holiday                     = staticmethod(is_holiday)

db_manager = _DBMgrShim()


# ── SQLite helper shims (wrap db_manager) ──────────────────────────────────
def _sqlite_ready():
    """True if attendance.db exists and has employees loaded."""
    try:
        if not os.path.exists(db_manager.DB_PATH): return False
        conn = db_manager.get_db()
        n = conn.execute("SELECT COUNT(*) FROM employees").fetchone()[0]
        conn.close()
        return n > 0
    except Exception:
        return False

def _sqlite():
    """Return a db_manager connection."""
    return db_manager.get_db()
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ==============================================================================
#  CONFIG
# ==============================================================================
# Device config loaded from settings.ini after _CFG is initialised below
# (temporary defaults used only if settings.ini is missing)
_DEFAULT_IPS = ["10.20.141.21","10.20.141.22","10.20.141.23","10.20.141.24",
                "10.20.141.25","10.20.141.26","10.20.141.27","10.20.141.28","10.20.141.29"]

def _get_script_dir():
    """UNC-safe directory resolution. Works on network shares via pushd."""
    for fn in (os.path.realpath, os.path.abspath):
        try:
            p = os.path.dirname(fn(__file__))
            if p and os.path.isdir(p): return p
        except Exception: pass
    # Fallback: use the actual script path from command-line argv
    try:
        p = os.path.dirname(os.path.realpath(sys.argv[0]))
        if p and os.path.isdir(p): return p
    except Exception: pass
    return os.getcwd()

SCRIPT_DIR     = _get_script_dir()
# Ensure SCRIPT_DIR is in Python path so db_manager and other local modules import correctly
if SCRIPT_DIR not in sys.path:
    sys.path.insert(0, SCRIPT_DIR)

EMPLOYEES_FILE = os.path.join(SCRIPT_DIR, "employees_export.csv")
USERS_FILE     = os.path.join(SCRIPT_DIR, "dashboard_users.json")
DEVICES_FILE   = os.path.join(SCRIPT_DIR, "device_names.json")

# ── Load settings.ini (falls back to defaults if file missing) ─────────────────
def _load_config():
    cfg = configparser.ConfigParser()
    ini = os.path.join(SCRIPT_DIR, 'settings.ini')
    if os.path.exists(ini):
        cfg.read(ini, encoding='utf-8')
        print("[Config] Loaded settings.ini"); sys.stdout.flush()
    else:
        print("[Config] settings.ini not found — using built-in defaults"); sys.stdout.flush()
    return cfg

_CFG = _load_config()

def _cfg(section, key, default):
    try:    return _CFG.get(section, key).strip()
    except: return default

def _cfg_list(section, key, default):
    raw = _cfg(section, key, '')
    if not raw: return default
    return [x.strip() for x in raw.split(',') if x.strip()]

def _cfg_int(section, key, default):
    try:    return int(_cfg(section, key, str(default)))
    except: return default

MDB_PATH    = _cfg('database', 'mdb_path', '') or None  # from settings.ini

# ── Config-driven constants (all read from settings.ini) ───────────────────────
DEVICE_IPS          = _cfg_list('devices', 'ips', _DEFAULT_IPS)
DEVICE_PORT         = _cfg_int ('devices', 'port',         4370)
DEVICE_TIMEOUT      = _cfg_int ('devices', 'timeout',        15)
DEVICE_PULL_TIMEOUT = _cfg_int ('devices', 'pull_timeout',  120)
CACHE_REFRESH_MINS  = _cfg_int ('server',  'cache_refresh_mins', 5)
APP_VERSION        = _cfg     ('app',     'version', '2.2')

DEFAULT_ADMIN_PASSWORD = _cfg('app', 'default_admin_password', 'Gaesous180')

EXCLUDE_DEPARTMENTS   = _cfg_list('departments', 'exclude',
                            ["DELETED EMPLOYEES","TRANSPORT","GAES","GULF ASIAN ENGLISH SCHOOL"])
COLLAPSED_DEPARTMENTS = _cfg_list('departments', 'collapsed', EXCLUDE_DEPARTMENTS[:])
DEPT_ORDER            = _cfg_list('departments', 'order',
                            ["ADMIN","SUPPORT","TEACHING","CONDUCTOR","DRIVER","CLEANING STAFF"])

ADMIN_INACTIVITY_MINS    = _cfg_int('sessions', 'admin_inactivity_mins',    60)
EMPLOYEE_INACTIVITY_MINS = _cfg_int('sessions', 'employee_inactivity_mins', 20)

BACKUP_DIR = os.path.join(SCRIPT_DIR, 'backups')

# Departments excluded from attendance tracking (absent/present reports)
# EXCLUDE_DEPARTMENTS defined above

CHECKINOUT_TABLE = "CHECKINOUT"
USERINFO_TABLE   = "USERINFO"
COL_USER_ID      = "USERID"
COL_CHECKTIME    = "CHECKTIME"
COL_BADGE_MDB    = "Badgenumber"

DEPT_WORKDAYS_FALLBACK = {
    "ADMIN":          {6, 0, 1, 2, 3},
    "SUPPORT":        {6, 0, 1, 2, 3},
    "DRIVER":         {6, 0, 1, 2, 3},
    "CONDUCTOR":      {6, 0, 1, 2, 3},
    "CLEANING STAFF": {6, 0, 1, 2, 3},
    "TEACHING":       {0, 1, 2, 3},
}
DEFAULT_WORKDAYS = {6, 0, 1, 2, 3}

# ── Telegram notifier (loaded lazily after DB is ready) ────────────────────────
_tg_notifier   = None   # type: ignore
_tg_init_error = None   # last error from _init_telegram(), exposed in test/save responses

def _init_telegram():
    """Instantiate TelegramNotifier from settings.ini / DB overrides."""
    global _tg_notifier, _tg_init_error
    try:
        from telegram_notifier import TelegramNotifier
        token   = _cfg('telegram', 'bot_token',   '') or db_manager.get_setting('tg_bot_token',   '')
        chat_id = _cfg('telegram', 'chat_id',     '') or db_manager.get_setting('tg_chat_id',     '')
        if not token or not chat_id:
            missing = ("bot_token, " if not token else "") + ("chat_id" if not chat_id else "")
            print("[Telegram] No token/chat_id configured — notifications disabled"); sys.stdout.flush()
            _tg_notifier   = None
            _tg_init_error = "Missing: {0}".format(missing.strip(", "))
            return
        # Prefer DB settings (written by UI) over settings.ini defaults
        def _tg_bool(db_key, ini_key, default='1'):
            db_val = db_manager.get_setting(db_key, None)
            if db_val is not None:
                return db_val == '1'
            return _cfg('telegram', ini_key, default) == '1'
        _tg_notifier = TelegramNotifier(
            bot_token=token,
            chat_id=chat_id,
            enabled=_tg_bool('tg_enabled', 'enabled'),
            notify_device_status=_tg_bool('tg_notify_device_status', 'notify_device_status'),
            notify_punches=_tg_bool('tg_notify_punches', 'notify_punches'),
            notify_daily_report=_tg_bool('tg_notify_daily_report', 'notify_daily_report'),
            system_name="Attendance",
        )
        _tg_init_error = None
        print("[Telegram] Notifier ready (chat {0})".format(chat_id)); sys.stdout.flush()
    except Exception as exc:
        _tg_notifier   = None
        # Classify error for a safe, user-friendly summary (no raw stack trace in responses).
        exc_type = type(exc).__name__
        if "ImportError" in exc_type or "ModuleNotFoundError" in exc_type:
            _tg_init_error = "Required package missing (httpx). Run: pip install httpx"
        elif "ConnectionError" in exc_type or "Timeout" in exc_type:
            _tg_init_error = "Network error connecting to Telegram API"
        else:
            _tg_init_error = "Initialization failed ({0})".format(exc_type)
        print("[Telegram] Init failed ({0}): {1}".format(exc_type, exc)); sys.stdout.flush()

# ==============================================================================
#  FLASK APP
# ==============================================================================
app = Flask(__name__, static_folder=os.path.join(SCRIPT_DIR, 'static'))

# ==============================================================================
#  FLASK-SOCKETIO  (VoIP signaling)
# ==============================================================================
try:
    from flask_socketio import SocketIO, emit, join_room, leave_room
    socketio = SocketIO(app, cors_allowed_origins="*", async_mode='threading',
                        logger=False, engineio_logger=False)
    SOCKETIO_AVAILABLE = True
    print("[VoIP] Flask-SocketIO loaded (async_mode=threading)")
except ImportError:
    SOCKETIO_AVAILABLE = False
    socketio = None
    print("[VoIP] flask-socketio not installed — VoIP disabled")

@app.route('/sw.js')
def serve_sw():
    return send_file(os.path.join(SCRIPT_DIR, 'sw.js'), mimetype='application/javascript')

@app.route('/manifest.json')
def serve_manifest():
    return send_file(os.path.join(SCRIPT_DIR, 'manifest.json'), mimetype='application/json')

@app.route('/api/ping')
def api_ping():
    return jsonify({"status": "ok"})

@app.before_request
def _before_req():
    """Touch session last_active. Re-registers session if server was restarted."""
    if 'username' in session:
        sid = session.get('sid')
        if sid:
            with _active_sessions_lock:
                known = sid in _active_sessions
            if not known:
                # Server restarted and lost in-memory sessions — re-register so
                # existing valid cookie sessions are not immediately logged out
                with _active_sessions_lock:
                    _active_sessions[sid] = {
                        'sid':         sid,
                        'username':    session.get('username',''),
                        'role':        session.get('role','employee'),
                        'ip':          request.remote_addr,
                        'login_time':  datetime.now().isoformat(),
                        'last_active': datetime.now().isoformat(),
                    }
        _touch_session()

@app.errorhandler(401)
def unauthorized(e):
    return jsonify({"error": "Not authenticated", "auth_required": True}), 401

@app.errorhandler(403)
def forbidden(e):
    return jsonify({"error": "Access denied"}), 403

@app.errorhandler(404)
def not_found(e):
    return jsonify({"error": "Not found"}), 404

@app.errorhandler(500)
def server_error(e):
    return jsonify({"error": str(e)}), 500
# Secret key: stored in attendance.db settings table so sessions survive restarts.
# No .secret_key file needed. Auto-generates on first run.
_SECRET_PATH = os.path.join(SCRIPT_DIR, ".flask_secret")
def _get_or_create_secret_key():
    try:
        if os.path.exists(_SECRET_PATH):
            with open(_SECRET_PATH, "r") as f:
                key = f.read().strip()
                if key:  # ensure file isn't empty
                    return key
        k = binascii.hexlify(os.urandom(32)).decode()
        with open(_SECRET_PATH, "w") as f: f.write(k)
        print("[SESSION] New secret key created at:", _SECRET_PATH)
        return k
    except Exception as e:
        print("[CRITICAL] Secret Key Error:", e)
        return "static_fallback_emergency_key_zkteco_erp_2024"
# ── Session / Cookie config ───────────────────────────────────────────────────
# SESSION_COOKIE_SECURE=False: Flask runs on HTTP internally (behind nginx/caddy HTTPS proxy).
# The browser communicates over HTTPS to the proxy, but Flask sees HTTP.
# Setting Secure=True would make Flask mark the cookie as HTTPS-only, but since
# Flask is on HTTP, it never sets it — sessions die on every reload.
app.config['SESSION_COOKIE_SECURE']   = False   # Flask is HTTP behind proxy
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_NAME']     = 'zk_session'
app.config['SESSION_COOKIE_PATH']     = '/'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=30)  # 30 days
app.secret_key = _get_or_create_secret_key()

# ProxyFix: tells Flask the real scheme/host when behind nginx/caddy
try:
    from werkzeug.middleware.proxy_fix import ProxyFix
    app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1, x_prefix=1)
except ImportError:
    pass

@app.after_request
def add_cache_control(response):
    if request.path.startswith('/api/'):
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    return response

# ── Active sessions registry ──────────────────────────────────────────────────
# session_id -> {username, role, ip, login_time, last_active}
_active_sessions      = {}
_active_sessions_lock = threading.Lock()

def _session_id():
    """Get or create a unique ID for this Flask session."""
    if 'sid' not in session:
        session['sid'] = binascii.hexlify(os.urandom(12)).decode()
    return session['sid']

def _register_session(username, role, ip):
    sid = _session_id()
    with _active_sessions_lock:
        _active_sessions[sid] = {
            'sid':         sid,
            'username':    username,
            'role':        role,
            'ip':          ip,
            'login_time':  datetime.now().isoformat(),
            'last_active': datetime.now().isoformat(),
        }

def _touch_session():
    """Update last_active timestamp for current session."""
    sid = session.get('sid')
    if sid:
        with _active_sessions_lock:
            if sid in _active_sessions:
                _active_sessions[sid]['last_active'] = datetime.now().isoformat()

def _remove_session(sid=None):
    sid = sid or session.get('sid')
    if sid:
        with _active_sessions_lock:
            _active_sessions.pop(sid, None)

def _kill_user_sessions(username):
    """Force-logout all sessions for a user (for remote logout)."""
    with _active_sessions_lock:
        to_kill = [sid for sid, s in _active_sessions.items() if s['username'] == username]
        for sid in to_kill:
            del _active_sessions[sid]
    return len(to_kill)

def _check_inactivity():
    """Background: purge sessions inactive beyond their role timeout."""
    while True:
        time.sleep(60)
        now = datetime.now()
        with _active_sessions_lock:
            to_kill = []
            for sid, s in _active_sessions.items():
                try:
                    last = datetime.fromisoformat(s['last_active'])
                    role = s.get('role','employee')
                    limit = ADMIN_INACTIVITY_MINS if role == 'admin' else EMPLOYEE_INACTIVITY_MINS
                    if limit > 0 and (now - last).total_seconds() > limit * 60:
                        to_kill.append(sid)
                except Exception:
                    pass
            for sid in to_kill:
                print("[Session] Inactivity timeout: {0}".format(_active_sessions[sid].get('username','?')))
                del _active_sessions[sid]

# Start inactivity checker thread
_inactivity_thread = threading.Thread(target=_check_inactivity, daemon=True)
_inactivity_thread.start()

# ==============================================================================
#  USER MANAGEMENT
# ==============================================================================
# Permission keys for employee logins
PERM_KEYS = [
    "view_overview",
    "view_today",
    "view_history",
    "view_employees",
    "view_devices",
    "view_own_calendar",    # can view their own employee calendar
    "view_all_calendars",   # can view any employee's calendar
    "export_reports",
    "force_refresh",
]

def _hash_pw(pw, salt=None):
    """PBKDF2-SHA256 with random salt. Format: pbkdf2:<salt>:<hex-hash>"""
    salt = salt or binascii.hexlify(os.urandom(16)).decode()
    h = hashlib.pbkdf2_hmac('sha256', pw.encode('utf-8'), salt.encode('utf-8'), 260000)
    return "pbkdf2:{0}:{1}".format(salt, binascii.hexlify(h).decode())

def _verify_pw(pw, stored):
    """Verify password. Handles PBKDF2 (new) and plain SHA-256 (legacy auto-migrates)."""
    if not stored or not pw:
        return False
    if stored.startswith("pbkdf2:"):
        try:
            _, salt, _ = stored.split(":", 2)
            return _hash_pw(pw, salt) == stored
        except Exception:
            return False
    return hashlib.sha256(pw.encode()).hexdigest() == stored

def _load_users():
    if os.path.exists(USERS_FILE):
        try:
            with open(USERS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    # Bootstrap: create default admin
    users = {
        "admin": {
            "password_hash": _hash_pw(DEFAULT_ADMIN_PASSWORD),
            "role": "admin",
            "name": "Administrator",
            "badge": None,
            "permissions": {k: True for k in PERM_KEYS},
            "theme": "dark",
            "must_change_password": False,
        }
    }
    _save_users(users)
    return users

def _save_users(users):
    try:
        with open(USERS_FILE, "w", encoding="utf-8") as f:
            json.dump(users, f, indent=2)
    except Exception as e:
        print("[Users] Save error: {0}".format(e))

_users_lock = threading.Lock()
_users = _load_users()

def _get_users():
    global _users
    return _users

def _reload_users():
    global _users
    _users = _load_users()

# ==============================================================================
#  DEVICE NAMES
# ==============================================================================
def _load_device_names():
    if os.path.exists(DEVICES_FILE):
        try:
            with open(DEVICES_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def _save_device_names(names):
    try:
        with open(DEVICES_FILE, "w", encoding="utf-8") as f:
            json.dump(names, f, indent=2)
    except Exception as e:
        print("[Devices] Save error: {0}".format(e))

_device_names = _load_device_names()

# ==============================================================================
#  AUTH DECORATORS
# ==============================================================================
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "username" not in session:
            return jsonify({"error": "Not authenticated", "auth_required": True}), 401
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "username" not in session:
            return jsonify({"error": "Not authenticated", "auth_required": True}), 401
        users = _get_users()
        u = users.get(session["username"], {})
        if u.get("role") != "admin":
            return jsonify({"error": "Admin access required"}), 403
        return f(*args, **kwargs)
    return decorated

def permission_required(module, action="read"):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if "username" not in session:
                return jsonify({"error": "Not authenticated", "auth_required": True}), 401
            users = _get_users()
            u = users.get(session["username"], {})
            if u.get("role") == "admin":
                return f(*args, **kwargs)
            
            perms = u.get("permissions", {})
            mod_perms = perms.get(module)
            if isinstance(mod_perms, bool):
                has_perm = mod_perms
            elif isinstance(mod_perms, list):
                has_perm = action in mod_perms or "all" in mod_perms
            else:
                has_perm = False

            if not has_perm:
                return jsonify({"error": f"Permission denied: {module}:{action}"}), 403
            return f(*args, **kwargs)
        return decorated
    return decorator

# ==============================================================================
# TICKETING Endpoints
# ==============================================================================

@app.route("/api/tickets", methods=["GET"])
@login_required
def api_get_tickets():
    conn = db_manager.get_db()
    rows = conn.execute("SELECT * FROM tickets ORDER BY id DESC").fetchall()
    conn.close()
    return jsonify({"tickets": [dict(r) for r in rows]})

@app.route("/api/tickets", methods=["POST"])
@login_required
def api_create_ticket():
    data = request.get_json() or {}
    title = data.get("title", "No Title")
    desc = data.get("desc", "")
    priority = data.get("priority", "Medium")
    requester = session.get("username", "unknown")
    conn = db_manager.get_db()
    conn.execute(
        "INSERT INTO tickets (title, desc, priority, status, requester_badge, created_at) VALUES (?,?,?,?,?,?)",
        (title, desc, priority, "Open", requester, datetime.now().isoformat())
    )
    conn.commit(); conn.close()
    try: db_manager.write_audit(requester, "CREATE_TICKET", f"Created ticket {title}", request.remote_addr)
    except Exception: pass
    return jsonify({"ok": True})

@app.route("/api/tickets/<int:tid>", methods=["PUT", "PATCH"])
@permission_required("tickets", "write")
def api_update_ticket(tid):
    data = request.get_json() or {}
    conn = db_manager.get_db()
    fields = []
    params = []
    for k in ["title", "desc", "status", "assigned_to", "priority", "due_time"]:
        if k in data:
            fields.append(f"{k} = ?")
            params.append(data[k])
            
    if not fields:
        return jsonify({"error": "No updates provided"}), 400
        
    params.append(tid)
    conn.execute(f"UPDATE tickets SET {', '.join(fields)} WHERE id=?", tuple(params))
    conn.commit(); conn.close()
    
    username = session.get("username", "unknown")
    import json as _j
    try: db_manager.write_audit(username, "UPDATE_TICKET", f"Updated ticket {tid}", request.remote_addr, _j.dumps(data))
    except Exception: pass
    
    return jsonify({"ok": True})

# ==============================================================================
#  AUTH ROUTES
# ==============================================================================
_login_fails = {}  # ip -> (fail_count, first_fail_time)

@app.route("/api/auth/login", methods=["POST"])
def auth_login():
    data = request.get_json() or {}
    username = (data.get("username") or "").strip().lower()
    password = data.get("password") or ""
    if not username or not password:
        return jsonify({"error": "Username and password required"}), 400
    ip = request.remote_addr
    cnt, ft = _login_fails.get(ip, (0, datetime.now()))
    if cnt >= 10 and (datetime.now() - ft).total_seconds() < 300:
        return jsonify({"error": "Too many failed attempts. Try again in 5 minutes."}), 429
    if (datetime.now() - ft).total_seconds() >= 300:
        _login_fails.pop(ip, None)
    users = _get_users()
    user = users.get(username)
    if not user:
        conn = get_db()
        try:
            emp = conn.execute(
                "SELECT badge, name, dept, active FROM employees WHERE badge=?",
                (username,)
            ).fetchone()
            if emp and emp["active"]:
                if password == username:
                    user = {"name": emp["name"], "role": "employee",
                            "password_hash": "dynamic", "badge": emp["badge"]}
        finally:
            conn.close()
    
    if not user:
        return jsonify({"error": "Invalid credentials"}), 401
    if user.get("password_hash") != "dynamic" and not _verify_pw(password, user["password_hash"]):
        ip = request.remote_addr
        cnt, ft = _login_fails.get(ip, (0, datetime.now()))
        _login_fails[ip] = (cnt + 1, ft)
        return jsonify({"error": "Invalid credentials"}), 401
    # Auto-upgrade legacy SHA-256 and store last_login
    if user.get("password_hash") != "dynamic":
        with _users_lock:
            _users_local = _get_users()
            if username in _users_local:
                if not _users_local[username]["password_hash"].startswith("pbkdf2:"):
                    _users_local[username]["password_hash"] = _hash_pw(password)
                _users_local[username]["last_login"] = datetime.now().isoformat()
                _save_users(_users_local)
    _login_fails.pop(request.remote_addr, None)
    session["username"]   = username
    session["role"]       = user.get("role", "employee")
    session.permanent     = True
    _register_session(username, user.get('role','employee'), request.remote_addr)
    try: db_manager.write_audit(username, "LOGIN", "Logged in", request.remote_addr)
    except Exception: pass
    _role  = user.get("role", "employee")
    _badge = user.get("badge", username)
    resp_data = {
        "ok": True,
        "username":    username,
        "name":        user.get("name", username),
        "role":        _role,
        "theme":       user.get("theme", "dark"),
        "permissions": user.get("permissions", {}),
        "badge":       _badge,
        "must_change_password": user.get("must_change_password", False),
        "avatar_id":   user.get("avatar_id", 1) if _role == "admin" else int(db_manager.get_setting(f"avatar_{_badge}", 1) or 1),
    }
    return jsonify(resp_data)

@app.route("/api/auth/logout", methods=["POST"])
def auth_logout():
    try: db_manager.write_audit(session.get("username","?"), "LOGOUT", "", request.remote_addr)
    except Exception: pass
    _remove_session()
    session.clear()
    return jsonify({"ok": True})

@app.route("/api/auth/me")
def auth_me():
    if "username" not in session:
        return jsonify({"authenticated": False})
    users = _get_users()
    u = users.get(session["username"], {})
    return jsonify({
        "authenticated": True,
        "username":    session["username"],
        "name":        u.get("name", session["username"]),
        "role":        u.get("role", "employee"),
        "theme":       u.get("theme", "dark"),
        "permissions": u.get("permissions", {}),
        "badge":       u.get("badge"),
        "must_change_password": u.get("must_change_password", False),
        "avatar_id":   u.get("avatar_id", 1) if session.get("role")=="admin" else int(db_manager.get_setting(f"avatar_{session['username']}", 1) or 1),
    })

@app.route("/api/auth/avatar", methods=["POST"])
def auth_set_avatar():
    if "username" not in session:
        return jsonify({"error": "Not authenticated"}), 401
    data = request.get_json() or {}
    avatar_id = int(data.get("avatar_id", 1))
    username = session["username"]
    role = session.get("role", "employee")
    
    if role == "admin":
        with _users_lock:
            users = _get_users()
            if username in users:
                users[username]["avatar_id"] = avatar_id
                _save_users(users)
    else:
        db_manager.set_setting(f"avatar_{username}", avatar_id)
        
    return jsonify({"ok": True, "avatar_id": avatar_id})

@app.route("/api/auth/change_password", methods=["POST"])
@login_required
def auth_change_password():
    data = request.get_json() or {}
    current = data.get("current_password", "")
    new_pw  = data.get("new_password", "")
    if not new_pw or len(new_pw) < 4:
        return jsonify({"error": "New password must be at least 4 characters"}), 400
    with _users_lock:
        users = _get_users()
        username = session["username"]
        u = users.get(username)
        if not u:
            return jsonify({"error": "User not found"}), 404
        # Admins can skip current-password check when resetting their own
        if u["role"] != "admin" or current:
            if not _verify_pw(current, u["password_hash"]):
                return jsonify({"error": "Current password is incorrect"}), 400
        u["password_hash"] = _hash_pw(new_pw)
        u["must_change_password"] = False
        _save_users(users)
    try: db_manager.write_audit(session.get("username","?"), "CHANGE_PASSWORD", "", request.remote_addr)
    except Exception: pass
    return jsonify({"ok": True})

@app.route("/api/auth/theme", methods=["POST"])
@login_required
def auth_set_theme():
    data = request.get_json() or {}
    theme = data.get("theme", "dark")
    if theme not in ("dark", "light"):
        return jsonify({"error": "Invalid theme"}), 400
    with _users_lock:
        users = _get_users()
        username = session["username"]
        if username in users:
            users[username]["theme"] = theme
            _save_users(users)
    return jsonify({"ok": True, "theme": theme})

# ==============================================================================
#  ADMIN USER MANAGEMENT ROUTES
# ==============================================================================
@app.route("/api/admin/users", methods=["GET"])
@permission_required("users", "read")
def admin_list_users():
    users = _get_users()
    out = []
    for uname, u in users.items():
        out.append({
            "username":    uname,
            "name":        u.get("name", uname),
            "role":        u.get("role", "employee"),
            "badge":       u.get("badge"),
            "permissions": u.get("permissions", {}),
            "theme":       u.get("theme", "dark"),
            "last_login":  u.get("last_login", "Never"),
            "avatar_id":   u.get("avatar_id", 1) if session.get("role")=="admin" else int(db_manager.get_setting("avatar_{0}".format(uname), 1) or 1),
        })
    return jsonify(out)

@app.route("/api/admin/users", methods=["POST"])
@permission_required("users", "write")
def admin_create_user():
    data = request.get_json() or {}
    username = (data.get("username") or "").strip().lower()
    password = data.get("password") or ""
    name     = data.get("name") or username
    badge    = data.get("badge")
    role     = data.get("role", "employee")
    perms    = data.get("permissions", {k: False for k in PERM_KEYS})
    if not username or not password:
        return jsonify({"error": "Username and password required"}), 400
    if len(password) < 4:
        return jsonify({"error": "Password must be at least 4 characters"}), 400
    with _users_lock:
        users = _get_users()
        if username in users:
            return jsonify({"error": "Username already exists"}), 409
        users[username] = {
            "password_hash": _hash_pw(password),
            "role": role,
            "name": name,
            "badge": badge,
            "permissions": {k: bool(perms.get(k, False)) for k in PERM_KEYS},
            "theme": "dark",
            "must_change_password": data.get("must_change_password", True),
            "last_login": "Never",
            "avatar_id": 1,
        }
        _save_users(users)
        import json as _j
        try: write_audit(session.get("username", "admin"), "CREATE_USER", f"Created user {username}", detail_json=_j.dumps(data))
        except Exception: pass
    return jsonify({"ok": True, "username": username})

@app.route("/api/admin/users/<username>", methods=["PUT"])
@admin_required
def admin_update_user(username):
    data = request.get_json() or {}
    with _users_lock:
        users = _get_users()
        if username not in users:
            return jsonify({"error": "User not found"}), 404
        u = users[username]
        if "name" in data:   u["name"] = data["name"]
        if "badge" in data:  u["badge"] = data["badge"]
        if "role" in data and username != "admin":
            u["role"] = data["role"]
        if "permissions" in data:
            u["permissions"] = {k: bool(data["permissions"].get(k, False)) for k in PERM_KEYS}
        if "new_password" in data and data["new_password"]:
            if len(data["new_password"]) < 6:
                return jsonify({"error": "Password must be at least 6 characters"}), 400
            u["password_hash"] = _hash_pw(data["new_password"])
            u["must_change_password"] = data.get("must_change_password", False)
        _save_users(users)
        import json as _j
        try: write_audit(session.get("username", "admin"), "UPDATE_USER", f"Updated user {username}", detail_json=_j.dumps(data))
        except Exception: pass
    return jsonify({"ok": True})

@app.route("/api/admin/users/<username>", methods=["DELETE"])
@admin_required
def admin_delete_user(username):
    if username == "admin":
        return jsonify({"error": "Cannot delete the admin account"}), 400
    with _users_lock:
        users = _get_users()
        if username not in users:
            return jsonify({"error": "User not found"}), 404
        del users[username]
        _save_users(users)
        try: write_audit(session.get("username", "admin"), "DELETE_USER", f"Deleted user {username}")
        except Exception: pass
    return jsonify({"ok": True})

@app.route("/api/admin/perm_keys")
@admin_required
def admin_perm_keys():
    return jsonify(PERM_KEYS)

# ==============================================================================
#  DEVICE NAME ROUTES
# ==============================================================================
# ── Device IP Management ────────────────────────────────────────────
def _load_device_ips():
    """Load IPs from attendance.db settings table; fall back to settings.ini values."""
    try:
        conn = get_db()
        row  = conn.execute("SELECT value FROM settings WHERE key='device_ips'").fetchone()
        conn.close()
        if row:
            data = json.loads(row[0])
            if data: return data
    except Exception: pass
    return list(DEVICE_IPS)

def _save_device_ips_file(ips):
    """Save IPs to attendance.db settings table (no file written)."""
    try:
        conn = get_db()
        conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES ('device_ips', ?)",
                     (json.dumps(ips),))
        conn.commit(); conn.close()
    except Exception as e:
        print("[DeviceIPs] Save error: {0}".format(e))

@app.route("/api/devices/ips", methods=["GET"])
@login_required
def get_device_ips():
    ips   = _load_device_ips()
    names = _device_names
    return jsonify([{"ip": ip, "name": names.get(ip, "")} for ip in ips])

@app.route("/api/devices/ips", methods=["POST"])
@admin_required
def save_device_ips():
    import re
    data = request.get_json() or {}
    pat  = re.compile(r"^\d{1,3}[.]\d{1,3}[.]\d{1,3}[.]\d{1,3}$")
    ips  = [ip.strip() for ip in data.get("ips", []) if pat.match(ip.strip())]
    seen = set()
    unique = [ip for ip in ips if not (ip in seen or seen.add(ip))]
    if not unique:
        return jsonify({"error": "No valid IP addresses provided"}), 400
    _save_device_ips_file(unique)
    global DEVICE_IPS
    DEVICE_IPS = unique
    try:
        db_manager.write_audit(session.get("username","?"), "UPDATE_DEVICE_IPS",
                               "{0} IPs saved".format(len(unique)), request.remote_addr)
    except Exception: pass
    return jsonify({"ok": True, "ips": unique})

@app.route("/api/devices/names", methods=["GET"])
@login_required
def get_device_names():
    return jsonify(_device_names)

@app.route("/api/devices/names", methods=["POST"])
@admin_required
def set_device_names():
    data = request.get_json() or {}
    global _device_names
    # Only accept keys that are known IPs
    new_names = {ip: str(v).strip() for ip, v in data.items() if ip in DEVICE_IPS}
    _device_names = new_names
    _save_device_names(_device_names)
    return jsonify({"ok": True, "names": _device_names})

# ==============================================================================
#  SETTINGS API ROUTES (Restored)
# ==============================================================================
@app.route("/api/settings/system", methods=["GET"])
@login_required
def get_system_settings():
    try:
        g = db_manager.get_setting
        return jsonify({
            "company_name": g("company_name", "GAES"),
            "email_domain":  g("email_domain",  "aman5z.in"),
            "company_logo":  g("company_logo",  ""),
            "gas_url":        g("gas_url",   ""),
            "gas_email":      g("gas_email",  ""),
            "gas_pass":       g("gas_pass",   ""),
            "zk_url":         g("zk_url",    ""),
            "term_ws_url":    g("term_ws_url", ""),
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/settings/system", methods=["POST"])
@admin_required
def set_system_settings():
    data = request.get_json() or {}
    try:
        s = db_manager.set_setting
        field_map = {
            "company_name": "company_name",
            "email_domain":  "email_domain",
            "company_logo":  "company_logo",
            "gas_url":        "gas_url",
            "gas_email":      "gas_email",
            "gas_pass":       "gas_pass",
            "zk_url":         "zk_url",
            "term_ws_url":    "term_ws_url",
        }
        for field, key in field_map.items():
            if field in data and data[field] is not None:
                s(key, str(data[field]))
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ==============================================================================
#  MDB HELPERS
# ==============================================================================
def _find_mdb():
    if MDB_PATH: return MDB_PATH
    for f in os.listdir(SCRIPT_DIR):
        if f.lower().endswith(".mdb") or f.lower().endswith(".accdb"):
            found = os.path.join(SCRIPT_DIR, f)
            print("[MDB] Auto-detected: {0}".format(found)); sys.stdout.flush()
            return found
    raise FileNotFoundError(
        "No .mdb / .accdb file found in: {0}\n"
        "Place the database in the same folder as server.py, or set MDB_PATH.".format(SCRIPT_DIR)
    )

def _is_linux():
    return sys.platform.startswith("linux")

class _MdbToolsConn:
    """Minimal MDB connection using mdbtools on Linux."""
    def __init__(self, mdb_path):
        import subprocess
        self.mdb_path = mdb_path
        r = subprocess.run(["mdb-tables", "-1", mdb_path],
                           capture_output=True, text=True, timeout=10)
        if r.returncode != 0:
            raise RuntimeError(
                "mdbtools error: " + r.stderr.strip() + "\n"
                "Install with: sudo apt install mdbtools"
            )
        self._tables = [t.strip() for t in r.stdout.splitlines() if t.strip()]

    def read_table(self, table_name):
        import subprocess
        import io
        r = subprocess.run(["mdb-export", self.mdb_path, table_name],
                           capture_output=True, text=True, timeout=60)
        if r.returncode != 0:
            raise RuntimeError("mdb-export failed for {0}: {1}".format(
                table_name, r.stderr.strip()))
        return pd.read_csv(io.StringIO(r.stdout), dtype=str)

    def close(self): pass

    def cursor(self): return self

    def tables(self, tableType=None):
        class Row:
            def __init__(self, name): self.table_name = name
        return [Row(t) for t in self._tables]

def _mdb_read(sql, conn):
    """Cross-platform pd.read_sql — uses mdbtools on Linux."""
    if not isinstance(conn, _MdbToolsConn):
        return pd.read_sql(sql, conn)
    # Parse table name from SQL
    import re
    m = re.search(r'FROM\s+\[?(\w+)\]?', sql, re.IGNORECASE)
    if not m:
        raise ValueError("Cannot parse table from SQL: " + sql)
    table = m.group(1)
    df = conn.read_table(table)
    # Apply basic date-range WHERE filter (Access # syntax)
    date_m = re.findall(r'\[?(\w+)\]?\s*(>=|<=|>|<)\s*#([^#]+)#', sql)
    for col, op, val in date_m:
        if col in df.columns:
            try:
                df[col] = pd.to_datetime(df[col], errors='coerce')
                dv = pd.to_datetime(val.strip())
                if   op == '>=': df = df[df[col] >= dv]
                elif op == '<=': df = df[df[col] <= dv]
                elif op == '>':  df = df[df[col] > dv]
                elif op == '<':  df = df[df[col] < dv]
            except Exception:
                pass
    return df

def connect_mdb(mdb_path=None):
    raw  = mdb_path or _find_mdb()
    path = os.path.abspath(raw)
    if not os.path.exists(path):
        raise FileNotFoundError("MDB not found: " + path)
    if _is_linux():
        print("[MDB] Linux — using mdbtools"); sys.stdout.flush()
        return _MdbToolsConn(path)
    import pyodbc
    return pyodbc.connect(
        "Driver={Microsoft Access Driver (*.mdb, *.accdb)};"
        "Dbq=" + path + ";Uid=Admin;Pwd=;"
    )

def get_uid_to_badge_map(conn):
    df = _mdb_read(
        "SELECT [{0}],[{1}] FROM [{2}]".format(COL_USER_ID, COL_BADGE_MDB, USERINFO_TABLE), conn
    )
    df[COL_USER_ID]   = df[COL_USER_ID].astype(str).str.strip()
    df[COL_BADGE_MDB] = df[COL_BADGE_MDB].astype(str).str.strip()
    return dict(zip(df[COL_USER_ID], df[COL_BADGE_MDB]))

def get_badge_to_uid_map(conn):
    df = _mdb_read(
        "SELECT [{0}],[{1}] FROM [{2}]".format(COL_USER_ID, COL_BADGE_MDB, USERINFO_TABLE), conn
    )
    df[COL_USER_ID]   = df[COL_USER_ID].astype(str).str.strip()
    df[COL_BADGE_MDB] = df[COL_BADGE_MDB].astype(str).str.strip()
    return dict(zip(df[COL_BADGE_MDB], df[COL_USER_ID]))

_local_sync_time = 0
_local_sync_lock = threading.Lock()
def _ensure_synced():
    global _local_sync_time
    # Sync every 5 minutes maximum
    if time.time() - _local_sync_time < 300:
        return
    with _local_sync_lock:
        if time.time() - _local_sync_time < 300:
            return
        try:
            db = get_db()
            db.execute("CREATE TABLE IF NOT EXISTS checkinout_cache (userid TEXT, checktime DATETIME, sn TEXT)")
            db.execute("CREATE INDEX IF NOT EXISTS idx_checktime ON checkinout_cache(checktime)")
            db.execute("CREATE INDEX IF NOT EXISTS idx_userid ON checkinout_cache(userid)")
            r = db.execute("SELECT MAX(checktime) as max_t FROM checkinout_cache").fetchone()
            max_t = r["max_t"] if r and r["max_t"] else "2000-01-01 00:00:00"
            
            conn = connect_mdb()
            sql = "SELECT [{0}], [{1}], [sn] FROM [{2}] WHERE [{1}] > #{3}#".format(
                COL_USER_ID, COL_CHECKTIME, CHECKINOUT_TABLE, max_t
            )
            try:
                df = _mdb_read(sql, conn)
            except:
                # Fallback without [sn] if column missing
                sql = "SELECT [{0}], [{1}] FROM [{2}] WHERE [{1}] > #{3}#".format(
                    COL_USER_ID, COL_CHECKTIME, CHECKINOUT_TABLE, max_t
                )
                df = _mdb_read(sql, conn)
            conn.close()
            
            if not df.empty:
                df[COL_CHECKTIME] = pd.to_datetime(df[COL_CHECKTIME], errors="coerce")
                df = df.dropna(subset=[COL_CHECKTIME])
                df["checktime_str"] = df[COL_CHECKTIME].dt.strftime("%Y-%m-%d %H:%M:%S")
                df = df.drop_duplicates(subset=[COL_USER_ID, "checktime_str"])
                
                rows = []
                for _, row in df.iterrows():
                    rows.append((str(row[COL_USER_ID]).strip(), row["checktime_str"], str(row.get("sn", ""))))
                if rows:
                    db.executemany("INSERT INTO checkinout_cache (userid, checktime, sn) VALUES (?, ?, ?)", rows)
                    db.commit()
            db.close()
            _local_sync_time = time.time()
        except Exception as e:
            print("[SYNC ERROR]", e)

def _get_checkinout(conn, date_from, date_to, uid=None):
    """Load punch records from fast local SQLite cache, syncing from MDB in the background."""
    _ensure_synced()
    fs = date_from.strftime("%Y-%m-%d 00:00:00")
    ts = date_to.strftime("%Y-%m-%d 23:59:59")
    db = get_db()
    if uid:
        rows = db.execute("SELECT userid, checktime, sn FROM checkinout_cache WHERE checktime >= ? AND checktime <= ? AND userid = ?", (fs, ts, str(uid))).fetchall()
    else:
        rows = db.execute("SELECT userid, checktime, sn FROM checkinout_cache WHERE checktime >= ? AND checktime <= ?", (fs, ts)).fetchall()
    db.close()
    
    df = pd.DataFrame([dict(r) for r in rows])
    if df.empty:
        return pd.DataFrame(columns=[COL_USER_ID, COL_CHECKTIME, "sn"])
    df = df.rename(columns={"userid": COL_USER_ID, "checktime": COL_CHECKTIME})
    return df

# ==============================================================================
#  SCHEDULE READER
# ==============================================================================
_badge_workdays    = {}
_schedule_loaded   = False
_serial_to_ip      = {}
_shift_cache       = {}   # dept -> {start, end, grace}  loaded from db_manager

def _get_shift(dept):
    global _shift_cache
    if not _shift_cache:
        try: _shift_cache = db_manager.get_shift_times()
        except Exception: pass
    d = dept.upper() if dept else ""
    s = _shift_cache.get(d) or _shift_cache.get("DEFAULT") or {"start":"07:30","end":"15:00","grace":15}
    return s

def _parse_hm(hm_str):
    """'07:30' -> timedelta from midnight"""
    try:
        h, m = hm_str.split(":")
        return timedelta(hours=int(h), minutes=int(m))
    except Exception:
        return timedelta(hours=7, minutes=30)

def _check_late_early(punches, dept):
    """
    Given sorted punch list [{time:'HH:MM:SS',...}] return dict:
      late: bool, early_departure: bool,
      late_mins: int, early_mins: int
    """
    result = {"late": False, "early_departure": False, "late_mins": 0, "early_mins": 0}
    if not punches: return result
    shift  = _get_shift(dept)
    start  = _parse_hm(shift["start"])
    end    = _parse_hm(shift["end"])
    grace  = timedelta(minutes=int(shift.get("grace", 15)))

    def _to_td(t_str):
        try:
            h, m, s = t_str.split(":")
            return timedelta(hours=int(h), minutes=int(m), seconds=int(s))
        except:
            return None

    first_punch = _to_td(punches[0]["time"])
    last_punch  = _to_td(punches[-1]["time"])
    if first_punch is not None and first_punch > start + grace:
        result["late"] = True
        result["late_mins"] = int((first_punch - start).total_seconds() / 60)
    if last_punch is not None and last_punch < end - timedelta(minutes=5):
        result["early_departure"] = True
        result["early_mins"] = int((end - last_punch).total_seconds() / 60)
    return result
_uid_to_badge_cache = {}   # populated from MDB for SQLite punch storage

def _ensure_uid_badge_cache():
    """Ensure _uid_to_badge_cache is populated.
    Priority: (1) already populated from MDB; (2) SQLite employees table
    where uid is assumed to equal badge (standard ZKTeco default behaviour).
    """
    global _uid_to_badge_cache
    if _uid_to_badge_cache:
        return
    try:
        conn = get_db()
        rows = conn.execute("SELECT badge FROM employees WHERE active=1").fetchall()
        conn.close()
        for r in rows:
            b = str(r["badge"]).strip()
            if b:
                _uid_to_badge_cache[b] = b   # uid == badge fallback
        if _uid_to_badge_cache:
            print("[Cache] uid-to-badge: loaded {0} entries from SQLite (uid==badge fallback)".format(
                len(_uid_to_badge_cache))); sys.stdout.flush()
    except Exception as e:
        print("[Cache] uid-to-badge fallback error: {0}".format(e)); sys.stdout.flush()

def _emp_cache_for_db():
    """Return badge set from CSV for unknown-user detection."""
    try:
        df = _read_employee_file(_find_employee_file())
        return set(df["Badgenumber"].tolist())
    except Exception:
        return set()

def _load_schedule_from_mdb():
    DAY_COLS = {"SUN":6,"MON":0,"TUES":1,"WED":2,"THUR":3,"FR":4,"SAT":5}
    try:
        conn    = connect_mdb()
        cursor  = conn.cursor()
        tables  = {r.table_name.upper() for r in cursor.tables(tableType="TABLE")}
        sch_to_days = {}
        if "SCHCLASS" in tables:
            try:
                df_sc  = _mdb_read("SELECT * FROM [SCHCLASS]", conn)
                id_col = next((c for c in df_sc.columns if c.upper() in ("SCHCLASSID","CLASSID","ID")), df_sc.columns[0])
                for _, row in df_sc.iterrows():
                    sid  = str(row[id_col]).strip(); days = set()
                    for col in df_sc.columns:
                        cu = col.upper()
                        for day_prefix, py_day in DAY_COLS.items():
                            if cu.startswith(day_prefix) and "SHIFT" in cu:
                                try:
                                    if int(float(str(row[col]))) != 0: days.add(py_day)
                                except: pass
                    if days: sch_to_days[sid] = days
            except Exception as e:
                print("[Schedule] SCHCLASS error: {0}".format(e))
        user_to_sch = {}
        if "USERUSEDSCLASSES" in tables:
            try:
                df_usc  = _mdb_read("SELECT * FROM [USERUSEDSCLASSES]", conn)
                uid_col = next((c for c in df_usc.columns if "userid" in c.lower()), None)
                cls_col = next((c for c in df_usc.columns if "classid" in c.lower() or "schclass" in c.lower()), None)
                if uid_col and cls_col:
                    for _, row in df_usc.iterrows():
                        uid = str(row[uid_col]).strip(); cid = str(row[cls_col]).strip()
                        if uid not in ("","nan") and cid not in ("","nan","0"):
                            user_to_sch[uid] = cid
            except Exception as e:
                print("[Schedule] USERUSEDSCLASSES error: {0}".format(e))
        dept_to_sch = {}
        if "DEPTUSEDSCHS" in tables:
            try:
                df_ds   = _mdb_read("SELECT * FROM [DEPTUSEDSCHS]", conn)
                did_col = next((c for c in df_ds.columns if "deptid" in c.lower()), None)
                sch_col = next((c for c in df_ds.columns if "schid" in c.lower() or "classid" in c.lower()), None)
                if did_col and sch_col:
                    for _, row in df_ds.iterrows():
                        did = str(row[did_col]).strip(); sid = str(row[sch_col]).strip()
                        if did not in dept_to_sch: dept_to_sch[did] = sid
            except Exception as e:
                print("[Schedule] DEPTUSEDSCHS error: {0}".format(e))
        badge_workdays = {}
        df_u = _mdb_read(
            "SELECT [{0}],[{1}],[DEFAULTDEPTID] FROM [{2}]".format(COL_USER_ID, COL_BADGE_MDB, USERINFO_TABLE), conn
        )
        for _, row in df_u.iterrows():
            uid   = str(row[COL_USER_ID]).strip()
            badge = str(row[COL_BADGE_MDB]).strip()
            dept  = str(row["DEFAULTDEPTID"]).strip()
            cid   = user_to_sch.get(uid) or dept_to_sch.get(dept)
            if cid and cid in sch_to_days:
                badge_workdays[badge] = sch_to_days[cid]
        conn.close()
        return badge_workdays
    except Exception as e:
        print("[Schedule] Error: {0}".format(e)); sys.stdout.flush()
        return {}

_dept_workdays_cache  = {}   # dept -> [weekday ints]
_emp_workdays_cache   = {}   # badge -> [weekday ints]
_workday_cache_loaded = False

def _load_workday_caches():
    global _dept_workdays_cache, _emp_workdays_cache, _workday_cache_loaded
    try:
        _dept_workdays_cache = get_dept_workdays()
        _emp_workdays_cache  = get_all_emp_workdays()
        _workday_cache_loaded = True
    except Exception as e:
        print("[Workdays] Cache load failed: {0}".format(e))

def _get_workdays_for_badge(badge, dept):
    # 1. Employee-specific override
    emp_wd = _emp_workdays_cache.get(badge)
    if emp_wd is not None:
        return set(emp_wd)
    # 2. Department setting from DB
    dept_wd = _dept_workdays_cache.get((dept or "").upper())
    if dept_wd is not None:
        return set(dept_wd)
    # 3. MDB-loaded schedule
    if badge in _badge_workdays:
        return _badge_workdays[badge]
    # 4. Hardcoded fallback
    return DEPT_WORKDAYS_FALLBACK.get((dept or "").upper(), DEFAULT_WORKDAYS)

def _is_working_day(d, badge, dept):
    on_holiday, _label = is_holiday(d, badge=badge, dept=dept)
    if on_holiday:
        return False
    return d.weekday() in _get_workdays_for_badge(badge, dept)

# ==============================================================================
#  EMPLOYEE LOADING
# ==============================================================================
def _read_employee_file(path):
    if path.lower().endswith(".csv"):
        df = pd.read_csv(path, dtype=str)
    else:
        raw  = pd.read_excel(path, header=None, dtype=str)
        hrow = 0
        for i, row in raw.iterrows():
            if any(str(v).strip() == "Badgenumber" for v in row.values):
                hrow = i; break
        df = pd.read_excel(path, header=hrow, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    df["Badgenumber"] = df["Badgenumber"].astype(str).str.strip()
    df["Name"]        = df["Name"].astype(str).str.strip()
    df["DEPTNAME"]    = df["DEPTNAME"].astype(str).str.strip()
    df = df[
        df["Badgenumber"].notna() & (df["Badgenumber"] != "") & (df["Badgenumber"] != "nan") &
        (df["Name"] != "") & (df["Name"] != "nan")
    ].copy()
    return df.reset_index(drop=True)

def _find_employee_file():
    path = EMPLOYEES_FILE
    if not os.path.exists(path):
        alt = path.replace(".csv",".xlsx") if path.endswith(".csv") else path.replace(".xlsx",".csv")
        if os.path.exists(alt): return alt
        raise FileNotFoundError("Employee file not found: {0}\nRun EXPORT_EMPLOYEES.py first.".format(path))
    return path

def load_employees():
    """Active employees only (excludes EXCLUDE_DEPARTMENTS)."""
    df   = _read_employee_file(_find_employee_file())
    excl = [d.upper() for d in EXCLUDE_DEPARTMENTS]
    return df[~df["DEPTNAME"].str.upper().isin(excl)].copy().reset_index(drop=True)

def load_employees_all():
    """All employees including excluded departments."""
    return _read_employee_file(_find_employee_file())

def _dept_sort(dept):
    try:    return DEPT_ORDER.index(dept.upper())
    except: return len(DEPT_ORDER)

# ==============================================================================
#  DEVICE PULL
# ==============================================================================
def _pull_single_device(ip, target_date):
    result_box = [{"ip": ip, "online": False, "error": "Timeout", "badges": set()}]
    def _do_pull():
        zk_conn = None
        try:
            from zk import ZK
            zk      = ZK(ip, port=DEVICE_PORT, timeout=DEVICE_TIMEOUT, verbose=False)
            zk_conn = zk.connect()
            zk_conn.disable_device()
            records    = zk_conn.get_attendance()
            today_bdgs = set(); day_hits = 0; full_records = []
            for r in records:
                uid = str(r.user_id).strip()
                ts  = r.timestamp.strftime("%Y-%m-%d %H:%M:%S")
                full_records.append((uid, ts))
                if r.timestamp.date() == target_date:
                    today_bdgs.add(uid); day_hits += 1
            zk_conn.enable_device()
            result_box[0] = {"ip": ip, "online": True, "punches_today": day_hits,
                             "badges": today_bdgs, "full_records": full_records}
        except Exception as e:
            result_box[0] = {"ip": ip, "online": False, "error": str(e), "badges": set()}
        finally:
            if zk_conn:
                try: zk_conn.enable_device(); zk_conn.disconnect()
                except: pass
    t = threading.Thread(target=_do_pull)
    t.daemon = True
    t.start()
    t.join(timeout=DEVICE_PULL_TIMEOUT)
    return result_box[0]

def _check_device_status(ip):
    global _serial_to_ip
    zk_conn = None
    try:
        from zk import ZK
        zk      = ZK(ip, port=DEVICE_PORT, timeout=10, verbose=False)
        zk_conn = zk.connect()
        sn = zk_conn.get_serialnumber() if hasattr(zk_conn, "get_serialnumber") else "N/A"
        if sn and sn != "N/A": _serial_to_ip[sn] = ip
        return {
            "ip": ip, "online": True,
            "serialno":   sn,
            "platform":   zk_conn.get_platform() if hasattr(zk_conn, "get_platform") else "N/A",
            "user_count": len(zk_conn.get_users()),
            "name":       _device_names.get(ip, ""),
        }
    except Exception as e:
        return {"ip": ip, "online": False, "error": str(e), "name": _device_names.get(ip, "")}
    finally:
        if zk_conn:
            try: zk_conn.disconnect()
            except: pass

# ==============================================================================
#  BACKGROUND CACHE
# ==============================================================================
_cache_lock = threading.Lock()
_cache = {"today": None, "last_updated": None, "refreshing": False, "error": None}

# Track device online state across refreshes so we only alert on *changes*
_device_online_state = {}   # ip -> bool

# Track which (badge, punch_time) pairs we've already notified about
_notified_punches = set()   # (badge, punch_time) — cleared daily

def _refresh_cache():
    global _device_online_state, _notified_punches
    with _cache_lock:
        if _cache["refreshing"]: return
        _cache["refreshing"] = True
    today = date.today()
    _ensure_uid_badge_cache()   # make sure UID→badge map is ready
    try:
        emp_df = load_employees()
    except Exception as e:
        with _cache_lock: _cache["refreshing"] = False; _cache["error"] = str(e)
        return

    # Build badge -> name lookup for punch notifications
    badge_name_map = {}
    try:
        for _, row in emp_df.iterrows():
            badge_name_map[str(row["Badgenumber"]).strip()] = str(row["Name"]).strip()
    except Exception:
        pass

    present_badges = set()
    att_results    = {}
    try:
        for idx, ip in enumerate(DEVICE_IPS, 1):
            print("[Cache] Polling {0}/{1}: {2}".format(idx, len(DEVICE_IPS), ip)); sys.stdout.flush()
            _sse_push({"type":"progress","device":ip,"idx":idx,"total":len(DEVICE_IPS),"status":"polling"})
            result = _pull_single_device(ip, today)
            is_online = result.get("online", False)
            _sse_push({"type":"progress","device":ip,"idx":idx,"total":len(DEVICE_IPS),
                       "status":"online" if is_online else "offline",
                       "punches":result.get("punches_today",0)})
            att_results[ip] = result
            present_badges.update(result.get("badges", set()))

            # ── Telegram: device up/down alert (only on state change) ──
            prev_online = _device_online_state.get(ip)
            if prev_online is not None and prev_online != is_online:
                try:
                    dev_name = _device_names.get(ip, "")
                    if _tg_notifier:
                        if is_online:
                            threading.Thread(
                                target=_tg_notifier.notify_device_online,
                                args=(ip, dev_name), daemon=True).start()
                        else:
                            threading.Thread(
                                target=_tg_notifier.notify_device_offline,
                                args=(ip, dev_name, result.get("error", "")),
                                daemon=True).start()
                except Exception:
                    pass
            _device_online_state[ip] = is_online

            # Store full punch records to SQLite
            full_records = result.get("full_records", [])
            if full_records:
                emp_badges  = _emp_cache_for_db()   # returns a set of badge strings
                punch_store = []
                new_today_punches = []   # (badge, name, ts) for Telegram
                today_str = today.strftime("%Y-%m-%d")
                for (uid, ts) in full_records:
                    # Prefer raw device user_id when it directly matches an employee
                    # badge (the common ZKTeco case: user_id == badge number).  Only
                    # use the MDB UID→badge map when the raw UID is NOT itself a badge,
                    # preventing cross-device sequential-UID collisions from causing
                    # punches to be attributed to the wrong employee.
                    if uid in emp_badges:
                        badge = uid
                    else:
                        badge = _uid_to_badge_cache.get(uid, "")
                    effective_badge = badge or uid
                    punch_store.append((effective_badge, ts, ip))
                    # Record unknown users (device UIDs not mapped to any employee badge)
                    if not effective_badge or effective_badge not in emp_badges:
                        try: db_manager.record_unknown_user(ip, uid)
                        except Exception: pass
                    # Collect new today-punches for Telegram notification
                    if ts.startswith(today_str) and (effective_badge, ts) not in _notified_punches:
                        new_today_punches.append((effective_badge, ts))
                        _notified_punches.add((effective_badge, ts))
                try: db_manager.store_punches(punch_store)
                except Exception as e:
                    print("[DB] store_punches error: {0}".format(e)); sys.stdout.flush()

                # ── Telegram: per-punch notifications (fire-and-forget thread) ──
                if new_today_punches and _tg_notifier:
                    def _send_punch_notifs(punches, device_ip):
                        for bdg, ts in punches:
                            try:
                                name = badge_name_map.get(bdg, "")
                                _tg_notifier.notify_punch(bdg, name, device_ip, ts)
                            except Exception:
                                pass
                    threading.Thread(
                        target=_send_punch_notifs,
                        args=(list(new_today_punches), ip),
                        daemon=True).start()

            time.sleep(0.3)
    except Exception as e:
        print("[Cache] Device loop error: {0}".format(e)); sys.stdout.flush()

    # Clear notified-punches set at day rollover so it doesn't grow forever
    today_prefix = today.strftime("%Y-%m-%d")
    _notified_punches = {(b, t) for (b, t) in _notified_punches if t.startswith(today_prefix)}
    device_status = []
    try:
        with ThreadPoolExecutor(max_workers=len(DEVICE_IPS)) as ex:
            ping_futures = {ex.submit(_check_device_status, ip): ip for ip in DEVICE_IPS}
            ping_results = {}
            for f in as_completed(ping_futures, timeout=12):
                r = f.result(); ping_results[r["ip"]] = r
    except Exception:
        ping_results = {}
    for ip in DEVICE_IPS:
        att  = att_results.get(ip, {"ip": ip, "online": False, "error": "No result", "badges": set()})
        ping = ping_results.get(ip, {})
        is_online = att.get("online", False) or ping.get("online", False)
        device_status.append({
            "ip":            ip,
            "name":          _device_names.get(ip, ""),
            "online":        is_online,
            "punches_today": att.get("punches_today", 0),
            "error":         att.get("error", ping.get("error", "")),
            "serialno":      ping.get("serialno", "N/A"),
            "platform":      ping.get("platform", "N/A"),
            "user_count":    ping.get("user_count", 0),
        })
    # Translate raw device UIDs → employee badge numbers.
    # _uid_to_badge_cache maps uid_str -> badge_str.
    # Also keep raw UIDs as fallback (some devices store badge number as UID).
    _emp_badges_set = _emp_cache_for_db()
    present_badges_resolved = set()
    for uid in present_badges:
        # ZKTeco devices normally enroll users with their badge/employee number as
        # user_id.  Prefer the raw device user_id when it IS a known employee badge
        # to avoid cross-device UID collisions: in multi-device setups different
        # devices can assign the same internal sequential UID to different employees,
        # so an MDB-derived UID→badge mapping from one device can be wrong for
        # another device.  Only fall back to the MDB map when the raw UID is not
        # itself a recognisable employee badge.
        if uid in _emp_badges_set:
            present_badges_resolved.add(uid)
        else:
            mapped = _uid_to_badge_cache.get(uid, "")
            present_badges_resolved.add(mapped if mapped else uid)
    # Supplement with badges already stored in the SQLite punches table for today.
    # This covers employees whose device UIDs did not match their badge at pull time
    # but whose data was correctly stored in a previous sync cycle.
    # Only add badges that belong to known employees to prevent phantom/wrong UIDs
    # (e.g., from UID collisions across devices) from showing as present.
    try:
        today_str = today.strftime("%Y-%m-%d")
        db_sup = get_db()
        sup_rows = db_sup.execute(
            "SELECT DISTINCT badge FROM punches WHERE punch_time >= ? AND punch_time <= ?",
            (today_str + " 00:00:00", today_str + " 23:59:59")
        ).fetchall()
        db_sup.close()
        for r in sup_rows:
            b = str(r["badge"]).strip()
            # Only add if it is a recognised employee badge (or emp set is empty which
            # means we could not load the file — in that case add unconditionally so we
            # don't accidentally hide everyone).
            if b and (not _emp_badges_set or b in _emp_badges_set):
                present_badges_resolved.add(b)
    except Exception as e:
        print("[Cache] DB punch supplement error: {0}".format(e)); sys.stdout.flush()

    absent, present, off_today = [], [], []
    for _, emp in emp_df.iterrows():
        badge = emp["Badgenumber"]; dept = emp["DEPTNAME"]
        rec = {"name": emp["Name"], "code": badge, "dept": dept}
        if not _is_working_day(today, badge, dept):
            off_today.append(rec)
        elif badge in present_badges_resolved:
            present.append(rec)
        else:
            absent.append(rec)
    # Also count anyone who punched in but is marked as off-day (holiday duty)
    absent.sort(key=lambda x: (_dept_sort(x["dept"]), x["name"]))
    present.sort(key=lambda x: (_dept_sort(x["dept"]), x["name"]))
    with _cache_lock:
        _cache["today"] = {
            "date": today.strftime("%d %B %Y"), "total": len(present)+len(absent)+len(off_today),
            "working_today": len(present)+len(absent),
            "present_count": len(present), "absent_count": len(absent),
            "off_today_count": len(off_today),
            "present": present, "absent": absent, "off_today": off_today,
            "devices": device_status,
            "last_updated": datetime.now().strftime("%H:%M:%S"), "cache_age_secs": 0,
        }
        _cache["last_updated"] = datetime.now()
        _cache["refreshing"]   = False
        _cache["error"]        = None
    _sse_push({"type":"done","present":len(present),"absent":len(absent)})
    print("[Cache] Done. Present={0}, Absent={1}".format(len(present), len(absent))); sys.stdout.flush()

def _background_loop():
    while True:
        try: _refresh_cache()
        except Exception:
            with _cache_lock: _cache["refreshing"] = False
        time.sleep(CACHE_REFRESH_MINS * 60)

def _auto_backup():
    """Silently backup attendance.db on startup."""
    try:
        db_path = db_manager.DB_PATH
        if not os.path.exists(db_path): return
        os.makedirs(BACKUP_DIR, exist_ok=True)
        ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
        dst = os.path.join(BACKUP_DIR, "attendance_{0}.db".format(ts))
        shutil.copy2(db_path, dst)
        # Keep last 30 backups only
        baks = sorted([f for f in os.listdir(BACKUP_DIR) if f.endswith(".db")])
        for old in baks[:-30]:
            try: os.remove(os.path.join(BACKUP_DIR, old))
            except: pass
        print("[Backup] Auto-backup saved: {0}".format(os.path.basename(dst))); sys.stdout.flush()
    except Exception as e:
        print("[Backup] Auto-backup failed: {0}".format(e)); sys.stdout.flush()

def start_background_refresh():
    global _badge_workdays, _schedule_loaded
    # Auto-backup existing DB before anything else
    _auto_backup()
    # Init SQLite database
    try:
        db_manager.init_db()
        # Auto-import employees from CSV on first run
        if db_manager.get_db_stats()["employees"] == 0:
            try:
                emp_path = _find_employee_file()
                db_manager.import_employees_from_csv(emp_path, exclude_depts=EXCLUDE_DEPARTMENTS)
            except Exception as e:
                print("[DB] Auto-import employees skipped: {0}".format(e)); sys.stdout.flush()
    except Exception as e:
        print("[DB] SQLite init error: {0}".format(e)); sys.stdout.flush()
    try:
        _badge_workdays  = _load_schedule_from_mdb()
        _schedule_loaded = True
        print("[Schedule] Loaded {0} employee schedules".format(len(_badge_workdays))); sys.stdout.flush()
    except Exception as e:
        print("[Schedule] Using fallback ({0})".format(e)); sys.stdout.flush()
    # Load DB-stored workday overrides
    _load_workday_caches()
    # Cache UID -> Badge mapping for SQLite punch storage
    global _uid_to_badge_cache
    try:
        mdb_conn = connect_mdb()
        _uid_to_badge_cache = get_uid_to_badge_map(mdb_conn)
        mdb_conn.close()
        print("[DB] UID-to-badge map cached: {0} entries".format(len(_uid_to_badge_cache))); sys.stdout.flush()
    except Exception as e:
        print("[DB] UID-to-badge cache skipped: {0}".format(e)); sys.stdout.flush()
    t = threading.Thread(target=_background_loop)
    t.daemon = True
    t.start()
    start_auto_sync()
    start_email_scheduler()
    _init_telegram()
    start_telegram_scheduler()

# ==============================================================================
#  DATABASE MANAGEMENT ROUTES
# ==============================================================================
@app.route("/api/db/status")
@login_required
def db_status():
    try:
        stats = db_manager.get_db_stats()
        # Add MDB presence flag
        stats["mdb_available"] = any(
            f.lower().endswith(".mdb") or f.lower().endswith(".accdb")
            for f in os.listdir(SCRIPT_DIR)
        )
        return jsonify(stats)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/sql", methods=["POST"])
@admin_required
def execute_sql():
    # This endpoint is an intentional admin-only SQL console.
    # Executing arbitrary SQL is its purpose; access is restricted to admins.
    req = request.json or {}
    query = req.get("query", "").strip()
    if not query:
        return jsonify({"error": "Empty query"}), 400
        
    # Map MySQL SHOW TABLES to SQLite equivalent
    if query.strip().upper().strip(';') == "SHOW TABLES":
        query = "SELECT name, type FROM sqlite_master WHERE type IN ('table','view') AND name NOT LIKE 'sqlite_%' ORDER BY 1;"
        
    try:
        conn = get_db()
        try:
            cursor = conn.execute(query)
            if query.lstrip().upper().startswith("SELECT") or query.lstrip().upper().startswith("PRAGMA") or query.lstrip().upper().startswith("WITH"):
                rows = cursor.fetchall()
                cols = [desc[0] for desc in cursor.description] if cursor.description else []
                return jsonify({"columns": cols, "rows": [list(r) for r in rows]})
            else:
                conn.commit()
                rowcount = cursor.rowcount
                return jsonify({"message": f"Query executed successfully. Rows affected: {rowcount}"})
        finally:
            conn.close()
    except Exception as e:
        return jsonify({"error": str(e)}), 400

@app.route("/api/db/backup")
@admin_required
def db_backup():
    """Download attendance.db with headers that prevent browser blocking."""
    db_path = db_manager.DB_PATH
    if not os.path.exists(db_path):
        return jsonify({"error": "No database file found. Run a sync first."}), 404
    now   = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = "attendance_backup_{0}.db".format(now)
    with open(db_path, "rb") as f:
        data = f.read()
    resp = Response(data, status=200, mimetype="application/octet-stream")
    resp.headers["Content-Disposition"] = "attachment; filename=" + fname
    resp.headers["Content-Length"]      = str(len(data))
    resp.headers["X-Content-Type-Options"] = "nosniff"
    resp.headers["Cache-Control"]       = "no-store"
    return resp

@app.route("/api/db/backup-local", methods=["POST"])
@admin_required
def db_backup_local():
    """Save a timestamped backup copy in the backups/ subfolder."""
    db_path = db_manager.DB_PATH
    if not os.path.exists(db_path):
        return jsonify({"error": "No database file found. Run a sync first."}), 404
    os.makedirs(BACKUP_DIR, exist_ok=True)
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    dst  = os.path.join(BACKUP_DIR, "attendance_{0}.db".format(ts))
    shutil.copy2(db_path, dst)
    # Keep only last 30 backups
    backups = sorted([f for f in os.listdir(BACKUP_DIR) if f.endswith(".db")])
    for old in backups[:-30]:
        try: os.remove(os.path.join(BACKUP_DIR, old))
        except: pass
    size_mb = round(os.path.getsize(dst) / (1024*1024), 2)
    return jsonify({"ok": True, "path": dst, "filename": os.path.basename(dst), "size_mb": size_mb})

@app.route("/api/db/import-csv", methods=["POST"])
@admin_required
def db_import_csv():
    try:
        emp_path = _find_employee_file()
        n = db_manager.import_employees_from_csv(emp_path, exclude_depts=EXCLUDE_DEPARTMENTS)
        return jsonify({"ok": True, "imported": n})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/db/import-mdb", methods=["POST"])
@admin_required
def db_import_mdb():
    try:
        conn     = connect_mdb()
        uid_map  = get_uid_to_badge_map(conn)
        mdb_path = _find_mdb()
        inserted, skipped = db_manager.import_punches_from_mdb(
            mdb_path, uid_map, CHECKINOUT_TABLE, COL_USER_ID, COL_CHECKTIME
        )
        conn.close()
        return jsonify({"ok": True, "inserted": inserted, "skipped_unknown": skipped})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/db/resolve-unknown", methods=["POST"])
@admin_required
def db_resolve_unknown():
    """
    Assign a badge to a device UID so attendance is correctly attributed.
    Optionally create/update the employee record with name + department.
    Also re-attributes any punch records stored under the raw UID.
    """
    data  = request.get_json() or {}
    ip    = (data.get("device_ip") or "").strip()
    uid   = (data.get("uid")       or "").strip()
    badge = (data.get("badge")     or "").strip()
    name  = (data.get("name")      or "").strip()
    dept  = (data.get("dept")      or "").strip()

    if not ip or not uid or not badge:
        return jsonify({"error": "device_ip, uid, and badge are all required"}), 400
    try:
        db_manager.resolve_unknown_user(ip, uid)

        # Update UID->badge cache so future device syncs map correctly
        global _uid_to_badge_cache
        _uid_to_badge_cache[uid] = badge

        conn = get_db()
        now  = datetime.now().isoformat()
        try:
            # Re-attribute punch records stored under the raw UID to the real badge.
            # Re-attribute across ALL devices (not just the specified one) because the
            # same employee may have punched on multiple devices with the same UID, and
            # all those records should be corrected at once.
            punches_updated = 0
            if uid != badge:
                conn.execute(
                    "UPDATE punches SET badge=? WHERE badge=?",
                    (badge, uid)
                )
                punches_updated = conn.execute("SELECT changes()").fetchone()[0]

            # Create or update the employee record
            existing = conn.execute(
                "SELECT badge FROM employees WHERE badge=?", (badge,)
            ).fetchone()

            if existing:
                # Update only fields that were provided
                sets = ["updated_at=?"]
                vals = [now]
                if name: sets.insert(0, "name=?"); vals.insert(0, name)
                if dept: sets.insert(0, "dept=?"); vals.insert(0, dept)
                vals.append(badge)
                conn.execute(
                    "UPDATE employees SET {0} WHERE badge=?".format(", ".join(sets)), vals
                )
            else:
                # Create new employee record
                conn.execute(
                    "INSERT INTO employees (badge, name, dept, active, updated_at) VALUES (?,?,?,?,?)",
                    (badge, name or uid, dept or "UNKNOWN", 1, now)
                )

            conn.commit()
        finally:
            conn.close()

        write_audit(session.get("username","?"), "RESOLVE_UNKNOWN",
                    "uid={0} ip={1} -> badge={2} name={3} dept={4} punches_updated={5}".format(
                        uid, ip, badge, name, dept, punches_updated),
                    request.remote_addr)
        return jsonify({"ok": True, "punches_updated": punches_updated})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/db/auto-map-unknown", methods=["POST"])
@admin_required
def db_auto_map_unknown():
    """
    Auto-map unmapped device UIDs where the UID exactly matches an employee badge.
    This covers the common case where the ZK device stores the badge number as the UID.
    Returns counts of mapped, skipped (no match), and already resolved.
    """
    try:
        conn = get_db()
        try:
            unresolved = conn.execute(
                "SELECT device_ip, uid FROM unknown_users WHERE resolved=0"
            ).fetchall()

            mapped = 0; skipped = 0
            now = datetime.now().isoformat()

            global _uid_to_badge_cache

            results = []

            for row in unresolved:
                ip  = row["device_ip"]
                uid = row["uid"]

                # Check if uid directly matches an employee badge
                emp = conn.execute(
                    "SELECT badge, name, dept FROM employees WHERE badge=?", (uid,)
                ).fetchone()

                if not emp:
                    skipped += 1
                    results.append({"uid": uid, "device_ip": ip, "status": "skipped",
                                     "reason": "No employee with badge matching this UID"})
                    continue

                badge = emp["badge"]

                # Mark resolved
                conn.execute(
                    "UPDATE unknown_users SET resolved=1 WHERE device_ip=? AND uid=?",
                    (ip, uid)
                )

                # Update UID->badge cache
                _uid_to_badge_cache[uid] = badge

                # Re-attribute any punch records stored under the raw UID across ALL devices.
                # An employee may have been enrolled on multiple devices with the same UID,
                # so we correct all punch records that are stored under the raw UID at once.
                punches_updated = 0
                if uid != badge:
                    conn.execute(
                        "UPDATE punches SET badge=? WHERE badge=?",
                        (badge, uid)
                    )
                    punches_updated = conn.execute(
                        "SELECT changes()"
                    ).fetchone()[0]

                mapped += 1
                results.append({"uid": uid, "device_ip": ip, "status": "mapped",
                                 "badge": badge, "name": emp["name"], "dept": emp["dept"],
                                 "punches_updated": punches_updated})

            conn.commit()
        finally:
            conn.close()

        write_audit(session.get("username","?"), "AUTO_MAP_UNKNOWN",
                    "mapped={0} skipped={1}".format(mapped, skipped),
                    request.remote_addr)

        return jsonify({
            "ok":      True,
            "mapped":  mapped,
            "skipped": skipped,
            "results": results,
            "message": "{0} UID(s) auto-mapped. {1} had no matching employee badge.".format(
                        mapped, skipped),
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/db/unknown-users")
@admin_required
def db_unknown_users():
    try:
        rows = db_manager.get_unknown_users()
        return jsonify(rows)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ==============================================================================
#  STATIC ROUTES
# ==============================================================================
@app.route("/")
def index():
    return send_from_directory(SCRIPT_DIR, "d.html")

@app.route("/d.html")
def dashboard_html():
    return send_from_directory(SCRIPT_DIR, "d.html")

@app.route("/d")
def dashboard_d():
    return send_from_directory(SCRIPT_DIR, "d.html")

@app.route("/dashboard")
@app.route("/dashboard.html")
def dashboard_alias():
    return send_from_directory(SCRIPT_DIR, "d.html")

@app.route("/zk")
def zk_page():
    return send_from_directory(SCRIPT_DIR, "zk.html")

@app.route("/ad")
def ad_page():
    return send_from_directory(SCRIPT_DIR, "ad.html")


@app.route("/static/<path:filename>")
def static_files(filename):
    return send_from_directory(os.path.join(SCRIPT_DIR, "static"), filename)

@app.route("/js/<path:filename>")
def js_files(filename):
    return send_from_directory(os.path.join(SCRIPT_DIR, "js"), filename)

# ==============================================================================
#  CACHE API
# ==============================================================================
@app.route("/api/cache/status")
@login_required
def cache_status():
    with _cache_lock:
        last = _cache["last_updated"]
        age  = int((datetime.now()-last).total_seconds()) if last else None
    return jsonify({
        "last_updated": last.strftime("%H:%M:%S") if last else None,
        "age_seconds":  age, "refreshing": _cache["refreshing"],
        "has_data":     _cache["today"] is not None, "error": _cache["error"]
    })

@app.route("/api/cache/refresh", methods=["POST"])
@permission_required("force_refresh")
def force_refresh():
    t = threading.Thread(target=_refresh_cache)
    t.daemon = True
    t.start()
    return jsonify({"status": "Refresh started"})

@app.route("/api/cache/force-refresh", methods=["POST"])
@permission_required("force_refresh")
def force_full_refresh():
    """Full sync: re-pulls from devices then rebuilds the cache."""
    global _local_sync_time
    _local_sync_time = 0  # reset sync timer to force re-pull on next _ensure_synced call
    t = threading.Thread(target=_refresh_cache)
    t.daemon = True
    t.start()
    return jsonify({"status": "Full refresh started"})

# ==============================================================================
#  TODAY
# ==============================================================================
@app.route("/api/today")
@permission_required("view_today")
def today_report():
    with _cache_lock:
        data = _cache["today"]; last = _cache["last_updated"]; refreshing = _cache["refreshing"]
    if data is None:
        if refreshing:
            return jsonify({"error": "Still loading -- first pull in progress...", "loading": True}), 202
        return jsonify({"error": "No data yet. Click Refresh to start.", "loading": False}), 503
    age  = int((datetime.now()-last).total_seconds()) if last else 0
    data = dict(data); data["cache_age_secs"] = age; data["refreshing"] = refreshing
    # Supplement cached present/absent lists with the latest SQLite punch data so
    # that punches arriving after the last background refresh are reflected
    # immediately -- this mirrors what history_report() does at query time.
    # Guard: only supplement when the cache was built for today.  After midnight
    # the cache may still hold yesterday's lists; mixing them with today's SQLite
    # punches would produce wrong present/absent counts until the next refresh.
    cache_is_for_today = (last is not None and last.date() == date.today())
    if cache_is_for_today:
        try:
            today_str = date.today().strftime("%Y-%m-%d")
            db_sup = get_db()
            sup_rows = db_sup.execute(
                "SELECT DISTINCT badge FROM punches WHERE punch_time >= ? AND punch_time <= ?",
                (today_str + " 00:00:00", today_str + " 23:59:59")
            ).fetchall()
            db_sup.close()
            latest_badges = {str(r["badge"]).strip() for r in sup_rows if r["badge"]}
            # Move any employee in absent whose badge now appears in SQLite to present
            cached_present_codes = {e["code"] for e in (data.get("present") or [])}
            newly_present = [e for e in (data.get("absent") or []) if e["code"] in latest_badges and e["code"] not in cached_present_codes]
            if newly_present:
                updated_present = list(data.get("present") or []) + newly_present
                updated_absent  = [e for e in (data.get("absent") or []) if e["code"] not in latest_badges]
                updated_present.sort(key=lambda x: (_dept_sort(x["dept"]), x["name"]))
                updated_absent.sort(key=lambda x:  (_dept_sort(x["dept"]), x["name"]))
                data["present"]       = updated_present
                data["absent"]        = updated_absent
                data["present_count"] = len(updated_present)
                data["absent_count"]  = len(updated_absent)
                data["working_today"] = len(updated_present) + len(updated_absent)
        except Exception as e:
            print("[Today] Live supplement error: {0}".format(e)); sys.stdout.flush()
    return jsonify(data)

# ==============================================================================
#  EMPLOYEES
# ==============================================================================
@app.route("/api/employees")
@permission_required("view_employees")
def employees_list():
    try:
        df = load_employees_all()
        records = df[["Badgenumber","Name","DEPTNAME"]].rename(
            columns={"Badgenumber":"code","Name":"name","DEPTNAME":"dept"}
        ).to_dict("records")
        # Merge active flag from SQLite DB
        try:
            db_conn = get_db()
            db_rows = db_conn.execute("SELECT badge, active FROM employees").fetchall()
            db_conn.close()
            active_map = {r["badge"]: r["active"] for r in db_rows}
            for rec in records:
                rec["active"] = active_map.get(rec["code"], 1)
        except Exception:
            for rec in records:
                rec["active"] = 1
        depts = sorted(df["DEPTNAME"].unique().tolist())
        excl = [d.upper() for d in EXCLUDE_DEPARTMENTS]
        active_count = len([r for r in records if r.get("active", 1) and
                            r["dept"].upper() not in excl])
        return jsonify({
            "total":          len(records),
            "active_count":   active_count,
            "employees":      records,
            "departments":    depts,
            "collapsed_depts": COLLAPSED_DEPARTMENTS,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ==============================================================================
#  SCHEDULE
# ==============================================================================
@app.route("/api/schedule/info")
@login_required
def schedule_info():
    global _badge_workdays, _schedule_loaded
    day_names = {0:"Mon",1:"Tue",2:"Wed",3:"Thu",4:"Fri",5:"Sat",6:"Sun"}
    summary = {b: sorted([day_names[d] for d in days]) for b, days in list(_badge_workdays.items())[:5]}
    return jsonify({
        "loaded":        _schedule_loaded,
        "total_mapped":  len(_badge_workdays),
        "fallback_used": len(_badge_workdays) == 0,
        "sample":        summary,
        "fallback_config": {dept: sorted([day_names[d] for d in days]) for dept, days in DEPT_WORKDAYS_FALLBACK.items()}
    })

@app.route("/api/schedule/reload", methods=["POST"])
@admin_required
def schedule_reload():
    global _badge_workdays, _schedule_loaded
    try:
        _badge_workdays  = _load_schedule_from_mdb()
        _schedule_loaded = True
        return jsonify({"status": "ok", "total_mapped": len(_badge_workdays)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ==============================================================================
#  DEVICES
# ==============================================================================
@app.route("/api/devices/list")
@login_required
def devices_list():
    return jsonify({"ips": DEVICE_IPS, "names": _device_names})

@app.route("/api/devices/status")
@permission_required("view_devices")
def devices_status():
    results = {}
    try:
        with ThreadPoolExecutor(max_workers=len(DEVICE_IPS)) as ex:
            futures = {ex.submit(_check_device_status, ip): ip for ip in DEVICE_IPS}
            for f in as_completed(futures, timeout=12):
                try:
                    r = f.result(); results[r["ip"]] = r
                except Exception: pass
    except Exception as e:
        print("[Devices] Status error: {0}".format(e)); sys.stdout.flush()
    out = []
    for ip in DEVICE_IPS:
        d = results.get(ip, {"ip": ip, "online": False, "error": "Timeout"})
        d["name"] = _device_names.get(ip, "")
        out.append(d)
    return jsonify(out)

@app.route("/api/device/<ip>/users")
@permission_required("view_devices")
def device_users(ip):
    if ip not in DEVICE_IPS: return jsonify({"error": "Unknown device"}), 400
    zk_conn = None
    try:
        from zk import ZK
        zk_conn = ZK(ip, port=DEVICE_PORT, timeout=30, verbose=False).connect()
        users   = zk_conn.get_users()
        data    = [{"uid": str(u.user_id), "name": u.name, "privilege": u.privilege} for u in users]
        return jsonify({"ip": ip, "count": len(data), "users": data})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if zk_conn:
            try: zk_conn.disconnect()
            except: pass

@app.route("/api/device/<ip>/attendance")
@permission_required("view_devices")
def device_attendance(ip):
    if ip not in DEVICE_IPS: return jsonify({"error": "Unknown device"}), 400
    date_str = request.args.get("date", date.today().strftime("%Y-%m-%d"))
    try: filter_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    except: return jsonify({"error": "Use ?date=YYYY-MM-DD"}), 400
    zk_conn = None
    try:
        from zk import ZK
        zk_conn = ZK(ip, port=DEVICE_PORT, timeout=DEVICE_TIMEOUT, verbose=False).connect()
        records = zk_conn.get_attendance()
        data    = [{"uid": str(r.user_id), "time": r.timestamp.strftime("%H:%M:%S"), "type": str(r.punch)}
                   for r in records if r.timestamp.date() == filter_date]
        data.sort(key=lambda x: x["time"])
        return jsonify({"ip": ip, "date": date_str, "count": len(data), "records": data})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if zk_conn:
            try: zk_conn.disconnect()
            except: pass

# ==============================================================================
#  EMPLOYEE INDIVIDUAL REPORT
# ==============================================================================
@app.route("/api/employee/<badge>/report")
@login_required
def employee_report(badge):
    # Permission check: employees can only see their own calendar unless view_all_calendars
    users = _get_users()
    u     = users.get(session.get("username", ""), {})
    if u.get("role") != "admin":
        own_badge = u.get("badge", "")
        can_all   = u.get("permissions", {}).get("view_all_calendars", False)
        can_own   = u.get("permissions", {}).get("view_own_calendar", False)
        if not can_all:
            if not (can_own and own_badge == badge):
                return jsonify({"error": "Permission denied"}), 403

    from_str = request.args.get("from")
    to_str   = request.args.get("to")
    source   = request.args.get("source", "mdb")   # "mdb" or "device"
    if not from_str or not to_str:
        return jsonify({"error": "Provide ?from=DD/MM/YYYY&to=DD/MM/YYYY"}), 400
    try:
        date_from = datetime.strptime(from_str, "%d/%m/%Y").date()
        date_to   = datetime.strptime(to_str,   "%d/%m/%Y").date()
    except:
        return jsonify({"error": "Invalid date format. Use DD/MM/YYYY"}), 400

    try:
        emp_df  = load_employees_all()
        emp_row = emp_df[emp_df["Badgenumber"] == badge]
        if emp_row.empty:
            db_emp = db_manager.get_employee(badge)
            if not db_emp:
                return jsonify({"error": "Employee {0} not found".format(badge)}), 404
            emp_name = db_emp["name"]
            emp_dept = db_emp["dept"]
        else:
            emp      = emp_row.iloc[0]
            emp_name = emp["Name"]
            emp_dept = emp["DEPTNAME"]

        if source == "device":
            import concurrent.futures
            # Pull directly from ZK devices
            punch_map = {}   # date -> list of times
            
            def fetch_from_device(ip):
                local_punches = []
                zk_conn = None
                try:
                    from zk import ZK
                    zk_conn = ZK(ip, port=DEVICE_PORT, timeout=DEVICE_TIMEOUT, verbose=False).connect()
                    records = zk_conn.get_attendance()
                    b2u_map = {}
                    try:
                        users_list = zk_conn.get_users()
                        for zu in users_list:
                            b2u_map[str(zu.user_id).strip()] = str(zu.user_id).strip()
                    except: pass
                    for r in records:
                        if r.timestamp.date() < date_from or r.timestamp.date() > date_to: continue
                        if str(r.user_id).strip() != badge and b2u_map.get(str(r.user_id).strip()) != badge: continue
                        local_punches.append({"date": r.timestamp.date(), "time": r.timestamp.strftime("%H:%M:%S"), "machine": _device_names.get(ip, ip)})
                except: pass
                finally:
                    if zk_conn:
                        try: zk_conn.disconnect()
                        except: pass
                return local_punches

            with concurrent.futures.ThreadPoolExecutor(max_workers=min(10, len(DEVICE_IPS) or 1)) as executor:
                futures = [executor.submit(fetch_from_device, ip) for ip in DEVICE_IPS]
                for future in concurrent.futures.as_completed(futures):
                    for p in future.result():
                        d_key = p["date"]
                        if d_key not in punch_map: punch_map[d_key] = []
                        punch_map[d_key].append({"time": p["time"], "machine": p["machine"]})

            # Build response
            all_dates = list(pd.date_range(date_from, date_to).date)
            days = []; working_days = 0; present_count = 0; holiday_duty_days = 0
            for d in all_dates:
                is_work   = _is_working_day(d, badge, emp_dept)
                punches   = sorted(punch_map.get(d, []), key=lambda x: x["time"])
                is_present = len(punches) > 0
                is_holiday_duty = is_present and not is_work
                if is_work:
                    working_days += 1
                    if is_present: present_count += 1
                if is_holiday_duty:
                    holiday_duty_days += 1
                late_info = _check_late_early(punches, emp_dept) if (is_present and is_work) else {"late":False,"early_departure":False,"late_mins":0,"early_mins":0}
                days.append({
                    "date": d.strftime("%d/%m/%Y"), "date_iso": d.strftime("%Y-%m-%d"),
                    "day": d.strftime("%A"), "present": is_present,
                    "working_day": is_work, "holiday_duty": is_holiday_duty, "punches": punches,
                    "late": late_info["late"], "late_mins": late_info["late_mins"],
                    "early_departure": late_info["early_departure"], "early_mins": late_info["early_mins"],
                })
        else:
            # Pull from SQLite punches table (populated from device pulls).
            start_s = date_from.strftime("%Y-%m-%d") + " 00:00:00"
            end_s   = date_to.strftime("%Y-%m-%d")   + " 23:59:59"
            db_punch = get_db()
            punch_rows_emp = db_punch.execute(
                "SELECT punch_time, device_ip FROM punches "
                "WHERE badge=? AND punch_time >= ? AND punch_time <= ? ORDER BY punch_time",
                (badge, start_s, end_s)
            ).fetchall()
            db_punch.close()
            punch_map = {}   # date -> list of {time, machine}
            for row in punch_rows_emp:
                try:
                    dt_obj = datetime.strptime(row["punch_time"], "%Y-%m-%d %H:%M:%S")
                    punch_map.setdefault(dt_obj.date(), []).append({
                        "time":    dt_obj.strftime("%H:%M:%S"),
                        "machine": _device_names.get(row["device_ip"], row["device_ip"] or "N/A"),
                    })
                except Exception:
                    pass
            all_dates = list(pd.date_range(date_from, date_to).date)
            days = []; working_days = 0; present_count = 0; holiday_duty_days = 0
            for d in all_dates:
                is_work   = _is_working_day(d, badge, emp_dept)
                punches   = sorted(punch_map.get(d, []), key=lambda x: x["time"])
                is_present = len(punches) > 0
                is_holiday_duty = is_present and not is_work
                if is_work:
                    working_days += 1
                    if is_present: present_count += 1
                if is_holiday_duty:
                    holiday_duty_days += 1
                late_info = _check_late_early(punches, emp_dept) if (is_present and is_work) else {"late":False,"early_departure":False,"late_mins":0,"early_mins":0}
                days.append({
                    "date": d.strftime("%d/%m/%Y"), "date_iso": d.strftime("%Y-%m-%d"),
                    "day": d.strftime("%A"), "present": is_present,
                    "working_day": is_work, "holiday_duty": is_holiday_duty, "punches": punches,
                    "late": late_info["late"], "late_mins": late_info["late_mins"],
                    "early_departure": late_info["early_departure"], "early_mins": late_info["early_mins"],
                })

    except Exception as e:
        return jsonify({"error": str(e)}), 500

    late_days  = sum(1 for d in days if d.get("late"))
    early_days = sum(1 for d in days if d.get("early_departure"))
    shift_info = _get_shift(emp_dept)
    return jsonify({
        "badge":             badge,
        "name":              emp_name,
        "dept":              emp_dept,
        "from":              date_from.strftime("%d %b %Y"),
        "to":                date_to.strftime("%d %b %Y"),
        "total_days":        len(days),
        "working_days":      working_days,
        "present_days":      present_count,
        "absent_days":       working_days - present_count,
        "holiday_duty_days": holiday_duty_days,
        "late_days":         late_days,
        "early_days":        early_days,
        "shift":             {"start": shift_info["start"], "end": shift_info["end"], "grace": shift_info["grace"]},
        "days":              days,
        "source":            source,
    })

# ==============================================================================
#  MONTHLY SUMMARY API
# ==============================================================================
@app.route("/api/employee/<badge>/monthly-summary")
@login_required
def employee_monthly_summary(badge):
    # Permission check: same as report
    users = _get_users()
    u     = users.get(session.get("username", ""), {})
    if u.get("role") != "admin":
        own_badge = u.get("badge", "")
        can_all   = u.get("permissions", {}).get("view_all_calendars", False)
        can_own   = u.get("permissions", {}).get("view_own_calendar", False)
        if not can_all:
            if not (can_own and own_badge == badge):
                return jsonify({"error": "Permission denied"}), 403

    month_str = request.args.get("month")   # format YYYY-MM
    if not month_str:
        return jsonify({"error": "Provide ?month=YYYY-MM"}), 400
    try:
        yr, mn = map(int, month_str.split('-'))
        date_from = date(yr, mn, 1)
        last_day = calendar.monthrange(yr, mn)[1]
        date_to = date(yr, mn, last_day)
    except:
        return jsonify({"error": "Invalid month format. Use YYYY-MM"}), 400

    try:
        emp_df   = load_employees()
        emp_row  = emp_df[emp_df["Badgenumber"] == badge]
        if emp_row.empty: return jsonify({"error": "Employee not found"}), 404
        emp_name = str(emp_row.iloc[0]["Name"])
        emp_dept = str(emp_row.iloc[0]["DEPTNAME"])

        # Load punch data from SQLite punches table (populated from device pulls).
        start_s = date_from.strftime("%Y-%m-%d") + " 00:00:00"
        end_s   = date_to.strftime("%Y-%m-%d")   + " 23:59:59"
        db_punch = get_db()
        punch_rows_mo = db_punch.execute(
            "SELECT punch_time FROM punches WHERE badge=? AND punch_time >= ? AND punch_time <= ? ORDER BY punch_time",
            (badge, start_s, end_s)
        ).fetchall()
        db_punch.close()
        # Build date → list of punch times
        day_punch_times = {}   # date obj -> list of "HH:MM:SS"
        for row in punch_rows_mo:
            try:
                dt_obj = datetime.strptime(row["punch_time"], "%Y-%m-%d %H:%M:%S")
                day_punch_times.setdefault(dt_obj.date(), []).append(dt_obj.strftime("%H:%M:%S"))
            except Exception:
                pass

        all_dates = list(pd.date_range(date_from, date_to).date)
        working_days = 0; present_count = 0; holiday_duty_days = 0
        late_days = 0; early_days = 0

        for d in all_dates:
            is_work    = _is_working_day(d, badge, emp_dept)
            times_list = day_punch_times.get(d, [])
            is_present = len(times_list) > 0
            is_holiday_duty = is_present and not is_work
            if is_work:
                working_days += 1
                if is_present: present_count += 1
            if is_holiday_duty:
                holiday_duty_days += 1

            punches = sorted([{"time": t} for t in times_list], key=lambda x: x["time"])
            if is_present and is_work:
                late_info = _check_late_early(punches, emp_dept)
                if late_info.get("late"): late_days += 1
                if late_info.get("early_departure"): early_days += 1

        return jsonify({
            "badge": badge,
            "name": emp_name,
            "month": month_str,
            "working_days": working_days,
            "present_days": present_count,
            "absent_days": working_days - present_count,
            "late_count": late_days,
            "early_departure_count": early_days,
            "holiday_duty_days": holiday_duty_days
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ==============================================================================
#  HISTORY
# ==============================================================================
@app.route("/api/history")
@permission_required("view_history")
def history_report():
    from_str = request.args.get("from"); to_str = request.args.get("to")
    if not from_str or not to_str:
        return jsonify({"error": "Provide ?from=DD/MM/YYYY&to=DD/MM/YYYY"}), 400
    try:
        date_from = datetime.strptime(from_str, "%d/%m/%Y").date()
        date_to   = datetime.strptime(to_str,   "%d/%m/%Y").date()
    except:
        return jsonify({"error": "Invalid date format. Use DD/MM/YYYY"}), 400
    try:
        emp_df = load_employees()
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    # Load punch data from SQLite punches table (populated from device pulls).
    start_dt = date_from.strftime("%Y-%m-%d") + " 00:00:00"
    end_dt   = date_to.strftime("%Y-%m-%d")   + " 23:59:59"
    date_badges = {}   # date obj -> set of badge strings
    try:
        db_punch = get_db()
        punch_rows = db_punch.execute(
            "SELECT badge, punch_time FROM punches WHERE punch_time >= ? AND punch_time <= ?",
            (start_dt, end_dt)
        ).fetchall()
        db_punch.close()
        for row in punch_rows:
            try:
                pt = datetime.strptime(row["punch_time"][:10], "%Y-%m-%d").date()
                date_badges.setdefault(pt, set()).add(str(row["badge"]).strip())
            except Exception:
                pass
    except Exception as e:
        print("[History] Punch load error: {0}".format(e))
    # Supplement today's punch badges with the in-memory cache (populated from
    # direct device polling), so the history report reflects the same present
    # count as the Today view rather than only punches stored to SQLite so far.
    today_date = date.today()
    if date_from <= today_date <= date_to:
        with _cache_lock:
            cached_today = _cache.get("today")
        if cached_today:
            cached_present_badges = {e["code"] for e in (cached_today.get("present") or [])}
            date_badges.setdefault(today_date, set()).update(cached_present_badges)

    all_dates = pd.date_range(date_from, date_to).date
    days_data = []

    # Get DB holidays across the date range *once*
    date_str_from = date_from.strftime("%Y-%m-%d")
    date_str_to   = date_to.strftime("%Y-%m-%d")
    db = get_db()
    holidays = db.execute("SELECT * FROM holidays WHERE date <= ? AND date_end >= ?", (date_str_to, date_str_from)).fetchall()
    db.close()

    def _check_holiday_quick(d_str, badge, dept):
        for r in holidays:
            if r["date"] <= d_str <= r["date_end"]:
                if r["scope"] == 'all': return True
                if r["scope"] == 'dept' and r["dept"] and dept and r["dept"].upper() == dept.upper(): return True
                if r["scope"] == 'employee' and r["employees"] and badge and badge in [c.strip() for c in r["employees"].split(',')]: return True
        return False

    for d in all_dates:
        d_str = d.strftime("%Y-%m-%d")
        present_badges = date_badges.get(d, set())
        absent, present = [], []
        for _, emp in emp_df.iterrows():
            badge = emp["Badgenumber"]; dept = emp["DEPTNAME"]
            is_work = not _check_holiday_quick(d_str, badge, dept) and (d.weekday() in _get_workdays_for_badge(badge, dept))
            if not is_work: continue
            rec = {"name": emp["Name"], "code": badge, "dept": dept}
            if badge in present_badges:
                present.append(rec)
            else:
                absent.append(rec)
        absent.sort(key=lambda x: (_dept_sort(x["dept"]), x["name"]))
        present.sort(key=lambda x: (_dept_sort(x["dept"]), x["name"]))
        days_data.append({
            "date": d.strftime("%d %b %Y"), "date_iso": d_str,
            "present_count": len(present), "absent_count": len(absent),
            "present": present, "absent": absent,
        })
    return jsonify({
        "from": date_from.strftime("%d %b %Y"), "to": date_to.strftime("%d %b %Y"),
        "total_emps": len(emp_df), "days": days_data
    })

# ==============================================================================
#  EXCEL / PDF HELPERS
# ==============================================================================
def _thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _build_absent_wb_single_sheet(absent_list, title):
    wb = Workbook(); ws = wb.active; ws.title = "Absent Report"
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:D1")
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name="Arial", bold=True, size=13, color="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.cell(row=2, column=1,
        value="Generated: {0}".format(datetime.now().strftime('%d %b %Y %H:%M'))).font = \
        Font(name="Arial", size=9, italic=True, color="888888")
    for col, hdr in enumerate(["Emp Code","Name","Department","Date"], 1):
        c = ws.cell(row=4, column=col, value=hdr)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", start_color="1F4E79")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _thin()
    ws.row_dimensions[4].height = 18
    sorted_list = sorted(absent_list, key=lambda r: (_dept_sort(r.get("dept","")), r.get("dept",""), r.get("name",""), r.get("date","")))
    row = 5; prev_dept = None
    for i, rec in enumerate(sorted_list):
        dept = rec.get("dept", "")
        if dept != prev_dept:
            for col in range(1, 5):
                c = ws.cell(row=row, column=col, value=dept if col==2 else "")
                c.fill = PatternFill("solid", start_color="DCE6F1")
                c.font = Font(name="Arial", bold=True, size=10, color="1F4E79")
                c.border = _thin()
            row += 1; prev_dept = dept
        bg = "EBF3FB" if i%2==0 else "FFFFFF"
        for col, val in enumerate([rec.get("code",""), rec.get("name",""), rec.get("dept",""), rec.get("date","")], 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(name="Arial", size=10)
            c.fill = PatternFill("solid", start_color=bg)
            c.border = _thin()
            c.alignment = Alignment(horizontal="center" if col in (1,4) else "left", vertical="center")
        row += 1
    for col, w in zip("ABCD", [12,30,24,14]): ws.column_dimensions[col].width = w
    return wb

@app.route("/api/today/export")
@permission_required("export_reports")
def export_today():
    with _cache_lock: data = _cache["today"]
    if not data: return jsonify({"error": "No data cached yet"}), 503
    absent = [dict(r, date=data["date"]) for r in data["absent"]]
    wb = _build_absent_wb_single_sheet(absent, "Absent Report -- {0}".format(data['date']))
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, download_name="absent_{0}.xlsx".format(date.today().strftime('%Y%m%d')),
                     as_attachment=True, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/api/punches")
@login_required
def get_raw_punches():
    """Return raw punch records for the Attendance Logs page. Accepts ?from=YYYY-MM-DD&to=YYYY-MM-DD&badge="""
    from_str  = request.args.get("from", "")
    to_str    = request.args.get("to",   "")
    badge_str = request.args.get("badge", "").strip()
    if not from_str or not to_str:
        return jsonify({"error": "Provide ?from=YYYY-MM-DD&to=YYYY-MM-DD"}), 400
    try:
        start_dt = from_str + " 00:00:00"
        end_dt   = to_str   + " 23:59:59"
    except Exception:
        return jsonify({"error": "Invalid date format. Use YYYY-MM-DD"}), 400
    try:
        conn = get_db()
        if badge_str:
            rows = conn.execute(
                "SELECT p.badge, p.punch_time, p.device_ip, e.name, e.dept "
                "FROM punches p LEFT JOIN employees e ON p.badge=e.badge "
                "WHERE p.punch_time >= ? AND p.punch_time <= ? AND p.badge=? "
                "ORDER BY p.punch_time DESC LIMIT 5000",
                (start_dt, end_dt, badge_str)
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT p.badge, p.punch_time, p.device_ip, e.name, e.dept "
                "FROM punches p LEFT JOIN employees e ON p.badge=e.badge "
                "WHERE p.punch_time >= ? AND p.punch_time <= ? "
                "ORDER BY p.punch_time DESC LIMIT 5000",
                (start_dt, end_dt)
            ).fetchall()
        conn.close()
        return jsonify([{
            "badge":      r["badge"],
            "punch_time": r["punch_time"],
            "device_ip":  r["device_ip"],
            "name":       r["name"]  or "",
            "dept":       r["dept"]  or "",
        } for r in rows])
    except Exception as e:
        return jsonify({"error": "Failed to load punch records"}), 500

@app.route("/api/history/export")
@permission_required("export_reports")
def export_history():
    from_str = request.args.get("from"); to_str = request.args.get("to")
    if not from_str or not to_str:
        return jsonify({"error": "Provide ?from=DD/MM/YYYY&to=DD/MM/YYYY"}), 400
    try:
        date_from = datetime.strptime(from_str, "%d/%m/%Y").date()
        date_to   = datetime.strptime(to_str,   "%d/%m/%Y").date()
    except (TypeError, ValueError):
        return jsonify({"error": "Invalid dates. Use DD/MM/YYYY"}), 400
    try:
        emp_df = load_employees()
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    # Load punch data from SQLite punches table (populated from device pulls).
    start_dt = date_from.strftime("%Y-%m-%d") + " 00:00:00"
    end_dt   = date_to.strftime("%Y-%m-%d")   + " 23:59:59"
    date_badges_ex = {}   # date obj -> set of badges
    try:
        db_punch = get_db()
        punch_rows_ex = db_punch.execute(
            "SELECT badge, punch_time FROM punches WHERE punch_time >= ? AND punch_time <= ?",
            (start_dt, end_dt)
        ).fetchall()
        db_punch.close()
        for row in punch_rows_ex:
            try:
                pt = datetime.strptime(row["punch_time"][:10], "%Y-%m-%d").date()
                date_badges_ex.setdefault(pt, set()).add(str(row["badge"]).strip())
            except Exception:
                pass
    except Exception as e:
        return jsonify({"error": "Could not load punch data: " + str(e)}), 500
    # Supplement today's badges with the in-memory cache (same fix as history_report).
    today_date = date.today()
    if date_from <= today_date <= date_to:
        with _cache_lock:
            cached_today = _cache.get("today")
        if cached_today:
            cached_present_badges = {e["code"] for e in (cached_today.get("present") or [])}
            date_badges_ex.setdefault(today_date, set()).update(cached_present_badges)
    all_absent = []
    for d in pd.date_range(date_from, date_to).date:
        present = date_badges_ex.get(d, set())
        for _, emp in emp_df.iterrows():
            badge = emp["Badgenumber"]; dept = emp["DEPTNAME"]
            if not _is_working_day(d, badge, dept): continue
            if badge not in present:
                all_absent.append({"code": badge, "name": emp["Name"], "dept": dept, "date": d.strftime("%d/%m/%Y")})
    wb = _build_absent_wb_single_sheet(
        all_absent,
        "Absent History -- {0} to {1}".format(date_from.strftime('%d %b'), date_to.strftime('%d %b %Y'))
    )
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = "absent_{0}_{1}.xlsx".format(date_from.strftime('%Y%m%d'), date_to.strftime('%Y%m%d'))
    return send_file(buf, download_name=fname, as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def _build_employee_xlsx(report):
    wb = Workbook(); ws = wb.active; ws.title = "Attendance"
    ws.sheet_view.showGridLines = False
    def thin():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)
    ws.merge_cells("A1:F1")
    c = ws.cell(row=1, column=1, value="Attendance Report -- {0}  ({1})".format(report['name'], report['badge']))
    c.font = Font(name="Arial", bold=True, size=13, color="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A2:F2")
    c = ws.cell(row=2, column=1, value="Department: {0}   |   {1} -> {2}".format(report['dept'], report['from'], report['to']))
    c.font = Font(name="Arial", size=10, italic=True, color="666666")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18
    ws.merge_cells("A3:F3")
    c = ws.cell(row=3, column=1,
        value="Present: {0} days   |   Absent: {1} days   |   Total: {2} days".format(report['present_days'], report['absent_days'], report['total_days']))
    c.font = Font(name="Arial", bold=True, size=11)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 20
    for col, hdr in enumerate(["Date","Day","Status","Punch Times","Machines","Count"], 1):
        c = ws.cell(row=5, column=col, value=hdr)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", start_color="1F4E79")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin()
    ws.row_dimensions[5].height = 18
    row = 6
    for i, day in enumerate(report["days"]):
        is_holiday_duty = day.get("holiday_duty", False)
        bg          = "EDE9FE" if is_holiday_duty else ("E8F5E9" if day["present"] else "FFEBEE")
        status      = "HOLIDAY DUTY" if is_holiday_duty else ("PRESENT" if day["present"] else "ABSENT")
        punch_times = " | ".join(p["time"] for p in day["punches"])
        machines    = " | ".join(dict.fromkeys(p["machine"] for p in day["punches"]))
        vals = [day["date"], day["day"], status, punch_times, machines, len(day["punches"])]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(name="Arial", size=10)
            c.fill = PatternFill("solid", start_color=bg)
            c.border = thin()
            c.alignment = Alignment(horizontal="center" if col in (1,2,3,6) else "left", vertical="center")
        if day.get("holiday_duty"):
            ws.cell(row=row, column=3).font = Font(name="Arial", size=10, bold=True, color="7C3AED")
        elif day["present"]:
            ws.cell(row=row, column=3).font = Font(name="Arial", size=10, bold=True, color="1B5E20")
        else:
            ws.cell(row=row, column=3).font = Font(name="Arial", size=10, bold=True, color="B71C1C")
        ws.row_dimensions[row].height = 16
        row += 1
    for col, w in zip("ABCDEF", [14,14,12,30,30,8]): ws.column_dimensions[col].width = w
    return wb

def _build_employee_pdf(report):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    except ImportError:
        return None
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=15*mm, bottomMargin=15*mm, leftMargin=15*mm, rightMargin=15*mm)
    styles = getSampleStyleSheet(); story = []
    title_style = ParagraphStyle("title", parent=styles["Title"], fontSize=14, textColor=colors.HexColor("#1F4E79"), spaceAfter=4)
    sub_style   = ParagraphStyle("sub",   parent=styles["Normal"], fontSize=10, textColor=colors.grey, spaceAfter=2)
    story.append(Paragraph("Attendance Report -- {0} ({1})".format(report['name'], report['badge']), title_style))
    story.append(Paragraph("Department: {0}  |  {1} -> {2}".format(report['dept'], report['from'], report['to']), sub_style))
    story.append(Paragraph("Present: <b>{0}</b> days  |  Absent: <b>{1}</b> days  |  Total: <b>{2}</b> days".format(
        report['present_days'], report['absent_days'], report['total_days']), sub_style))
    story.append(Spacer(1, 8*mm))
    hdr  = ["Date","Day","Status","Punch Times","Machines","#"]
    rows = [hdr]
    for day in report["days"]:
        punch_times = "\n".join(p["time"] for p in day["punches"]) or "--"
        machines    = "\n".join(dict.fromkeys(p["machine"] for p in day["punches"])) or "--"
        rows.append([day["date"], day["day"], "PRESENT" if day["present"] else "ABSENT", punch_times, machines, str(len(day["punches"]))])
    t = Table(rows, colWidths=[28*mm,28*mm,22*mm,60*mm,60*mm,12*mm])
    style = TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1F4E79")),("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),8),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),("ALIGN",(3,1),(4,-1),"LEFT"),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),("GRID",(0,0),(-1,-1),0.4,colors.HexColor("#CCCCCC")),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#F5F5F5")]),
        ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
    ])
    for i, day in enumerate(report["days"], 1):
        bg = colors.HexColor("#E8F5E9") if day["present"] else colors.HexColor("#FFEBEE")
        fc = colors.HexColor("#1B5E20") if day["present"] else colors.HexColor("#B71C1C")
        style.add("BACKGROUND",(0,i),(1,i),bg); style.add("BACKGROUND",(2,i),(2,i),bg)
        style.add("TEXTCOLOR",(2,i),(2,i),fc); style.add("FONTNAME",(2,i),(2,i),"Helvetica-Bold")
    t.setStyle(style); story.append(t); doc.build(story); buf.seek(0)
    return buf

@app.route("/api/employee/<badge>/report/export")
@permission_required("export_reports")
def employee_report_export(badge):
    from_str = request.args.get("from"); to_str = request.args.get("to")
    fmt      = request.args.get("format", "xlsx").lower()
    source   = request.args.get("source", "mdb")
    try:
        date_from = datetime.strptime(from_str, "%d/%m/%Y").date()
        date_to   = datetime.strptime(to_str,   "%d/%m/%Y").date()
    except:
        return jsonify({"error": "Invalid dates"}), 400
    # Get report data by calling the function directly within the active request context
    # (session is already validated by the @permission_required decorator above)
    resp = employee_report(badge)
    if isinstance(resp, tuple): return resp
    report = resp.get_json()
    if fmt == "pdf":
        buf = _build_employee_pdf(report)
        if buf is None: return jsonify({"error": "PDF requires reportlab: pip install reportlab"}), 500
        fname = "attendance_{0}_{1}_{2}.pdf".format(badge, date_from.strftime('%Y%m%d'), date_to.strftime('%Y%m%d'))
        return send_file(buf, download_name=fname, as_attachment=True, mimetype="application/pdf")
    else:
        wb  = _build_employee_xlsx(report)
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        fname = "attendance_{0}_{1}_{2}.xlsx".format(badge, date_from.strftime('%Y%m%d'), date_to.strftime('%Y%m%d'))
        return send_file(buf, download_name=fname, as_attachment=True,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ==============================================================================
#  SHIFT TIMES ROUTES
# ==============================================================================
@app.route("/api/shifts", methods=["GET"])
@login_required
def get_shifts():
    try:
        return jsonify(db_manager.get_shift_times())
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/shifts", methods=["POST"])
@admin_required
def save_shifts():
    data = request.get_json() or []
    try:
        db_manager.save_shift_times(data)
        global _shift_cache
        _shift_cache = {}  # clear cache so next request reloads
        db_manager.write_audit(session.get("username","?"), "UPDATE_SHIFTS",
                               "{0} depts".format(len(data)), request.remote_addr)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ==============================================================================
#  AUDIT LOG ROUTE
# ==============================================================================
@app.route("/api/audit")
@admin_required
def audit_log():
    limit = int(request.args.get("limit", 200))
    try:
        return jsonify(db_manager.get_audit_log(limit))
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ==============================================================================
#  EMPLOYEE ACTIVE TOGGLE
# ==============================================================================
@app.route("/api/employee/<badge>/active", methods=["POST"])
@admin_required
def toggle_employee_active(badge):
    data   = request.get_json() or {}
    active = bool(data.get("active", True))
    try:
        db_manager.set_employee_active(badge, active)
        db_manager.write_audit(session.get("username","?"), "SET_ACTIVE",
                               "{0} active={1}".format(badge, active), request.remote_addr)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ==============================================================================
#  AUTO-SYNC SCHEDULE
# ==============================================================================
_auto_sync_thread = None

def _auto_sync_loop():
    """Background thread: pull from devices at scheduled hour every day."""
    while True:
        try:
            sync_hour = int(db_manager.get_setting("auto_sync_hour", "6"))
            now = datetime.now()
            target = now.replace(hour=sync_hour, minute=0, second=0, microsecond=0)
            if target <= now:
                target += timedelta(days=1)
            wait_secs = (target - now).total_seconds()
            print("[AutoSync] Next sync at {0} (in {1:.0f}m)".format(
                target.strftime("%Y-%m-%d %H:%M"), wait_secs/60)); sys.stdout.flush()
            time.sleep(wait_secs)
            print("[AutoSync] Starting scheduled device pull..."); sys.stdout.flush()
            _refresh_cache()
            db_manager.write_audit("system", "AUTO_SYNC", "Scheduled sync at hour {0}".format(sync_hour))
        except Exception as e:
            print("[AutoSync] Error: {0}".format(e)); sys.stdout.flush()
            time.sleep(3600)

def start_auto_sync():
    try:
        enabled = db_manager.get_setting("auto_sync_enabled", "1")
        if str(enabled) == "1":
            t = threading.Thread(target=_auto_sync_loop, daemon=True); t.start()
            print("[AutoSync] Scheduled sync enabled"); sys.stdout.flush()
        else:
            print("[AutoSync] Disabled (change in Admin → Settings)"); sys.stdout.flush()
    except Exception as e:
        print("[AutoSync] Could not start: {0}".format(e)); sys.stdout.flush()

@app.route("/api/settings/auto-sync", methods=["GET"])
@admin_required
def get_auto_sync():
    return jsonify({
        "enabled": db_manager.get_setting("auto_sync_enabled","1") == "1",
        "hour":    int(db_manager.get_setting("auto_sync_hour","6")),
    })

@app.route("/api/settings/auto-sync", methods=["POST"])
@admin_required
def set_auto_sync():
    data = request.get_json() or {}
    db_manager.set_setting("auto_sync_enabled", "1" if data.get("enabled", True) else "0")
    db_manager.set_setting("auto_sync_hour", str(int(data.get("hour", 6))))
    db_manager.write_audit(session.get("username","?"), "UPDATE_AUTO_SYNC", str(data), request.remote_addr)
    return jsonify({"ok": True})

# ==============================================================================
#  SSE DEVICE POLL PROGRESS
# ==============================================================================
_sse_clients = []
_sse_lock    = threading.Lock()

def _sse_push(data_dict):
    msg = "data: {0}\n\n".format(json.dumps(data_dict))
    with _sse_lock:
        dead = []
        for q in _sse_clients:
            try: q.put_nowait(msg)
            except Exception: dead.append(q)
        for q in dead: _sse_clients.remove(q)

@app.route("/api/refresh/stream")
@login_required
def refresh_stream():
    q = queue.Queue(maxsize=50)
    with _sse_lock: _sse_clients.append(q)
    def generate():
        try:
            while True:
                try:
                    msg = q.get(timeout=30)
                    yield msg
                except queue.Empty:
                    yield "data: {}\n\n"  # keepalive
        finally:
            with _sse_lock:
                if q in _sse_clients: _sse_clients.remove(q)
    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control":"no-cache","X-Accel-Buffering":"no"})

# ==============================================================================
#  WHATSAPP / PLAIN-TEXT ABSENT EXPORT
# ==============================================================================
@app.route("/api/today/absent-text")
@permission_required("export_reports")
def export_today_text():
    with _cache_lock: data = _cache["today"]
    if not data: return jsonify({"error": "No data cached yet"}), 503
    dept_groups = {}
    for emp in data["absent"]:
        dept_groups.setdefault(emp["dept"], []).append(emp["name"])
    lines = ["*Absent Today — {0}*".format(data["date"]), ""]
    for dept in sorted(dept_groups.keys()):
        lines.append("*{0}* ({1})".format(dept, len(dept_groups[dept])))
        for name in sorted(dept_groups[dept]):
            lines.append("  • {0}".format(name))
        lines.append("")
    lines.append("Total: {0} absent  |  {1} present".format(data["absent_count"], data["present_count"]))
    return jsonify({"text": "\n".join(lines), "absent_count": data["absent_count"]})


# ==============================================================================
#  EMAIL REPORT  (Gmail SMTP with App Password)
# ==============================================================================
import smtplib
import ssl
from email.mime.multipart import MIMEMultipart
from email.mime.text      import MIMEText

def _get_email_cfg():
    """Load email config from settings table. Falls back to hardcoded defaults."""
    g = db_manager.get_setting
    return {
        "enabled":      g("email_enabled",      "1") == "1",
        "smtp_host":    g("email_smtp_host",     "smtp.gmail.com"),
        "smtp_port":    int(g("email_smtp_port", "465")),
        "sender":       g("email_sender",        "print@gulfasian.com"),
        "sender_name":  g("email_sender_name",   "ZKTeco Attendance"),
        "app_password": g("email_app_password",  ""),
        "recipients":   g("email_recipients",    "print@gulfasian.com"),
        "send_hour":    int(g("email_send_hour", "9")),
        "send_minute":  int(g("email_send_min",  "30")),
        "subject":      g("email_subject",       "Daily Absent Report — {date}"),
    }

def _build_email_html(data):
    """Build a clean HTML email body from today's absent data."""
    date_str = data.get("date", datetime.now().strftime("%d %B %Y"))
    absent   = data.get("absent", [])
    present_c = data.get("present_count", 0)
    absent_c  = data.get("absent_count",  0)
    total_c   = data.get("total",          0)

    # Group by department
    dept_groups = {}
    for emp in absent:
        dept_groups.setdefault(emp["dept"], []).append(emp["name"])

    dept_rows = ""
    for dept in sorted(dept_groups.keys()):
        names = dept_groups[dept]
        dept_rows += """
        <tr>
          <td colspan="2" style="background:#1F4E79;color:#fff;font-weight:700;
              padding:8px 14px;font-size:12px;letter-spacing:.5px">{dept} &nbsp;({n})</td>
        </tr>""".format(dept=html.escape(str(dept)), n=len(names))
        for i, name in enumerate(sorted(names)):
            bg = "#EBF3FB" if i % 2 == 0 else "#FFFFFF"
            dept_rows += """
        <tr>
          <td style="padding:7px 14px;font-size:13px;background:{bg}">{name}</td>
          <td style="padding:7px 14px;font-size:12px;color:#666;background:{bg}">{dept}</td>
        </tr>""".format(bg=bg, name=html.escape(str(name)), dept=html.escape(str(dept)))

    html = """<!DOCTYPE html>
<html><head><meta charset="UTF-8"></head>
<body style="font-family:Arial,sans-serif;background:#f0f4f8;margin:0;padding:20px">
  <div style="max-width:600px;margin:0 auto;background:#fff;border-radius:12px;
              overflow:hidden;box-shadow:0 4px 20px rgba(0,0,0,.1)">
    <!-- Header -->
    <div style="background:#1F4E79;padding:24px 28px">
      <div style="color:#fff;font-size:20px;font-weight:700">
        &#9889; Attendance Report
      </div>
      <div style="color:#a8c8e8;font-size:13px;margin-top:4px">{date}</div>
    </div>
    <!-- Stats -->
    <div style="display:flex;padding:20px 28px;gap:16px;background:#f8fafc;
                border-bottom:1px solid #e0e8f0">
      <div style="flex:1;text-align:center">
        <div style="font-size:28px;font-weight:800;color:#cc2200">{absent_c}</div>
        <div style="font-size:11px;color:#666;text-transform:uppercase;letter-spacing:.5px">Absent</div>
      </div>
      <div style="flex:1;text-align:center">
        <div style="font-size:28px;font-weight:800;color:#1a7a3c">{present_c}</div>
        <div style="font-size:11px;color:#666;text-transform:uppercase;letter-spacing:.5px">Present</div>
      </div>
      <div style="flex:1;text-align:center">
        <div style="font-size:28px;font-weight:800;color:#1F4E79">{total_c}</div>
        <div style="font-size:11px;color:#666;text-transform:uppercase;letter-spacing:.5px">Total</div>
      </div>
    </div>
    <!-- Absent table -->
    <div style="padding:20px 28px">
      <div style="font-size:13px;font-weight:700;color:#333;margin-bottom:12px">
        Absent Employees
      </div>
      {absent_section}
    </div>
    <!-- Footer -->
    <div style="background:#f0f4f8;padding:14px 28px;font-size:11px;color:#999;text-align:center">
      Sent automatically by ZKTeco Attendance Dashboard &bull; {date}
    </div>
  </div>
</body></html>""".format(
        date=date_str,
        absent_c=absent_c, present_c=present_c, total_c=total_c,
        absent_section=(
            """<table width="100%" cellpadding="0" cellspacing="0" style="border-collapse:collapse;
               border:1px solid #e0e8f0;border-radius:8px;overflow:hidden">{rows}</table>""".format(rows=dept_rows)
            if dept_rows else
            '<div style="color:#1a7a3c;font-size:14px;text-align:center;padding:20px">&#10003; No absences today!</div>'
        )
    )
    return html

def _build_email_text(data):
    """Plain-text fallback for email."""
    date_str  = data.get("date", datetime.now().strftime("%d %B %Y"))
    absent    = data.get("absent", [])
    present_c = data.get("present_count", 0)
    absent_c  = data.get("absent_count",  0)

    dept_groups = {}
    for emp in absent:
        dept_groups.setdefault(emp["dept"], []).append(emp["name"])

    lines = [
        "Attendance Report — {0}".format(date_str),
        "Absent: {0}  |  Present: {1}".format(absent_c, present_c),
        "",
    ]
    for dept in sorted(dept_groups.keys()):
        lines.append("{0} ({1})".format(dept, len(dept_groups[dept])))
        for name in sorted(dept_groups[dept]):
            lines.append("  - {0}".format(name))
        lines.append("")
    if not dept_groups:
        lines.append("No absences today!")
    return "\n".join(lines)

def send_email_report(data=None, test_recipient=None):
    """
    Send the daily absent report via Gmail SMTP.
    data: cache dict (uses live cache if None)
    test_recipient: if set, send only to this address (for test button)
    Returns (ok:bool, message:str)
    """
    cfg = _get_email_cfg()
    if not cfg["sender"] or not cfg["app_password"]:
        return False, "Email not configured. Set sender and App Password in Admin → Email Settings."

    recipients = [r.strip() for r in (test_recipient or cfg["recipients"]).split(",") if r.strip()]
    if not recipients:
        return False, "No recipients configured."

    # Get attendance data
    if data is None:
        with _cache_lock:
            data = _cache.get("today")
    if not data:
        return False, "No attendance data available yet. Run a refresh first."

    date_str = data.get("date", datetime.now().strftime("%d %B %Y"))
    subject  = cfg["subject"].replace("{date}", date_str)

    # Build message
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"]    = "{0} <{1}>".format(cfg.get("sender_name") or "ZKTeco Attendance", cfg["sender"])
    msg["To"]      = ", ".join(recipients)
    msg.attach(MIMEText(_build_email_text(data), "plain"))
    msg.attach(MIMEText(_build_email_html(data), "html"))

    # Send
 # Send
    try:
        context = ssl.create_default_context()
        with smtplib.SMTP_SSL(cfg["smtp_host"], cfg["smtp_port"], context=context) as server:
            server.login(cfg["sender"], cfg["app_password"])
            server.sendmail(cfg["sender"], recipients, msg.as_string())
        print("[Email] Sent to {0}: {1}".format(recipients, subject)); sys.stdout.flush()
        db_manager.write_audit("system", "EMAIL_SENT",
                               "To: {0} | {1}".format(",".join(recipients), subject))
        return True, "Email sent to {0}".format(", ".join(recipients))
    except smtplib.SMTPAuthenticationError:
        return False, "Authentication failed. Check your Gmail address and App Password."
    except smtplib.SMTPException as e:
        return False, "SMTP error: {0}".format(str(e))
    except Exception as e:
        return False, "Error: {0}".format(str(e))

def _get_absent_data_for_date(target_date):
    """
    Fetch absent/present data for a specific date from the SQLite punches table.
    target_date: a datetime.date object
    Returns a data dict in the same format as _cache["today"].
    """
    try:
        emp_df = load_employees()
    except Exception as e:
        raise RuntimeError("Could not load employees: {0}".format(e))

    # Load punches from SQLite for the target date
    start_s = target_date.strftime("%Y-%m-%d") + " 00:00:00"
    end_s   = target_date.strftime("%Y-%m-%d") + " 23:59:59"
    try:
        db_punch = get_db()
        punch_rows = db_punch.execute(
            "SELECT DISTINCT badge FROM punches WHERE punch_time >= ? AND punch_time <= ?",
            (start_s, end_s)
        ).fetchall()
        db_punch.close()
        present_badges = {str(r["badge"]).strip() for r in punch_rows if r["badge"]}
    except Exception as e:
        raise RuntimeError("Could not fetch punch data: {0}".format(e))

    absent, present = [], []
    for _, emp in emp_df.iterrows():
        badge = emp["Badgenumber"]; dept = emp["DEPTNAME"]
        if not _is_working_day(target_date, badge, dept):
            continue
        rec = {"name": emp["Name"], "code": badge, "dept": dept}
        (present if badge in present_badges else absent).append(rec)

    absent.sort(key=lambda x: (_dept_sort(x["dept"]), x["name"]))
    present.sort(key=lambda x: (_dept_sort(x["dept"]), x["name"]))

    return {
        "date":          target_date.strftime("%d %B %Y"),
        "total":         len(present) + len(absent),
        "present_count": len(present),
        "absent_count":  len(absent),
        "present":       present,
        "absent":        absent,
        "devices":       [],
    }

# ── Email scheduler ────────────────────────────────────────────────────────────
def _email_scheduler_loop():
    """Background thread: send email at configured time every day."""
    print("[Email] Scheduler started"); sys.stdout.flush()
    while True:
        try:
            cfg = _get_email_cfg()
            if not cfg["enabled"]:
                time.sleep(300); continue   # check again in 5 min

            now    = datetime.now()
            target = now.replace(hour=cfg["send_hour"], minute=cfg["send_minute"],
                                 second=0, microsecond=0)
            if target <= now:
                target += timedelta(days=1)
            wait = (target - now).total_seconds()
            print("[Email] Next send at {0} (in {1:.0f}m)".format(
                target.strftime("%H:%M"), wait/60)); sys.stdout.flush()
            time.sleep(wait)

            # Send
            ok, msg = send_email_report()
            if ok:
                print("[Email] Scheduled send OK: {0}".format(msg)); sys.stdout.flush()
            else:
                print("[Email] Scheduled send FAILED: {0}".format(msg)); sys.stdout.flush()
        except Exception as e:
            print("[Email] Scheduler error: {0}".format(e)); sys.stdout.flush()
            time.sleep(3600)

def start_email_scheduler():
    t = threading.Thread(target=_email_scheduler_loop, daemon=True)
    t.start()

# ── Telegram scheduler (daily absent report at configured time) ───────────────
def _telegram_scheduler_loop():
    """Background thread: send Telegram absent report at configured time every day."""
    print("[Telegram] Daily report scheduler started"); sys.stdout.flush()
    while True:
        try:
            if not _tg_notifier or not _tg_notifier.notify_daily_report:
                time.sleep(300); continue

            rpt_hour   = int(db_manager.get_setting('tg_daily_report_hour',   '') or _cfg_int('telegram', 'daily_report_hour',   8))
            rpt_minute = int(db_manager.get_setting('tg_daily_report_minute', '') or _cfg_int('telegram', 'daily_report_minute', 10))

            now    = datetime.now()
            target = now.replace(hour=rpt_hour, minute=rpt_minute, second=0, microsecond=0)
            if target <= now:
                target += timedelta(days=1)
            wait = (target - now).total_seconds()
            print("[Telegram] Next daily report at {0} (in {1:.0f}m)".format(
                target.strftime("%H:%M"), wait / 60)); sys.stdout.flush()
            time.sleep(wait)

            # Fetch absent data fresh from DB (today so far)
            try:
                report_data = _get_absent_data_for_date(date.today())
            except Exception as exc:
                print("[Telegram] Could not build report data: {0}".format(exc)); sys.stdout.flush()
                time.sleep(3600); continue

            absent       = report_data.get("absent", [])
            present_count = report_data.get("present_count", 0)
            total         = report_data.get("total", 0)
            date_str      = report_data.get("date", datetime.now().strftime("%d %B %Y"))

            ok = _tg_notifier.send_daily_absent_report(
                absent=absent,
                present_count=present_count,
                total=total,
                date_str=date_str,
                dept_order=DEPT_ORDER,
            )
            if ok:
                print("[Telegram] Daily report sent OK"); sys.stdout.flush()
            else:
                print("[Telegram] Daily report send failed"); sys.stdout.flush()

        except Exception as e:
            print("[Telegram] Scheduler error: {0}".format(e)); sys.stdout.flush()
            time.sleep(3600)


def start_telegram_scheduler():
    t = threading.Thread(target=_telegram_scheduler_loop, daemon=True)
    t.start()

# ── Notice Board (Announcements) ────────────────────────────────────────────────
@app.route("/api/announcements", methods=["GET"])
@login_required
def get_announcements():
    try:
        conn = get_db()
        rows = conn.execute("SELECT * FROM announcements WHERE active=1 ORDER BY created_at DESC").fetchall()
        conn.close()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/announcements", methods=["POST"])
@admin_required
def post_announcement():
    data = request.json or {}
    message = data.get("message", "").strip()
    if not message: return jsonify({"error": "Empty message"}), 400
    try:
        conn = get_db()
        conn.execute("INSERT INTO announcements (message, active, created_at, created_by) VALUES (?, 1, ?, ?)",
                     (message, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), session.get("username", "admin")))
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/announcements/<int:ann_id>", methods=["DELETE"])
@admin_required
def delete_announcement(ann_id):
    try:
        conn = get_db()
        conn.execute("UPDATE announcements SET active=0 WHERE id=?", (ann_id,))
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── Internal Messaging ──────────────────────────────────────────────────────────
@app.route("/api/messages", methods=["GET"])
@login_required
def get_messages():
    username = session.get("username")
    # Also match on badge so the frontend can use either identity
    users = _get_users()
    badge = (users.get(username) or {}).get("badge") or username
    try:
        conn = get_db()
        rows = conn.execute(
            "SELECT * FROM messages WHERE receiver IN (?,?) OR sender IN (?,?) ORDER BY timestamp ASC",
            (username, badge, username, badge)
        ).fetchall()
        conn.close()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/users/online", methods=["GET"])
@login_required
def get_online_users():
    now = datetime.now()
    online = set()
    with _active_sessions_lock:
        for sid, data in list(_active_sessions.items()):
            try:
                la = datetime.fromisoformat(data['last_active'])
                if (now - la).total_seconds() < 300:
                    online.add(data['username'])
            except Exception:
                pass
    return jsonify(list(online))

@app.route("/api/messages", methods=["POST"])
@login_required
def send_message():
    data = request.json or {}
    receiver = data.get("receiver", "").strip()
    msg = data.get("message", "").strip()
    if not receiver or not msg:
        return jsonify({"error": "Missing receiver or message"}), 400

    sender = session.get("username")
    # Normalise receiver: if a badge was passed, resolve to username
    users = _get_users()
    badge_to_user = {(v.get("badge") or k): k for k, v in users.items()}
    receiver_resolved = badge_to_user.get(receiver, receiver)

    try:
        conn = get_db()
        conn.execute(
            "INSERT INTO messages (sender, receiver, message, timestamp) VALUES (?, ?, ?, ?)",
            (sender, receiver_resolved, msg, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        )
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/messages/<int:msg_id>/read", methods=["POST"])
@login_required
def mark_message_read(msg_id):
    username = session.get("username")
    users = _get_users()
    badge = (users.get(username) or {}).get("badge") or username
    try:
        conn = get_db()
        conn.execute(
            "UPDATE messages SET is_read=1 WHERE id=? AND receiver IN (?,?)",
            (msg_id, username, badge)
        )
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/messages/<int:msg_id>", methods=["DELETE"])
@login_required
def delete_message(msg_id):
    username = session.get("username")
    users = _get_users()
    badge = (users.get(username) or {}).get("badge") or username
    try:
        conn = get_db()
        conn.execute(
            "DELETE FROM messages WHERE id=? AND (sender IN (?,?) OR receiver IN (?,?))",
            (msg_id, username, badge, username, badge)
        )
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── Personal Notes ──────────────────────────────────────────────────────────────
@app.route("/api/notes", methods=["GET"])
@login_required
def get_notes():
    user = session.get("username")
    try:
        conn = get_db()
        rows = conn.execute("SELECT * FROM notes WHERE user_badge=? ORDER BY timestamp DESC", (user,)).fetchall()
        conn.close()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/notes", methods=["POST"])
@login_required
def save_note():
    data = request.json or {}
    text = data.get("text", "").strip()
    title = data.get("title", "").strip()
    color = data.get("color", "purple").strip()
    if not text: return jsonify({"error": "Empty note"}), 400
    user = session.get("username")
    try:
        conn = get_db()
        # Migrate schema if columns don't exist (wrapped per-column to avoid race on concurrent requests)
        try:
            conn.execute("ALTER TABLE notes ADD COLUMN title TEXT DEFAULT ''")
        except Exception:
            pass  # column already exists
        try:
            conn.execute("ALTER TABLE notes ADD COLUMN color TEXT DEFAULT 'purple'")
        except Exception:
            pass  # column already exists
        conn.execute("INSERT INTO notes (user_badge, text, title, color, timestamp) VALUES (?, ?, ?, ?, ?)",
                     (user, text, title, color, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/notes/<int:note_id>", methods=["DELETE"])
@login_required
def delete_note(note_id):
    user = session.get("username")
    try:
        conn = get_db()
        conn.execute("DELETE FROM notes WHERE id=? AND user_badge=?", (note_id, user))
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── Email API routes ───────────────────────────────────────────────────────────
@app.route("/api/settings/email", methods=["GET"])
@admin_required
def get_email_settings():
    cfg = _get_email_cfg()
    # Never return the actual app password — just whether it's set
    cfg["app_password"] = "••••••••" if cfg["app_password"] else ""
    cfg["password_set"] = db_manager.get_setting("email_app_password","") != ""
    return jsonify(cfg)

@app.route("/api/settings/email", methods=["POST"])
@admin_required
def save_email_settings():
    data = request.get_json() or {}
    s    = db_manager.set_setting
    s("email_enabled",      "1" if data.get("enabled") else "0")
    s("email_smtp_host",    data.get("smtp_host",    "smtp.gmail.com"))
    s("email_smtp_port",    str(data.get("smtp_port", 465)))
    s("email_sender",       data.get("sender",       ""))
    s("email_sender_name",  data.get("sender_name",  ""))
    s("email_recipients",   data.get("recipients",   ""))
    s("email_send_hour",    str(data.get("send_hour",   9)))
    s("email_send_min",     str(data.get("send_minute", 30)))
    s("email_subject",      data.get("subject", "Daily Absent Report — {date}"))
    # Only update password if a new one was provided (not the masked placeholder)
    if data.get("app_password") and "•" not in data["app_password"]:
        s("email_app_password", data["app_password"])
    db_manager.write_audit(session.get("username","?"), "UPDATE_EMAIL_SETTINGS",
                           "enabled={0}".format(data.get("enabled")), request.remote_addr)
    return jsonify({"ok": True})

# ── Telegram settings API ──────────────────────────────────────────────────────
@app.route("/api/settings/telegram", methods=["GET"])
@admin_required
def get_telegram_settings():
    s = db_manager.get_setting
    token = s("tg_bot_token", "") or _cfg("telegram", "bot_token", "")
    chat  = s("tg_chat_id",   "") or _cfg("telegram", "chat_id",   "")
    return jsonify({
        "enabled":              s("tg_enabled",              "1") == "1",
        "bot_token":            "••••" + token[-4:] if len(token) > 4 else ("••••" if token else ""),
        "token_set":            bool(token),
        "chat_id":              chat,
        "notify_device_status": s("tg_notify_device_status", "1") == "1",
        "notify_punches":       s("tg_notify_punches",       "1") == "1",
        "notify_daily_report":  s("tg_notify_daily_report",  "1") == "1",
        "daily_report_hour":    int(s("tg_daily_report_hour",   "8")),
        "daily_report_minute":  int(s("tg_daily_report_minute", "10")),
    })

@app.route("/api/settings/telegram", methods=["POST"])
@admin_required
def save_telegram_settings():
    global _tg_notifier
    data = request.get_json() or {}
    s = db_manager.set_setting
    s("tg_enabled",              "1" if data.get("enabled", True) else "0")
    s("tg_chat_id",              str(data.get("chat_id", "")).strip())
    s("tg_notify_device_status", "1" if data.get("notify_device_status", True) else "0")
    s("tg_notify_punches",       "1" if data.get("notify_punches",       True) else "0")
    s("tg_notify_daily_report",  "1" if data.get("notify_daily_report",  True) else "0")
    s("tg_daily_report_hour",    str(int(data.get("daily_report_hour",   8))))
    s("tg_daily_report_minute",  str(int(data.get("daily_report_minute", 10))))
    # Only update token if a new one was provided (not the masked placeholder)
    new_token = str(data.get("bot_token", "")).strip()
    if new_token and "•" not in new_token:
        s("tg_bot_token", new_token)
    db_manager.write_audit(session.get("username","?"), "UPDATE_TELEGRAM_SETTINGS",
                           "enabled={0}".format(data.get("enabled")), request.remote_addr)
    # Re-init notifier with new settings
    _init_telegram()
    resp = {"ok": True}
    if _tg_notifier is None:
        resp["warning"] = _not_configured_msg()
    return jsonify(resp)

def _ensure_telegram_initialized():
    """Re-initialize Telegram notifier if not active (e.g. after server restart)."""
    global _tg_notifier
    if not _tg_notifier:
        _init_telegram()
    return _tg_notifier

def _not_configured_msg():
    """Human-readable reason why Telegram is not configured."""
    if _tg_init_error:
        return "Telegram not configured: {0}".format(_tg_init_error)
    return "Telegram not configured. Set bot_token and chat_id first."

@app.route("/api/settings/telegram/test", methods=["POST"])
@admin_required
def test_telegram():
    notifier = _ensure_telegram_initialized()
    if not notifier:
        return jsonify({"ok": False, "message": _not_configured_msg()})
    # Distinguish "disabled" from actual send failure
    if not notifier.enabled:
        return jsonify({"ok": False, "message": "Telegram notifications are disabled. Enable them first."})
    ok = notifier.test_connection()
    return jsonify({"ok": ok, "message": "Test message sent!" if ok else "Send failed — check token and chat_id."})

@app.route("/api/settings/telegram/test-report", methods=["POST"])
@admin_required
def test_telegram_report():
    """Send a test daily report via Telegram (using today's data)."""
    notifier = _ensure_telegram_initialized()
    if not notifier:
        return jsonify({"ok": False, "message": _not_configured_msg()})
    try:
        report_data = _get_absent_data_for_date(date.today())
    except Exception as exc:
        print("[Telegram] test-report error: {0}".format(exc)); sys.stdout.flush()
        return jsonify({"ok": False, "message": "Could not build report data. Check server logs."}), 500
    ok = notifier.send_daily_absent_report(
        absent=report_data.get("absent", []),
        present_count=report_data.get("present_count", 0),
        total=report_data.get("total", 0),
        date_str=report_data.get("date", ""),
        dept_order=DEPT_ORDER,
    )
    return jsonify({"ok": ok, "message": "Report sent!" if ok else "Send failed."})


@app.route("/api/settings/email/test", methods=["POST"])
@admin_required
def test_email():
    req  = request.get_json() or {}
    recipient  = req.get("recipient", "").strip()
    date_param = req.get("date", "").strip()   # optional "dd/mm/yyyy"

    report_data = None

    # If a specific date was requested, fetch it from MDB
    if date_param:
        try:
            target_date = datetime.strptime(date_param, "%d/%m/%Y").date()
        except ValueError:
            return jsonify({"ok": False, "message": "Invalid date format — expected dd/mm/yyyy"}), 400
        try:
            report_data = _get_absent_data_for_date(target_date)
        except Exception as e:
            return jsonify({"ok": False, "message": str(e)}), 500
    # Otherwise fall back to live cache (today)
    else:
        with _cache_lock:
            report_data = _cache.get("today")

    try:
        ok, msg = send_email_report(data=report_data, test_recipient=recipient or None)
    except Exception as e:
        import traceback
        traceback.print_exc()
        ok, msg = False, "Error: " + str(e)

    db_manager.write_audit(
        session.get("username", "?"), "EMAIL_TEST",
        "date={0} result={1}".format(date_param or "today", msg),
        request.remote_addr
    )
    if ok:
        return jsonify({"ok": True,  "message": msg})
    else:
        return jsonify({"error": msg}), 500

# ==============================================================================
#  SESSION MANAGEMENT ROUTES
# ==============================================================================
@app.route("/api/sessions")
@admin_required
def list_sessions():
    with _active_sessions_lock:
        sessions = list(_active_sessions.values())
    return jsonify(sessions)

@app.route("/api/sessions/<sid>", methods=["DELETE"])
@admin_required
def kill_session(sid):
    username = None
    with _active_sessions_lock:
        s = _active_sessions.pop(sid, None)
        if s: username = s.get('username')
    if not username:
        return jsonify({"error": "Session not found"}), 404
    db_manager.write_audit(session.get("username","?"), "KILL_SESSION",
                           "Killed session for {0}".format(username), request.remote_addr)
    return jsonify({"ok": True, "killed": username})

@app.route("/api/sessions/user/<username>", methods=["DELETE"])
@admin_required
def kill_user_sessions(username):
    n = _kill_user_sessions(username)
    db_manager.write_audit(session.get("username","?"), "KILL_ALL_SESSIONS",
                           "Killed {0} sessions for {1}".format(n, username), request.remote_addr)
    return jsonify({"ok": True, "killed": n})

@app.route("/api/auth/inactivity-config")
@admin_required
def get_inactivity_config():
    return jsonify({
        "admin_mins":    ADMIN_INACTIVITY_MINS,
        "employee_mins": EMPLOYEE_INACTIVITY_MINS,
    })

# ==============================================================================
#  WORKDAY CONFIGURATION API
# ==============================================================================
@app.route("/api/workdays/dept", methods=["GET"])
@login_required
def get_workdays_dept():
    """Return workday config for all departments."""
    try:
        db_cfg = get_dept_workdays()
        # Merge with fallback so all known depts are present
        result = {}
        all_depts = set(list(DEPT_WORKDAYS_FALLBACK.keys()) + list(db_cfg.keys()))
        for dept in sorted(all_depts):
            if dept in db_cfg:
                result[dept] = db_cfg[dept]
            else:
                result[dept] = sorted(DEPT_WORKDAYS_FALLBACK.get(dept, DEFAULT_WORKDAYS))
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/workdays/dept", methods=["POST"])
@admin_required
def save_workdays_dept():
    """Save workday config for departments. Body: {dept: [weekday_ints], ...}
    Also accepts legacy wrapped body {workdays: {dept: [weekday_ints], ...}}.
    """
    data = request.get_json() or {}
    # Unwrap legacy format sent as {workdays: {...}}
    if "workdays" in data and isinstance(data.get("workdays"), dict) and len(data) == 1:
        data = data["workdays"]
    try:
        saved = 0
        for dept, days in data.items():
            if isinstance(days, list) and all(isinstance(d, int) and 0 <= d <= 6 for d in days):
                save_dept_workday(dept, days)
                saved += 1
        _load_workday_caches()   # Refresh in-memory cache
        db_manager.write_audit(session.get("username","?"), "UPDATE_DEPT_WORKDAYS",
                               "{0} depts updated".format(saved), request.remote_addr)
        return jsonify({"ok": True, "saved": saved})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/workdays/employee", methods=["GET"])
@login_required
def get_workdays_employee():
    """Return all employee-specific workday overrides."""
    try:
        return jsonify(get_all_emp_workdays())
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/workdays/employee", methods=["POST"])
@admin_required
def save_workdays_employee():
    """Save employee-specific workday override. Body: {badge: [weekday_ints] or null, ...}"""
    data = request.get_json() or {}
    try:
        for badge, days in data.items():
            if days is None:
                save_emp_workday(badge, None)   # Remove override
            elif isinstance(days, list) and all(isinstance(d, int) and 0 <= d <= 6 for d in days):
                save_emp_workday(badge, days)
        _load_workday_caches()
        db_manager.write_audit(session.get("username","?"), "UPDATE_EMP_WORKDAYS",
                               "{0} employees updated".format(len(data)), request.remote_addr)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ==============================================================================
#  HOLIDAYS API
# ==============================================================================
@app.route("/api/holidays", methods=["GET"])
@login_required
def api_get_holidays():
    year = request.args.get("year")
    try:
        return jsonify(get_holidays(year=year))
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/holidays", methods=["POST"])
@admin_required
def api_add_holiday():
    data = request.get_json() or {}
    date_val = (data.get("date") or "").strip()
    date_end = (data.get("date_end") or date_val).strip()
    label    = (data.get("label") or "Holiday").strip()
    scope    = (data.get("scope") or "all").strip()
    dept     = (data.get("dept") or "").strip()
    employees = (data.get("employees") or "").strip()
    if not date_val:
        return jsonify({"error": "date is required"}), 400
    if not label:
        return jsonify({"error": "label is required"}), 400
    if scope not in ("all", "dept", "employee"):
        return jsonify({"error": "scope must be all, dept, or employee"}), 400
    try:
        hid = add_holiday(date_val, date_end, label, scope, dept, employees)
        db_manager.write_audit(session.get("username","?"), "ADD_HOLIDAY",
                               "{0} on {1} scope={2}".format(label, date_val, scope),
                               request.remote_addr)
        return jsonify({"ok": True, "id": hid})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/holidays/<int:hid>", methods=["DELETE"])
@admin_required
def api_delete_holiday(hid):
    try:
        delete_holiday(hid)
        db_manager.write_audit(session.get("username","?"), "DELETE_HOLIDAY",
                               "Holiday id={0}".format(hid), request.remote_addr)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ==============================================================================
#  DEVICE DATA DOWNLOAD  (users, fingerprint templates, face data)
# ==============================================================================
@app.route("/api/device/<ip>/download/users")
@admin_required
def device_download_users(ip):
    """Download user list from device as JSON."""
    if ip not in DEVICE_IPS: return jsonify({"error": "Unknown device"}), 400
    zk_conn = None
    try:
        from zk import ZK
        zk_conn = ZK(ip, port=DEVICE_PORT, timeout=30, verbose=False).connect()
        users = zk_conn.get_users()
        user_list = [{"uid": str(u.user_id), "name": u.name or "", "privilege": u.privilege,
                      "card": u.card or "", "group_id": u.group_id or ""} for u in users]
        db_manager.write_audit(session.get("username","?"), "DOWNLOAD_USERS",
                               "Device {0}: {1} users".format(ip, len(user_list)), request.remote_addr)
        buf = io.BytesIO(json.dumps(user_list, indent=2).encode())
        fname = "users_{0}_{1}.json".format(ip.replace(".","_"), date.today().strftime("%Y%m%d"))
        return send_file(buf, download_name=fname, as_attachment=True, mimetype="application/json")
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if zk_conn:
            try: zk_conn.disconnect()
            except: pass

@app.route("/api/device/<ip>/download/fingerprints")
@admin_required
def device_download_fingerprints(ip):
    """Download fingerprint templates from device as JSON (base64-encoded raw data)."""
    if ip not in DEVICE_IPS: return jsonify({"error": "Unknown device"}), 400
    zk_conn = None
    try:
        import base64
        from zk import ZK
        zk_conn = ZK(ip, port=DEVICE_PORT, timeout=60, verbose=False).connect()
        templates = []
        try:
            fps = zk_conn.get_templates() if hasattr(zk_conn, "get_templates") else []
        except Exception:
            fps = []
        for fp in fps:
            templates.append({
                "uid":   str(fp.uid),
                "fid":   fp.fid,
                "valid": fp.valid,
                "data":  base64.b64encode(fp.template).decode() if fp.template else ""
            })
        db_manager.write_audit(session.get("username","?"), "DOWNLOAD_FINGERPRINTS",
                               "Device {0}: {1} templates".format(ip, len(templates)), request.remote_addr)
        buf = io.BytesIO(json.dumps(templates, indent=2).encode())
        fname = "fingerprints_{0}_{1}.json".format(ip.replace(".","_"), date.today().strftime("%Y%m%d"))
        return send_file(buf, download_name=fname, as_attachment=True, mimetype="application/json")
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if zk_conn:
            try: zk_conn.disconnect()
            except: pass

@app.route("/api/device/<ip>/download/attendance-raw")
@admin_required
def device_download_attendance_raw(ip):
    """Download full attendance log from device as CSV."""
    if ip not in DEVICE_IPS: return jsonify({"error": "Unknown device"}), 400
    zk_conn = None
    try:
        from zk import ZK
        zk_conn = ZK(ip, port=DEVICE_PORT, timeout=60, verbose=False).connect()
        zk_conn.disable_device()
        records = zk_conn.get_attendance()
        zk_conn.enable_device()
        lines = ["uid,timestamp,punch_type,status"]
        for r in records:
            lines.append("{0},{1},{2},{3}".format(
                r.user_id, r.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
                r.punch, r.status
            ))
        csv_data = "\n".join(lines)
        db_manager.write_audit(session.get("username","?"), "DOWNLOAD_ATTENDANCE_RAW",
                               "Device {0}: {1} records".format(ip, len(records)), request.remote_addr)
        buf = io.BytesIO(csv_data.encode())
        fname = "attendance_raw_{0}_{1}.csv".format(ip.replace(".","_"), date.today().strftime("%Y%m%d"))
        return send_file(buf, download_name=fname, as_attachment=True, mimetype="text/csv")
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if zk_conn:
            try: zk_conn.enable_device(); zk_conn.disconnect()
            except: pass

# ==============================================================================
#  PWA ICON ROUTES  (generated dynamically, no static PNG files needed)
# ==============================================================================
def _make_png_icon(size):
    """Generate a minimal valid PNG icon (colored square with lightning bolt)."""
    import struct
    import zlib
    W = H = size
    # Background color #0a0e1a, accent #00d4ff
    bg = (10, 14, 26)
    ac = (0, 212, 255)
    # Build RGBA pixel data
    rows = []
    for y in range(H):
        row = b"\x00"   # filter type none
        for x in range(W):
            # Simple lightning bolt shape
            cx = x / W; cy = y / H
            # Outer rounded rect (corner radius ~20%)
            margin = 0.1
            in_rect = cx > margin and cx < 1-margin and cy > margin and cy < 1-margin
            # Lightning bolt path approximation
            bx = (cx - 0.25) * 2; by = (cy - 0.1) * 1.2
            top_tri   = (by < bx * 1.3 + 0.5) and (bx > -0.5) and (by > 0) and (by < 0.55)
            bot_tri   = (by > bx * 1.3 + 0.3) and (bx < 0.5) and (by < 1.1) and (by > 0.45)
            is_bolt   = (top_tri or bot_tri) and in_rect
            if is_bolt:
                r, g, b, a = ac[0], ac[1], ac[2], 255
            elif in_rect:
                r, g, b, a = bg[0], bg[1], bg[2], 255
            else:
                r, g, b, a = 0, 0, 0, 0
            row += bytes([r, g, b, a])
        rows.append(row)
    raw = b"".join(rows)
    def chunk(name, data):
        c = struct.pack(">I", len(data)) + name + data
        return c + struct.pack(">I", zlib.crc32(name + data) & 0xffffffff)
    ihdr = struct.pack(">IIBBBBB", W, H, 8, 6, 0, 0, 0)
    idat = zlib.compress(raw)
    png  = b"\x89PNG\r\n\x1a\n" + chunk(b"IHDR", ihdr) + chunk(b"IDAT", idat) + chunk(b"IEND", b"")
    return png

@app.route("/static/icon-<int:size>.png")
def serve_icon(size):
    allowed = [192, 512]
    sz = size if size in allowed else 192
    data = _make_png_icon(sz)
    resp = Response(data, mimetype="image/png")
    resp.headers["Cache-Control"] = "public, max-age=86400"
    return resp

# ==============================================================================
#  DEVICE ADVANCED ROUTES  (clock sync, reboot, user sync)
# ==============================================================================
@app.route("/api/device/<ip>/sync-clock", methods=["POST"])
@admin_required
def device_sync_clock(ip):
    if ip not in DEVICE_IPS: return jsonify({"error": "Unknown device"}), 400
    zk_conn = None
    try:
        from zk import ZK
        zk_conn = ZK(ip, port=DEVICE_PORT, timeout=30, verbose=False).connect()
        now = datetime.now()
        zk_conn.set_time(now)
        db_manager.write_audit(session.get("username","?"), "CLOCK_SYNC",
                               "Synced {0} to {1}".format(ip, now.strftime("%Y-%m-%d %H:%M:%S")),
                               request.remote_addr)
        return jsonify({"ok": True, "ip": ip,
                        "synced_to": now.strftime("%Y-%m-%d %H:%M:%S")})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if zk_conn:
            try: zk_conn.disconnect()
            except: pass

@app.route("/api/devices/sync-all-clocks", methods=["POST"])
@admin_required
def devices_sync_all_clocks():
    results = []
    import concurrent.futures
    
    def sync_clock(ip):
        zk_conn = None
        try:
            from zk import ZK
            zk_conn = ZK(ip, port=DEVICE_PORT, timeout=15, verbose=False).connect()
            now = datetime.now()
            zk_conn.set_time(now)
            return {"ip": ip, "ok": True, "synced_to": now.strftime("%Y-%m-%d %H:%M:%S")}
        except Exception as e:
            return {"ip": ip, "ok": False, "error": str(e)}
        finally:
            if zk_conn:
                try: zk_conn.disconnect()
                except: pass

    with concurrent.futures.ThreadPoolExecutor(max_workers=min(10, len(DEVICE_IPS) or 1)) as executor:
        futures = [executor.submit(sync_clock, ip) for ip in DEVICE_IPS]
        for future in concurrent.futures.as_completed(futures):
            results.append(future.result())
    ok_count = sum(1 for r in results if r["ok"])
    db_manager.write_audit(session.get("username","?"), "CLOCK_SYNC_ALL",
                           "{0}/{1} devices synced".format(ok_count, len(DEVICE_IPS)),
                           request.remote_addr)
    return jsonify({"results": results, "synced": ok_count, "total": len(DEVICE_IPS)})

@app.route("/api/device/<ip>/reboot", methods=["POST"])
@admin_required
def device_reboot(ip):
    if ip not in DEVICE_IPS: return jsonify({"error": "Unknown device"}), 400
    zk_conn = None
    try:
        from zk import ZK
        zk_conn = ZK(ip, port=DEVICE_PORT, timeout=15, verbose=False).connect()
        zk_conn.restart()
        db_manager.write_audit(session.get("username","?"), "DEVICE_REBOOT",
                               "Rebooted {0}".format(ip), request.remote_addr)
        return jsonify({"ok": True, "ip": ip,
                        "message": "Reboot command sent. Device will be offline for ~30 seconds."})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if zk_conn:
            try: zk_conn.disconnect()
            except: pass

@app.route("/api/devices/sync-users", methods=["POST"])
@permission_required("devices", "write")
def devices_sync_users():
    """
    Sync all users from every device to every other device.
    Strategy:
      1. Pull user list from all online devices
      2. Build a unified user set (union of all users)
      3. Push any missing users to each device
    Returns a detailed result per device.
    """
    data = request.get_json() or {}
    source_ip = data.get("source_ip")   # if set, only push FROM this device

    # ── Step 1: Collect users from all devices ────────────────────────────────
    all_device_users = {}   # ip -> list of ZK user objects serialized
    online_ips = []
    import concurrent.futures

    def fetch_users(ip):
        zk_conn = None
        try:
            from zk import ZK
            zk_conn = ZK(ip, port=DEVICE_PORT, timeout=15, verbose=False).connect()
            users = zk_conn.get_users()
            return ip, [
                {"uid": str(u.user_id), "name": u.name or "",
                 "privilege": u.privilege, "password": u.password or "",
                 "group_id": u.group_id or ""}
                for u in users
            ]
        except Exception as e:
            print("[UserSync] {0} fetch failed: {1}".format(ip, e))
            return ip, None
        finally:
            if zk_conn:
                try: zk_conn.disconnect()
                except: pass

    with concurrent.futures.ThreadPoolExecutor(max_workers=min(10, len(DEVICE_IPS) or 1)) as executor:
        futures = [executor.submit(fetch_users, ip) for ip in DEVICE_IPS]
        for future in concurrent.futures.as_completed(futures):
            res_ip, res_users = future.result()
            all_device_users[res_ip] = res_users
            if res_users is not None:
                online_ips.append(res_ip)

    if not online_ips:
        return jsonify({"error": "No devices are online"}), 503

    # ── Step 2: Build unified user map uid -> user data ───────────────────────
    if source_ip and source_ip in all_device_users and all_device_users[source_ip]:
        # Push only from source device
        master_users = {u["uid"]: u for u in all_device_users[source_ip]}
    else:
        # Union from all online devices
        master_users = {}
        for ip in online_ips:
            for u in (all_device_users.get(ip) or []):
                if u["uid"] not in master_users:
                    master_users[u["uid"]] = u

    # ── Step 3: Push missing users to each device ─────────────────────────────
    results = []
    
    def push_users_to_device(ip):
        if source_ip and ip == source_ip:
            return {"ip": ip, "ok": True, "added": 0, "note": "Source device"}
        existing_uids = {u["uid"] for u in (all_device_users.get(ip) or [])}
        missing = [u for uid, u in master_users.items() if uid not in existing_uids]
        if not missing:
            return {"ip": ip, "ok": True, "added": 0, "note": "Already up to date"}
        zk_conn = None
        added = 0; errors = []
        try:
            from zk import ZK
            zk_conn = ZK(ip, port=DEVICE_PORT, timeout=30, verbose=False).connect()
            zk_conn.disable_device()
            for u in missing:
                try:
                    zk_conn.set_user(
                        uid=int(u["uid"]),
                        name=u["name"],
                        privilege=u.get("privilege", 0),
                        password=u.get("password", ""),
                        group_id=u.get("group_id", ""),
                    )
                    added += 1
                except Exception as e:
                    errors.append("{0}: {1}".format(u["uid"], str(e)[:60]))
            zk_conn.enable_device()
            return {"ip": ip, "ok": True, "added": added, "errors": errors, "note": "{0} new users pushed".format(added)}
        except Exception as e:
            return {"ip": ip, "ok": False, "added": 0, "error": str(e)}
        finally:
            if zk_conn:
                try: zk_conn.enable_device(); zk_conn.disconnect()
                except: pass

    with concurrent.futures.ThreadPoolExecutor(max_workers=min(10, len(online_ips) or 1)) as executor:
        futures = [executor.submit(push_users_to_device, ip) for ip in online_ips]
        for future in concurrent.futures.as_completed(futures):
            results.append(future.result())

    total_added = sum(r.get("added",0) for r in results)
    db_manager.write_audit(session.get("username","?"), "USER_SYNC",
                           "Synced users across {0} devices, {1} users pushed".format(
                               len(online_ips), total_added),
                           request.remote_addr)
    return jsonify({
        "ok": True,
        "online_devices": len(online_ips),
        "total_users_in_master": len(master_users),
        "total_pushed": total_added,
        "results": results,
    })

@app.route("/api/devices/new-enrollments")
@admin_required
def devices_new_enrollments():
    """
    Compare current device user lists against known employees.
    Returns users on devices that are not in the employee database.
    """
    try:
        conn = _sqlite() if _sqlite_ready() else None
        known_badges = set()
        if conn:
            rows = conn.execute("SELECT badge FROM employees").fetchall()
            known_badges = {r["badge"] for r in rows}
            conn.close()
        else:
            try:
                emp_df = load_employees_all()
                known_badges = set(emp_df["Badgenumber"].tolist())
            except Exception:
                pass

        new_enrollments = []
        for ip in DEVICE_IPS:
            zk_conn = None
            try:
                from zk import ZK
                zk_conn = ZK(ip, port=DEVICE_PORT, timeout=15, verbose=False).connect()
                users = zk_conn.get_users()
                for u in users:
                    uid = str(u.user_id).strip()
                    if uid not in known_badges:
                        # Check if already in unmapped table
                        new_enrollments.append({
                            "uid":    uid,
                            "name":   u.name or "",
                            "device": ip,
                            "device_name": _device_names.get(ip, ip),
                        })
                        try: db_manager.record_unknown_user(ip, uid)
                        except Exception: pass
            except Exception:
                pass
            finally:
                if zk_conn:
                    try: zk_conn.disconnect()
                    except: pass

        return jsonify({
            "new_enrollments": new_enrollments,
            "count": len(new_enrollments),
            "known_employees": len(known_badges),
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ==============================================================================
#  CONFIG / SETTINGS ROUTES
# ==============================================================================
@app.route("/api/config")
@login_required
def get_config():
    """Return safe (non-secret) config for the frontend."""
    return jsonify({
        "version":        APP_VERSION,
        "device_count":   len(DEVICE_IPS),
        "device_ips":     DEVICE_IPS,
        "refresh_mins":   CACHE_REFRESH_MINS,
        "device_timeout": DEVICE_TIMEOUT,
        "pull_timeout":   DEVICE_PULL_TIMEOUT,
        "admin_inactivity_mins":    ADMIN_INACTIVITY_MINS,
        "employee_inactivity_mins": EMPLOYEE_INACTIVITY_MINS,
    })

# ==============================================================================
#  EMPLOYEE CRUD  (update name/badge/dept, add new, list departments)
# ==============================================================================

@app.route("/api/employee/<badge>", methods=["PUT"])
@admin_required
def update_employee(badge):
    """Update an existing employee's name, department, and/or badge code."""
    data      = request.get_json() or {}
    new_badge = (data.get("new_badge") or "").strip()
    name      = (data.get("name")      or "").strip()
    dept      = (data.get("dept")      or "").strip()
    active    = data.get("active")

    conn = get_db()
    existing = conn.execute("SELECT * FROM employees WHERE badge=?", (badge,)).fetchone()
    if not existing:
        conn.close()
        return jsonify({"error": "Employee not found: {0}".format(badge)}), 404

    try:
        now = datetime.now().isoformat()

        if new_badge and new_badge != badge:
            # Badge rename: check no collision
            clash = conn.execute("SELECT badge FROM employees WHERE badge=?", (new_badge,)).fetchone()
            if clash:
                conn.close()
                return jsonify({"error": "Badge {0} already exists".format(new_badge)}), 409
            # Update all punch records to new badge
            conn.execute("UPDATE punches       SET badge=? WHERE badge=?", (new_badge, badge))
            conn.execute("UPDATE emp_workdays  SET badge=? WHERE badge=?", (new_badge, badge))
            conn.execute("UPDATE unknown_users SET uid=?   WHERE uid=?",   (new_badge, badge))
            # Insert with new badge then delete old
            conn.execute(
                "INSERT INTO employees (badge, name, dept, active, updated_at) VALUES (?,?,?,?,?)",
                (new_badge,
                 name      if name      else existing["name"],
                 dept      if dept      else existing["dept"],
                 int(active) if active is not None else existing["active"],
                 now)
            )
            conn.execute("DELETE FROM employees WHERE badge=?", (badge,))
            # Update UID-to-badge in-memory cache
            global _uid_to_badge_cache
            for uid, b in list(_uid_to_badge_cache.items()):
                if b == badge:
                    _uid_to_badge_cache[uid] = new_badge
        else:
            # In-place update
            sets = ["updated_at=?"]
            vals = [now]
            if name:               sets.insert(0, "name=?");   vals.insert(0, name)
            if dept:               sets.insert(0, "dept=?");   vals.insert(0, dept)
            if active is not None: sets.insert(0, "active=?"); vals.insert(0, int(active))
            vals.append(badge)
            conn.execute("UPDATE employees SET {0} WHERE badge=?".format(", ".join(sets)), vals)

        conn.commit()
        write_audit(session.get("username","?"), "UPDATE_EMPLOYEE",
                    "old_badge={0} new_badge={1} name={2} dept={3}".format(
                        badge, new_badge or badge, name, dept),
                    request.remote_addr)
        return jsonify({"ok": True, "badge": new_badge or badge})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/employees/add", methods=["POST"])
@admin_required
def add_employee():
    """Add a brand-new employee to the database."""
    data   = request.get_json() or {}
    badge  = (data.get("badge") or "").strip()
    name   = (data.get("name")  or "").strip()
    dept   = (data.get("dept")  or "").strip()
    active = int(data.get("active", 1))

    if not badge or not name:
        return jsonify({"error": "badge and name are required"}), 400

    conn = get_db()
    try:
        clash = conn.execute("SELECT badge FROM employees WHERE badge=?", (badge,)).fetchone()
        if clash:
            conn.close()
            return jsonify({"error": "Employee with badge {0} already exists".format(badge)}), 409
        conn.execute(
            "INSERT INTO employees (badge, name, dept, active, updated_at) VALUES (?,?,?,?,?)",
            (badge, name, dept, active, datetime.now().isoformat())
        )
        conn.commit()
        write_audit(session.get("username","?"), "ADD_EMPLOYEE",
                    "badge={0} name={1} dept={2}".format(badge, name, dept),
                    request.remote_addr)
        return jsonify({"ok": True, "badge": badge})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/employees/departments")
@login_required
def list_departments():
    """Return distinct department names from the employees table."""
    conn = get_db()
    rows = conn.execute(
        "SELECT DISTINCT dept FROM employees WHERE dept != '' ORDER BY dept"
    ).fetchall()
    conn.close()
    return jsonify([r["dept"] for r in rows])


# ==============================================================================
#  BULK EMPLOYEE IMPORT
# ==============================================================================
@app.route("/api/employees/import", methods=["POST"])
@admin_required
def import_employees_upload():
    """Accept an uploaded Excel/CSV and import employees from it."""
    if 'file' not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f    = request.files['file']
    name = f.filename.lower()
    if not (name.endswith('.xlsx') or name.endswith('.xls') or name.endswith('.csv')):
        return jsonify({"error": "Please upload an Excel (.xlsx) or CSV file"}), 400
    try:
        buf = io.BytesIO(f.read())
        if name.endswith('.csv'):
            df = pd.read_csv(buf, dtype=str)
        else:
            raw  = pd.read_excel(buf, header=None, dtype=str)
            hrow = 0
            for i, row in raw.iterrows():
                if any(str(v).strip() == "Badgenumber" for v in row.values):
                    hrow = i; break
            buf.seek(0)
            df = pd.read_excel(buf, header=hrow, dtype=str)
        df.columns = [str(c).strip() for c in df.columns]
        required = ['Badgenumber', 'Name', 'DEPTNAME']
        missing  = [c for c in required if c not in df.columns]
        if missing:
            return jsonify({"error": "Missing columns: {0}. Need: Badgenumber, Name, DEPTNAME".format(missing)}), 400
        df["Badgenumber"] = df["Badgenumber"].astype(str).str.strip()
        df["Name"]        = df["Name"].astype(str).str.strip()
        df["DEPTNAME"]    = df["DEPTNAME"].astype(str).str.strip()
        df = df[df["Badgenumber"].notna() & (df["Badgenumber"]!="") & (df["Badgenumber"]!="nan") &
                (df["Name"]!="") & (df["Name"]!="nan")].copy()
        excl = [d.upper() for d in EXCLUDE_DEPARTMENTS]
        emp_list = [(r["Badgenumber"], r["Name"], r["DEPTNAME"],
                     0 if r["DEPTNAME"].upper() in excl else 1)
                    for _, r in df.iterrows()]
        if not _sqlite_ready():
            db_manager.init_db()
        n = db_manager.upsert_employees(emp_list)
        # Ensure uid==badge fallback entries are in the cache for newly imported employees.
        # This is important when pyzk returns badge numbers directly as user_id (common case).
        global _uid_to_badge_cache
        for badge, _name, _dept, active in emp_list:
            if badge and badge not in _uid_to_badge_cache:
                _uid_to_badge_cache[badge] = badge
        # Also save as CSV for MDB-free mode
        csv_path = os.path.join(SCRIPT_DIR, "employees_export.csv")
        df[["Badgenumber","Name","DEPTNAME"]].to_csv(csv_path, index=False, encoding="utf-8-sig")
        db_manager.write_audit(session.get("username","?"), "BULK_IMPORT",
                               "{0} employees imported from {1}".format(len(emp_list), f.filename),
                               request.remote_addr)
        return jsonify({"ok": True, "imported": len(emp_list), "total_in_db": n})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ==============================================================================
#  MAIN
# ==============================================================================

# ==============================================================================
#  REMOTE PUNCH ENGINE  (inlined from remote_punch.py)
# ==============================================================================

REMOTE_PUNCH_SCHEMA = """
CREATE TABLE IF NOT EXISTS remote_punches (
    id                   INTEGER PRIMARY KEY AUTOINCREMENT,
    badge                TEXT NOT NULL,
    employee_name        TEXT NOT NULL DEFAULT '',
    punch_time           TEXT NOT NULL,
    punch_type           TEXT NOT NULL DEFAULT 'self',
    requested_by         TEXT NOT NULL DEFAULT '',
    reason               TEXT NOT NULL DEFAULT '',
    approval_status      TEXT NOT NULL DEFAULT 'auto',
    approved_by          TEXT NOT NULL DEFAULT '',
    approved_at          TEXT NOT NULL DEFAULT '',
    device_ip            TEXT NOT NULL DEFAULT '',
    device_status        TEXT NOT NULL DEFAULT 'pending',
    device_error         TEXT NOT NULL DEFAULT '',
    device_attempts      INTEGER NOT NULL DEFAULT 0,
    device_last_attempt  TEXT NOT NULL DEFAULT '',
    created_at           TEXT NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_rp_badge    ON remote_punches(badge);
CREATE INDEX IF NOT EXISTS idx_rp_status   ON remote_punches(approval_status);
CREATE INDEX IF NOT EXISTS idx_rp_created  ON remote_punches(created_at);
"""

def init_punch_db():
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.executescript(REMOTE_PUNCH_SCHEMA)
    conn.commit(); conn.close()

def _rp_get_db():
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn

def rp_create_remote_punch(badge, employee_name, punch_time, punch_type,
                            requested_by, reason, approval_status='auto'):
    conn = _rp_get_db()
    conn.execute(
        "INSERT INTO remote_punches (badge,employee_name,punch_time,punch_type,"
        "requested_by,reason,approval_status,created_at) VALUES (?,?,?,?,?,?,?,?)",
        (badge, employee_name, punch_time, punch_type, requested_by,
         reason, approval_status, datetime.now().isoformat())
    )
    rid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.commit(); conn.close()
    return rid

def rp_set_approval(punch_id, status, approved_by):
    conn = _rp_get_db()
    conn.execute(
        "UPDATE remote_punches SET approval_status=?,approved_by=?,approved_at=? WHERE id=?",
        (status, approved_by, datetime.now().isoformat(), punch_id)
    )
    conn.commit(); conn.close()

def rp_set_device_status(punch_id, status, device_ip='', error=''):
    conn = _rp_get_db()
    conn.execute(
        "UPDATE remote_punches SET device_status=?,device_ip=?,device_error=?,"
        "device_attempts=device_attempts+1,device_last_attempt=? WHERE id=?",
        (status, device_ip, error[:300], datetime.now().isoformat(), punch_id)
    )
    conn.commit(); conn.close()

def rp_get_pending_approvals():
    conn = _rp_get_db()
    rows = conn.execute(
        "SELECT * FROM remote_punches WHERE approval_status='pending' ORDER BY created_at ASC"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def rp_get_pending_device_pushes():
    conn = _rp_get_db()
    rows = conn.execute(
        "SELECT * FROM remote_punches WHERE approval_status IN ('auto','approved')"
        " AND device_status IN ('pending','failed') AND device_attempts < 3"
        " ORDER BY punch_time ASC"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def rp_get_my_requests(badge):
    conn = _rp_get_db()
    rows = conn.execute(
        "SELECT * FROM remote_punches WHERE badge=? ORDER BY created_at DESC LIMIT 60",
        (badge,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def rp_get_all_requests(limit=200):
    conn = _rp_get_db()
    rows = conn.execute(
        "SELECT * FROM remote_punches ORDER BY created_at DESC LIMIT ?", (limit,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def rp_store_in_punches_table(badge, punch_time_iso, source='remote-punch'):
    conn = _rp_get_db()
    conn.execute(
        "INSERT OR IGNORE INTO punches (badge,punch_time,device_ip) VALUES (?,?,?)",
        (badge, punch_time_iso, source)
    )
    conn.commit(); conn.close()

def rp_push_punch_to_device(ip, port, badge, timestamp, timeout=15):
    """Write attendance to device using raw TCP (zk_writer logic inlined in sync_db.py)."""
    try:
        from sync_db import write_attendance_best_effort
        return write_attendance_best_effort(ip, int(port), badge, timestamp, timeout)
    except ImportError:
        return False, "sync_db.py not found"
    except Exception as e:
        return False, str(e)

def rp_push_to_best_device(device_ips, port, badge, timestamp, timeout=15):
    last_err = "no devices configured"
    for ip in device_ips:
        ok, msg = rp_push_punch_to_device(ip, port, badge, timestamp, timeout)
        if ok:
            return True, ip, msg
        last_err = "{}: {}".format(ip, msg)
    return False, None, last_err

_rp_push_queue    = []
_rp_push_lock     = threading.Lock()
_rp_retry_running = False

def rp_enqueue_device_push(punch_id, badge, punch_time_dt, device_ips, port):
    with _rp_push_lock:
        _rp_push_queue.append({
            "id": punch_id, "badge": badge,
            "punch_time": punch_time_dt,
            "device_ips": device_ips, "port": port,
        })

def _rp_background_pusher(device_ips, port):
    global _rp_retry_running
    _rp_retry_running = True
    while True:
        with _rp_push_lock:
            if not _rp_push_queue: break
            item = _rp_push_queue.pop(0)
        ok, ip_used, msg = rp_push_to_best_device(
            item["device_ips"] or device_ips, item["port"] or port,
            item["badge"], item["punch_time"]
        )
        if ok:
            rp_set_device_status(item["id"], "success", device_ip=ip_used)
        else:
            rp_set_device_status(item["id"], "failed", error=msg)
    _rp_retry_running = False

def rp_flush_push_queue(device_ips, port):
    global _rp_retry_running
    if not _rp_retry_running:
        t = threading.Thread(target=_rp_background_pusher, args=(device_ips, port), daemon=True)
        t.start()

def rp_retry_failed_pushes(device_ips, port):
    items = rp_get_pending_device_pushes()
    for item in items:
        try:
            dt = datetime.fromisoformat(item["punch_time"])
        except Exception:
            try: dt = datetime.strptime(item["punch_time"], "%Y-%m-%d %H:%M:%S")
            except Exception: continue
        rp_enqueue_device_push(item["id"], item["badge"], dt, device_ips, port)
    rp_flush_push_queue(device_ips, port)
    return len(items)

# Compatibility alias so old rp.xxx calls still work
class _RPShim:
    DB_PATH                 = DB_PATH
    init_punch_db           = staticmethod(init_punch_db)
    create_remote_punch     = staticmethod(rp_create_remote_punch)
    set_approval            = staticmethod(rp_set_approval)
    set_device_status       = staticmethod(rp_set_device_status)
    get_pending_approvals   = staticmethod(rp_get_pending_approvals)
    get_pending_device_pushes = staticmethod(rp_get_pending_device_pushes)
    get_my_requests         = staticmethod(rp_get_my_requests)
    get_all_requests        = staticmethod(rp_get_all_requests)
    store_in_punches_table  = staticmethod(rp_store_in_punches_table)
    enqueue_device_push     = staticmethod(rp_enqueue_device_push)
    flush_push_queue        = staticmethod(rp_flush_push_queue)
    retry_failed_pushes     = staticmethod(rp_retry_failed_pushes)
rp = _RPShim()

# ==============================================================================
#  PUNCH ROUTES  (inlined from punch_routes.py)
# ==============================================================================

def _punch_device_cfg():
    """Remote punches write to all configured biometric devices."""
    return list(get_device_ips()), int(DEVICE_PORT)

def _punch_badge_for_session():
    users = _get_users()
    uname = session.get("username", "")
    return users.get(uname, {}).get("badge")

def _punch_is_admin():
    users = _get_users()
    uname = session.get("username", "")
    return users.get(uname, {}).get("role") == "admin"

def _punch_get_employee_name(badge):
    try:
        conn = get_db()
        row  = conn.execute("SELECT name FROM employees WHERE badge=?", (badge,)).fetchone()
        conn.close()
        return row["name"] if row else badge
    except Exception:
        return badge

def _punch_today_punches(badge):
    today_str = date.today().strftime("%Y-%m-%d")
    start = today_str + " 00:00:00"
    end   = today_str + " 23:59:59"
    try:
        conn  = get_db()
        rows1 = conn.execute(
            "SELECT punch_time, device_ip FROM punches "
            "WHERE badge=? AND punch_time BETWEEN ? AND ? ORDER BY punch_time",
            (badge, start, end)
        ).fetchall()
        rows2 = conn.execute(
            "SELECT punch_time, device_status, approval_status FROM remote_punches "
            "WHERE badge=? AND punch_time BETWEEN ? AND ? AND approval_status!='rejected' "
            "ORDER BY punch_time",
            (badge, start, end)
        ).fetchall()
        conn.close()
        result = []; seen = set()
        for r in rows1:
            t = r["punch_time"]
            if t not in seen:
                result.append({"time": t, "source": r["device_ip"] or "device", "status": "confirmed"})
                seen.add(t)
        for r in rows2:
            t = r["punch_time"]
            if t not in seen:
                status = "pending" if r["approval_status"] == "pending" else (
                    "device_ok" if r["device_status"] == "success" else r["device_status"])
                result.append({"time": t, "source": "remote", "status": status})
                seen.add(t)
        result.sort(key=lambda x: x["time"])
        return result
    except Exception:
        return []


@app.route("/api/punch/status")
def punch_status():
    if "username" not in session:
        return jsonify({"error": "Not authenticated", "auth_required": True}), 401
    badge = _punch_badge_for_session()
    if not badge:
        return jsonify({"error": "No badge linked to your account."}), 400
    punches = _punch_today_punches(badge)
    return jsonify({"badge": badge, "date": date.today().isoformat(),
                    "punches": punches, "count": len(punches)})


@app.route("/api/punch/self", methods=["POST"])
def punch_self():
    if "username" not in session:
        return jsonify({"error": "Not authenticated", "auth_required": True}), 401
    badge = _punch_badge_for_session()
    if not badge:
        return jsonify({"error": "No badge linked to your account."}), 400
    now = datetime.now(); now_iso = now.strftime("%Y-%m-%d %H:%M:%S")
    emp_name = _punch_get_employee_name(badge)
    device_ips, port = _punch_device_cfg()
    punch_id = rp.create_remote_punch(badge, emp_name, now_iso, "self",
                                       session["username"], "Self punch via dashboard", "auto")
    rp.store_in_punches_table(badge, now_iso, source="remote-self")
    rp.enqueue_device_push(punch_id, badge, now, device_ips, port)
    rp.flush_push_queue(device_ips, port)
    write_audit(session["username"], "SELF_PUNCH", "badge={} time={}".format(badge, now_iso),
                request.remote_addr)
    return jsonify({"ok": True, "punch_id": punch_id, "time": now_iso,
                    "message": "Punched in at {}".format(now.strftime("%H:%M"))})


@app.route("/api/punch/request", methods=["POST"])
def punch_request():
    if "username" not in session:
        return jsonify({"error": "Not authenticated", "auth_required": True}), 401
    badge = _punch_badge_for_session()
    if not badge:
        return jsonify({"error": "No badge linked to your account."}), 400
    data = request.get_json() or {}
    dt_str = (data.get("punch_time") or "").strip()
    reason = (data.get("reason") or "").strip()
    if not reason:
        return jsonify({"error": "reason is required"}), 400
    punch_dt = None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M"):
        try: punch_dt = datetime.strptime(dt_str, fmt); break
        except ValueError: continue
    if not punch_dt:
        return jsonify({"error": "Invalid date/time format. Use YYYY-MM-DD HH:MM"}), 400
    if punch_dt > datetime.now():
        return jsonify({"error": "Cannot request a punch for a future time"}), 400
    punch_iso = punch_dt.strftime("%Y-%m-%d %H:%M:%S")
    emp_name = _punch_get_employee_name(badge)
    punch_id = rp.create_remote_punch(badge, emp_name, punch_iso, "request",
                                       session["username"], reason, "pending")
    write_audit(session["username"], "PUNCH_REQUEST",
                "badge={} time={} reason={}".format(badge, punch_iso, reason[:80]),
                request.remote_addr)
    return jsonify({"ok": True, "punch_id": punch_id,
                    "message": "Punch request submitted for {}. Waiting for admin approval.".format(
                        punch_dt.strftime("%d %b %Y %H:%M"))})


@app.route("/api/punch/my-requests")
def punch_my_requests():
    if "username" not in session:
        return jsonify({"error": "Not authenticated", "auth_required": True}), 401
    badge = _punch_badge_for_session()
    if not badge:
        return jsonify({"error": "No badge linked to your account."}), 400
    return jsonify(rp.get_my_requests(badge))


@app.route("/api/punch/admin/pending")
def punch_admin_pending():
    if "username" not in session:
        return jsonify({"error": "Not authenticated", "auth_required": True}), 401
    if not _punch_is_admin():
        return jsonify({"error": "Admin access required"}), 403
    return jsonify(rp.get_pending_approvals())


@app.route("/api/punch/admin/approve/<int:punch_id>", methods=["POST"])
def punch_admin_approve(punch_id):
    if "username" not in session:
        return jsonify({"error": "Not authenticated", "auth_required": True}), 401
    if not _punch_is_admin():
        return jsonify({"error": "Admin access required"}), 403
    data   = request.get_json() or {}
    action = (data.get("action") or "").lower()
    if action not in ("approve", "reject"):
        return jsonify({"error": "action must be 'approve' or 'reject'"}), 400
    conn = get_db()
    row  = conn.execute("SELECT * FROM remote_punches WHERE id=?", (punch_id,)).fetchone()
    conn.close()
    if not row:
        return jsonify({"error": "Punch request not found"}), 404
    if row["approval_status"] != "pending":
        return jsonify({"error": "This request is already {}".format(row["approval_status"])}), 409
    admin_user = session["username"]
    if action == "reject":
        rp.set_approval(punch_id, "rejected", admin_user)
        write_audit(admin_user, "PUNCH_REJECT",
                    "id={} badge={} time={}".format(punch_id, row["badge"], row["punch_time"]),
                    request.remote_addr)
        return jsonify({"ok": True, "message": "Request rejected."})
    rp.set_approval(punch_id, "approved", admin_user)
    rp.store_in_punches_table(row["badge"], row["punch_time"], source="admin-approved")
    device_ips, port = _punch_device_cfg()
    try:    punch_dt = datetime.fromisoformat(row["punch_time"])
    except: punch_dt = datetime.strptime(row["punch_time"], "%Y-%m-%d %H:%M:%S")
    rp.enqueue_device_push(punch_id, row["badge"], punch_dt, device_ips, port)
    rp.flush_push_queue(device_ips, port)
    write_audit(admin_user, "PUNCH_APPROVE",
                "id={} badge={} time={}".format(punch_id, row["badge"], row["punch_time"]),
                request.remote_addr)
    return jsonify({"ok": True, "message": "Approved. Punch stored and being sent to device."})


@app.route("/api/punch/admin/direct", methods=["POST"])
def punch_admin_direct():
    if "username" not in session:
        return jsonify({"error": "Not authenticated", "auth_required": True}), 401
    if not _punch_is_admin():
        return jsonify({"error": "Admin access required"}), 403
    data   = request.get_json() or {}
    badge  = (data.get("badge") or "").strip()
    dt_str = (data.get("punch_time") or "").strip()
    reason = (data.get("reason") or "Admin manual entry").strip()
    if not badge:  return jsonify({"error": "badge is required"}), 400
    if not dt_str: return jsonify({"error": "punch_time is required (YYYY-MM-DD HH:MM)"}), 400
    punch_dt = None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M"):
        try: punch_dt = datetime.strptime(dt_str, fmt); break
        except ValueError: continue
    if not punch_dt:
        return jsonify({"error": "Invalid date/time. Use YYYY-MM-DD HH:MM"}), 400
    emp_name  = _punch_get_employee_name(badge)
    punch_iso = punch_dt.strftime("%Y-%m-%d %H:%M:%S")
    punch_id  = rp.create_remote_punch(badge, emp_name, punch_iso, "admin",
                                        session["username"], reason, "auto")
    rp.store_in_punches_table(badge, punch_iso, source="admin-direct")
    device_ips, port = _punch_device_cfg()
    rp.enqueue_device_push(punch_id, badge, punch_dt, device_ips, port)
    rp.flush_push_queue(device_ips, port)
    write_audit(session["username"], "ADMIN_DIRECT_PUNCH",
                "badge={} name={} time={} reason={}".format(badge, emp_name, punch_iso, reason[:80]),
                request.remote_addr)
    return jsonify({"ok": True, "punch_id": punch_id, "badge": badge, "name": emp_name,
                    "time": punch_iso,
                    "message": "Punch added for {} at {}. Sending to device...".format(
                        emp_name, punch_dt.strftime("%d %b %Y %H:%M"))})


@app.route("/api/punch/admin/all")
def punch_admin_all():
    if "username" not in session:
        return jsonify({"error": "Not authenticated", "auth_required": True}), 401
    if not _punch_is_admin():
        return jsonify({"error": "Admin access required"}), 403
    limit = int(request.args.get("limit", 200))
    return jsonify(rp.get_all_requests(limit))


@app.route("/api/punch/admin/device-queue")
def punch_device_queue():
    if "username" not in session:
        return jsonify({"error": "Not authenticated", "auth_required": True}), 401
    if not _punch_is_admin():
        return jsonify({"error": "Admin access required"}), 403
    return jsonify(rp.get_pending_device_pushes())


@app.route("/api/punch/admin/retry-device", methods=["POST"])
def punch_retry_device():
    if "username" not in session:
        return jsonify({"error": "Not authenticated", "auth_required": True}), 401
    if not _punch_is_admin():
        return jsonify({"error": "Admin access required"}), 403
    device_ips, port = _punch_device_cfg()
    count = rp.retry_failed_pushes(device_ips, port)
    write_audit(session.get("username","?"), "RETRY_DEVICE_PUSH",
                "{} records re-queued".format(count), request.remote_addr)
    return jsonify({"ok": True, "queued": count,
                    "message": "{} punch(es) queued for device push.".format(count)})

# ==============================================================================
#  VOIP — Database schema + in-memory registry
# ==============================================================================

VOIP_DB_SCHEMA = """
CREATE TABLE IF NOT EXISTS voip_call_log (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    caller      TEXT NOT NULL,
    callee      TEXT NOT NULL,
    started_at  TEXT NOT NULL,
    ended_at    TEXT,
    duration_s  INTEGER DEFAULT 0,
    status      TEXT NOT NULL DEFAULT 'initiated'
);
CREATE INDEX IF NOT EXISTS idx_voip_caller ON voip_call_log(caller);
CREATE INDEX IF NOT EXISTS idx_voip_callee ON voip_call_log(callee);
"""

def _init_voip_db():
    try:
        conn = get_db()
        conn.executescript(VOIP_DB_SCHEMA)
        conn.commit(); conn.close()
        print("[VoIP] DB schema ready")
    except Exception as e:
        print("[VoIP] DB init error:", e)

def _log_call(caller, callee, status='ended', duration_s=0, call_id=None):
    """Insert or update a call log entry."""
    try:
        conn = get_db()
        if call_id:
            conn.execute(
                "UPDATE voip_call_log SET ended_at=?, duration_s=?, status=? WHERE id=?",
                (datetime.now().isoformat(), duration_s, status, call_id)
            )
        else:
            conn.execute(
                "INSERT INTO voip_call_log (caller, callee, started_at, status) VALUES (?,?,?,?)",
                (caller, callee, datetime.now().isoformat(), status)
            )
        conn.commit()
        cid = conn.execute("SELECT last_insert_rowid()").fetchone()[0] if not call_id else call_id
        conn.close()
        return cid
    except Exception as e:
        print("[VoIP] log_call error:", e)
        return None

# In-memory: username -> socket-id / info
_voip_online      = {}   # username -> sid
_voip_online_info = {}   # username -> {name, badge, dept}
_voip_online_lock = threading.Lock()
_voip_active_calls = {}  # call_id -> {caller, callee, started_at, db_id}

def _voip_broadcast_online():
    """Push updated online list to all connected clients."""
    if not SOCKETIO_AVAILABLE or not socketio:
        return
    with _voip_online_lock:
        users = [
            {"username": u, **(dict(_voip_online_info.get(u, {})))}
            for u in _voip_online
        ]
    socketio.emit('voip_online_update', {"users": users})


# ==============================================================================
#  VOIP — Socket.IO events
# ==============================================================================

if SOCKETIO_AVAILABLE and socketio:

    @socketio.on('connect')
    def _voip_on_connect():
        pass  # registration happens via voip_register

    @socketio.on('disconnect')
    def _voip_on_disconnect():
        from flask_socketio import request as sio_req
        sid = sio_req.sid
        with _voip_online_lock:
            to_remove = [u for u, s in _voip_online.items() if s == sid]
            for u in to_remove:
                del _voip_online[u]
                _voip_online_info.pop(u, None)
        if to_remove:
            _voip_broadcast_online()

    @socketio.on('voip_register')
    def _voip_register(data):
        """Client registers itself: {username, name, badge, dept}"""
        from flask_socketio import request as sio_req
        username = data.get('username', '').strip()
        if not username:
            return
        with _voip_online_lock:
            _voip_online[username] = sio_req.sid
            _voip_online_info[username] = {
                'name':  data.get('name', username),
                'badge': data.get('badge', ''),
                'dept':  data.get('dept', ''),
            }
        _voip_broadcast_online()

    @socketio.on('voip_call')
    def _voip_call(data):
        """Caller initiates: {caller, caller_name, callee, offer_sdp}"""
        caller = data.get('caller', '')
        callee = data.get('callee', '')
        if not caller or not callee:
            return
        with _voip_online_lock:
            callee_sid = _voip_online.get(callee)
        if not callee_sid:
            emit('voip_error', {'message': '{} is not online'.format(callee), 'callee': callee})
            return
        call_id = str(int(time.time() * 1000))
        db_id   = _log_call(caller, callee, status='ringing')
        _voip_active_calls[call_id] = {
            'caller':     caller,
            'callee':     callee,
            'started_at': time.time(),
            'db_id':      db_id,
        }
        socketio.emit('voip_incoming', {
            'call_id':     call_id,
            'caller':      caller,
            'caller_name': data.get('caller_name', caller),
            'offer_sdp':   data.get('offer_sdp', ''),
        }, to=callee_sid)
        emit('voip_calling', {'call_id': call_id, 'callee': callee})

    @socketio.on('voip_answer')
    def _voip_answer(data):
        """Callee accepts. {call_id, answer_sdp}"""
        call_id = data.get('call_id', '')
        call    = _voip_active_calls.get(call_id)
        if not call: return
        with _voip_online_lock:
            caller_sid = _voip_online.get(call['caller'])
        if caller_sid:
            socketio.emit('voip_answered', {
                'call_id':    call_id,
                'answer_sdp': data.get('answer_sdp', ''),
            }, to=caller_sid)
        if call.get('db_id'):
            _log_call(call['caller'], call['callee'], status='active', call_id=call['db_id'])

    @socketio.on('voip_reject')
    def _voip_reject(data):
        """Callee rejects. {call_id}"""
        call_id = data.get('call_id', '')
        call    = _voip_active_calls.pop(call_id, None)
        if not call: return
        with _voip_online_lock:
            caller_sid = _voip_online.get(call['caller'])
        if caller_sid:
            socketio.emit('voip_rejected', {'call_id': call_id}, to=caller_sid)
        if call.get('db_id'):
            _log_call(call['caller'], call['callee'], status='rejected', call_id=call['db_id'])

    @socketio.on('voip_end')
    def _voip_end(data):
        """Either party ends. {call_id, initiator}"""
        call_id  = data.get('call_id', '')
        call     = _voip_active_calls.pop(call_id, None)
        if not call: return
        duration = int(time.time() - call.get('started_at', time.time()))
        initiator = data.get('initiator', '')
        with _voip_online_lock:
            caller_sid = _voip_online.get(call['caller'])
            callee_sid = _voip_online.get(call['callee'])
        payload = {'call_id': call_id, 'initiator': initiator, 'duration': duration}
        if caller_sid: socketio.emit('voip_ended', payload, to=caller_sid)
        if callee_sid: socketio.emit('voip_ended', payload, to=callee_sid)
        if call.get('db_id'):
            _log_call(call['caller'], call['callee'],
                      status='completed', duration_s=duration, call_id=call['db_id'])

    @socketio.on('voip_ice')
    def _voip_ice(data):
        """Relay ICE candidate. {call_id, target, candidate}"""
        call_id = data.get('call_id', '')
        call    = _voip_active_calls.get(call_id)
        if not call: return
        target  = data.get('target', '')
        with _voip_online_lock:
            target_sid = _voip_online.get(target)
        if target_sid:
            socketio.emit('voip_ice', {
                'call_id':   call_id,
                'candidate': data.get('candidate', ''),
            }, to=target_sid)


# ==============================================================================
#  VOIP — REST endpoints
# ==============================================================================

@app.route('/api/voip/online')
def voip_api_online():
    if 'username' not in session:
        return jsonify({'error': 'Not authenticated', 'auth_required': True}), 401
    me = session.get('username', '')
    with _voip_online_lock:
        users = [
            {"username": u, **(dict(_voip_online_info.get(u, {})))}
            for u in _voip_online if u != me
        ]
    return jsonify({"users": users})

@app.route('/api/voip/history')
def voip_api_history():
    if 'username' not in session:
        return jsonify({'error': 'Not authenticated', 'auth_required': True}), 401
    me    = session.get('username', '')
    limit = int(request.args.get('limit', 50))
    try:
        conn = get_db()
        rows = conn.execute(
            """SELECT id, caller, callee, started_at, ended_at, duration_s, status
               FROM voip_call_log
               WHERE caller=? OR callee=?
               ORDER BY id DESC LIMIT ?""",
            (me, me, limit)
        ).fetchall()
        conn.close()
        return jsonify({"calls": [dict(r) for r in rows]})
    except Exception as e:
        return jsonify({"calls": [], "error": str(e)})

@app.route('/api/voip/status')
def voip_api_status():
    return jsonify({"enabled": SOCKETIO_AVAILABLE})


# ==============================================================================
#  MAIN
# ==============================================================================

init_punch_db()


if __name__ == "__main__":
    import socket
    def _get_local_ip():
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect(("8.8.8.8", 80))
            ip = s.getsockname()[0]; s.close(); return ip
        except Exception:
            return "127.0.0.1"

    local_ip = _get_local_ip()
    print("\n" + "="*58)
    print("  ZKTeco Attendance Dashboard  v2.1")
    print("  Folder        : {0}".format(SCRIPT_DIR))
    print("  Open browser  -> http://localhost:5000/d")
    print("  Also works    -> http://127.0.0.1:5000/d")
    print("  Default login -> admin / {0}".format(DEFAULT_ADMIN_PASSWORD))
    print("="*58 + "\n")
    start_background_refresh()
    _init_voip_db()
    if SOCKETIO_AVAILABLE and socketio:
        print("  VoIP calling  -> enabled (WebRTC/SocketIO)")
        socketio.run(app, host="0.0.0.0", port=5000, debug=False,
                     allow_unsafe_werkzeug=True)
    else:
        app.run(host="0.0.0.0", port=5000, debug=False, threaded=True)