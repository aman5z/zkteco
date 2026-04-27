# -*- coding: utf-8 -*-
"""
ZKTeco MDB Tools  --  Merged File
==================================
Combines:
  scan_mdb.py         : Inspect an MDB file to find table/column names
  EXPORT_EMPLOYEES.py : Export employee list from MDB to XLSX + CSV
  attendance_tool.py  : Generate absent reports (today from devices, or history from MDB)

Usage:
  python mdb_tools.py scan                                           # scan MDB structure
  python mdb_tools.py scan path\\to\\backup.mdb                      # scan specific MDB
  python mdb_tools.py export                                        # export employees
  python mdb_tools.py export path\\to\\backup.mdb                    # export from specific MDB
  python mdb_tools.py today                                         # today's absent report
  python mdb_tools.py history backup.mdb 01/03/2026 12/03/2026     # history absent report

Requirements:
  pip install pyodbc openpyxl pandas pyzk
  Microsoft Access Database Engine:
  https://www.microsoft.com/en-us/download/details.aspx?id=54920
"""

import sys, os, warnings
warnings.filterwarnings("ignore")

# ==============================================================================
#  SHARED HELPERS
# ==============================================================================

def _get_script_dir():
    """UNC-safe script directory resolution."""
    for fn in (os.path.realpath, os.path.abspath):
        try:
            p = os.path.dirname(fn(__file__))
            if p and os.path.isdir(p): return p
        except Exception: pass
    try:
        p = os.path.dirname(os.path.realpath(sys.argv[0]))
        if p and os.path.isdir(p): return p
    except Exception: pass
    return os.getcwd()

SCRIPT_DIR = _get_script_dir()

# ==============================================================================
#  PUNCH DEDUPLICATION  (same logic as server.py)
# ==============================================================================
_DEDUP_GAP_SECS = 60

def _dedupe_times(times_list, gap_secs=_DEDUP_GAP_SECS):
    """Keep only punch times >= gap_secs apart. Input: sorted HH:MM:SS strings."""
    from datetime import datetime as _dt
    result = []; last = None
    for t_str in sorted(str(t) for t in times_list if str(t) not in ("nan","None","")):
        try:
            t = _dt.strptime(t_str[:8], "%H:%M:%S").replace(year=2000, month=1, day=1)
        except Exception:
            result.append(t_str); continue
        if last is None or (t - last).total_seconds() >= gap_secs:
            result.append(t_str); last = t
    return result




def find_mdb(explicit=None):
    """Find .mdb/.accdb: use explicit path, then scan SCRIPT_DIR."""
    if explicit:
        if not os.path.exists(explicit):
            print("[ERROR] File not found: {0}".format(explicit))
            sys.exit(1)
        return explicit
    for f in os.listdir(SCRIPT_DIR):
        if f.lower().endswith(".mdb") or f.lower().endswith(".accdb"):
            return os.path.join(SCRIPT_DIR, f)
    return None


def connect_mdb(path):
    """Open an Access MDB/ACCDB connection via pyodbc."""
    try:
        import pyodbc
        return pyodbc.connect(
            r"Driver={Microsoft Access Driver (*.mdb, *.accdb)};"
            "Dbq={0};Uid=Admin;Pwd=;".format(os.path.abspath(path))
        )
    except ImportError:
        print("[ERROR] pyodbc not installed: pip install pyodbc"); sys.exit(1)
    except Exception as e:
        print("[ERROR] Cannot open database: {0}".format(e))
        print("Install driver: https://www.microsoft.com/en-us/download/details.aspx?id=54920")
        sys.exit(1)


# ==============================================================================
#  SCAN MODE  (was scan_mdb.py)
#  Prints all tables, columns, and row counts — use this first to learn MDB layout
# ==============================================================================

def cmd_scan(argv):
    """python mdb_tools.py scan [path_to.mdb]"""
    mdb = argv[0] if argv else find_mdb()
    if not mdb:
        print("[ERROR] No .mdb file found. Place it in the same folder or pass path as argument.")
        print("Usage: python mdb_tools.py scan path\\to\\backup.mdb")
        input("\nPress Enter to exit..."); sys.exit(1)

    print("\nScanning: {0}\n".format(mdb))
    conn   = connect_mdb(mdb)
    cursor = conn.cursor()
    tables = [r.table_name for r in cursor.tables(tableType="TABLE")]

    print("=" * 60)
    print("  Found {0} tables".format(len(tables)))
    print("=" * 60)

    for table in tables:
        try:
            cols  = [r.column_name for r in cursor.columns(table=table)]
            try:
                count = cursor.execute("SELECT COUNT(*) FROM [{0}]".format(table)).fetchone()[0]
            except:
                count = "?"
            print("\n  TABLE: {0}  ({1} rows)".format(table, count))
            print("  " + "-" * 50)
            for col in cols:
                print("    - {0}".format(col))
        except Exception as e:
            print("\n  TABLE: {0}  [Could not read: {1}]".format(table, e))

    print("\n" + "=" * 60)
    print("  Copy the correct table/column names into settings or attendance_tool calls.")
    print("  Look for:")
    print("    - A table with punch/check-in records (has timestamps)")
    print("    - A table with employee names")
    print("=" * 60)
    conn.close()
    input("\nPress Enter to exit...")


# ==============================================================================
#  EXPORT MODE  (was EXPORT_EMPLOYEES.py)
#  Reads DEPARTMENTS + USERINFO from MDB and produces employees_export.xlsx + .csv
# ==============================================================================

# Dept grouping for Excel output
SECTION_CONFIG = [
    ("ADMIN",          ["ADMIN"]),
    ("SUPPORT",        ["SUPPORT"]),
    ("TEACHING",       ["TEACHING"]),
    ("TRANSPORT DEPT", ["CONDUCTOR", "DRIVER"]),
    ("CLEANING STAFF", ["CLEANING STAFF"]),
    ("DELETED",        ["DELETED EMPLOYEES", "TRANSPORT"]),
    ("GAES",           ["GAES", "GULF ASIAN ENGLISH SCHOOL"]),
]


def _build_sort_keys():
    dept_to_section = {}
    dept_sort_order = {}
    for sec_idx, (section, depts) in enumerate(SECTION_CONFIG):
        for dept_idx, dept in enumerate(depts):
            u = dept.upper()
            dept_to_section[u] = section
            dept_sort_order[u] = (sec_idx, dept_idx)
    return dept_to_section, dept_sort_order


def _xl_thin():
    from openpyxl.styles import Border, Side
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _xl_hdr_col(ws, row, col, val):
    from openpyxl.styles import Font, PatternFill, Alignment
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    c.fill      = PatternFill("solid", start_color="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = _xl_thin()


def _xl_section_hdr(ws, row, label):
    from openpyxl.styles import Font, PatternFill, Alignment
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c = ws.cell(row=row, column=1, value=label)
    c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    c.fill      = PatternFill("solid", start_color="1F4E79")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border    = _xl_thin()
    ws.row_dimensions[row].height = 20


def _xl_sub_hdr(ws, row, label):
    from openpyxl.styles import Font, PatternFill, Alignment
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c = ws.cell(row=row, column=1, value=label)
    c.font      = Font(name="Arial", bold=True, color="1F4E79", size=10)
    c.fill      = PatternFill("solid", start_color="DCE6F1")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    c.border    = _xl_thin()
    ws.row_dimensions[row].height = 17


def _xl_data_cell(ws, row, col, val, bg, align="left"):
    from openpyxl.styles import Font, PatternFill, Alignment
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Arial", size=10)
    c.fill      = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = _xl_thin()
    ws.row_dimensions[row].height = 16


def cmd_export(argv):
    """python mdb_tools.py export [path_to.mdb]"""
    import pandas as pd
    from openpyxl import Workbook
    from datetime import datetime as _dt

    mdb = find_mdb(argv[0] if argv else None)
    if not mdb:
        print("[ERROR] No .mdb / .accdb file found.")
        print("  Place it in the same folder or: python mdb_tools.py export path\\to\\backup.mdb")
        input("\nPress Enter to exit..."); sys.exit(1)

    print("\n" + "=" * 55)
    print("  MDB Employee Exporter")
    print("=" * 55)
    print("  Source : {0}".format(mdb))

    conn = connect_mdb(mdb)

    print("\n[1/4] Loading DEPARTMENTS table...")
    try:
        dept_df = pd.read_sql("SELECT [DEPTID], [DEPTNAME] FROM [DEPARTMENTS]", conn)
    except Exception as e:
        print("[ERROR] Could not read DEPARTMENTS: {0}".format(e))
        input("\nPress Enter to exit..."); sys.exit(1)

    id_to_name = dict(zip(
        dept_df["DEPTID"].astype(str).str.strip(),
        dept_df["DEPTNAME"].astype(str).str.strip().str.upper()
    ))
    print("  Found {0} departments".format(len(id_to_name)))

    print("\n[2/4] Loading USERINFO table...")
    try:
        user_df = pd.read_sql(
            "SELECT [Badgenumber], [Name], [DEFAULTDEPTID] FROM [USERINFO]", conn
        )
    except Exception as e:
        print("[ERROR] Could not read USERINFO: {0}".format(e))
        input("\nPress Enter to exit..."); sys.exit(1)
    conn.close()
    print("  Found {0} employees".format(len(user_df)))

    print("\n[3/4] Mapping and sorting...")
    user_df["Badgenumber"]   = user_df["Badgenumber"].astype(str).str.strip()
    user_df["Name"]          = user_df["Name"].astype(str).str.strip()
    user_df["DEFAULTDEPTID"] = user_df["DEFAULTDEPTID"].astype(str).str.strip()
    user_df["DEPTNAME"]      = user_df["DEFAULTDEPTID"].map(id_to_name).fillna("UNKNOWN")

    dept_to_section, dept_sort_order = _build_sort_keys()
    user_df["_SECTION"]  = user_df["DEPTNAME"].map(dept_to_section).fillna("UNKNOWN")
    user_df["_SORT"]     = user_df["DEPTNAME"].map(lambda d: dept_sort_order.get(d.upper(), (999, 999)))
    user_df["_SEC_IDX"]  = user_df["_SORT"].map(lambda x: x[0])
    user_df["_DEPT_IDX"] = user_df["_SORT"].map(lambda x: x[1])
    user_df = user_df.sort_values(["_SEC_IDX", "_DEPT_IDX", "Name"]).reset_index(drop=True)
    out_df  = user_df[["Badgenumber", "Name", "DEPTNAME", "_SECTION"]].copy()

    print("\n  Breakdown:")
    for section, depts in SECTION_CONFIG:
        sec_total = 0
        for dept in depts:
            count     = len(out_df[out_df["DEPTNAME"] == dept.upper()])
            sec_total += count
            if len(depts) > 1:
                print("    {0:20s}  + {1:30s} {2:4d}".format(section, dept, count))
            else:
                print("    {0:20s}  {1:4d}".format(section, count))
        if len(depts) > 1:
            print("    {0:20s}    SUBTOTAL {1:4d}".format("", sec_total))
    print("\n  Total: {0} employees".format(len(out_df)))

    print("\n[4/4] Saving files...")
    xlsx_path = os.path.join(SCRIPT_DIR, "employees_export.xlsx")
    csv_path  = os.path.join(SCRIPT_DIR, "employees_export.csv")

    out_df[["Badgenumber", "Name", "DEPTNAME"]].to_csv(csv_path, index=False, encoding="utf-8-sig")
    print("  OK CSV  -> {0}".format(csv_path))

    wb = Workbook()
    ws = wb.active
    ws.title = "EMPLOYEES"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:C1")
    from openpyxl.styles import Font, Alignment
    c = ws.cell(row=1, column=1,
                value="Employee Export  --  {0}".format(_dt.now().strftime('%d %b %Y %H:%M')))
    c.font      = Font(name="Arial", bold=True, size=13, color="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for col, hdr in enumerate(["Badgenumber", "Name", "Department"], 1):
        _xl_hdr_col(ws, 2, col, hdr)
    ws.row_dimensions[2].height = 20

    xl_row = 3; prev_section = None; prev_dept = None; alt = 0
    for _, rec in out_df.iterrows():
        section = rec["_SECTION"]
        dept    = rec["DEPTNAME"]
        if section != prev_section:
            _xl_section_hdr(ws, xl_row, section)
            xl_row += 1; prev_section = section; prev_dept = None; alt = 0
        section_depts = next(d for s, d in SECTION_CONFIG if s == section)
        if len(section_depts) > 1 and dept != prev_dept:
            _xl_sub_hdr(ws, xl_row, dept)
            xl_row += 1; prev_dept = dept; alt = 0
        bg = "EBF3FB" if alt % 2 == 0 else "FFFFFF"
        _xl_data_cell(ws, xl_row, 1, rec["Badgenumber"], bg, align="center")
        _xl_data_cell(ws, xl_row, 2, rec["Name"],        bg)
        _xl_data_cell(ws, xl_row, 3, rec["DEPTNAME"],    bg)
        xl_row += 1; alt += 1

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 28
    wb.save(xlsx_path)
    print("  OK XLSX -> {0}".format(xlsx_path))

    print("\n" + "=" * 55)
    print("  Done! {0} employees exported.".format(len(out_df)))
    print("=" * 55)
    input("\nPress Enter to exit...")


# ==============================================================================
#  ATTENDANCE REPORT MODE  (was attendance_tool.py)
#  Generates absent reports as XLSX
# ==============================================================================

EMPLOYEES_FILE      = os.path.join(SCRIPT_DIR, "employees_export.csv")
OUTPUT_TODAY        = os.path.join(SCRIPT_DIR, "absent_today.xlsx")
OUTPUT_HISTORY      = os.path.join(SCRIPT_DIR, "absent_history.xlsx")
EXCLUDE_DEPARTMENTS = ["DELETED EMPLOYEES", "TRANSPORT", "GAES", "GULF ASIAN ENGLISH SCHOOL"]
DEPT_ORDER          = ["ADMIN", "SUPPORT", "TEACHING", "CONDUCTOR", "DRIVER", "CLEANING STAFF"]

# ZK device connection config (read from settings.ini if present)
import configparser as _cp
_cfg = _cp.ConfigParser()
_ini = os.path.join(SCRIPT_DIR, "settings.ini")
if os.path.exists(_ini):
    _cfg.read(_ini, encoding="utf-8")

def _cfg_get(section, key, default):
    try:    return _cfg.get(section, key).strip()
    except: return default

def _cfg_list(section, key, default):
    raw = _cfg_get(section, key, "")
    return [x.strip() for x in raw.split(",") if x.strip()] if raw else default

DEVICE_IPS     = _cfg_list("devices", "ips",
                    ["10.20.141.21","10.20.141.22","10.20.141.23","10.20.141.24"])
DEVICE_PORT    = int(_cfg_get("devices", "port",    "4370"))
DEVICE_TIMEOUT = int(_cfg_get("devices", "timeout", "10"))

# ── Excel styling helpers ──────────────────────────────────────────────────────
_CLR_HDR_BG = "1F4E79"; _CLR_HDR_FG = "FFFFFF"; _CLR_ALT = "DCE6F1"

def _rpt_thin():
    from openpyxl.styles import Border, Side
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _rpt_header_cell(ws, row, col, value):
    from openpyxl.styles import Font, PatternFill, Alignment
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=True, color=_CLR_HDR_FG, size=10)
    c.fill      = PatternFill("solid", start_color=_CLR_HDR_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = _rpt_thin()
    return c

def _rpt_data_cell(ws, row, col, value, bg=None, align="left"):
    from openpyxl.styles import Font, PatternFill, Alignment
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", size=10)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = _rpt_thin()
    if bg: c.fill = PatternFill("solid", start_color=bg)
    return c

def _dept_sort_key(dept):
    try:    return DEPT_ORDER.index(dept.upper())
    except: return len(DEPT_ORDER)


def _load_employees():
    import pandas as pd
    path = EMPLOYEES_FILE
    if not os.path.exists(path):
        alt = path.replace(".csv", ".xlsx") if path.endswith(".csv") else path.replace(".xlsx", ".csv")
        if os.path.exists(alt): path = alt
        else:
            print("[ERROR] Employee file not found: {0}".format(path))
            print("  Run: python mdb_tools.py export")
            sys.exit(1)
    if path.lower().endswith(".csv"):
        df = pd.read_csv(path, dtype=str)
    else:
        raw  = pd.read_excel(path, header=None, dtype=str)
        hrow = 0
        for i, row in raw.iterrows():
            if any(str(v).strip() == "Badgenumber" for v in row.values): hrow = i; break
        df = pd.read_excel(path, header=hrow, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    df["Badgenumber"] = df["Badgenumber"].astype(str).str.strip()
    df["Name"]        = df["Name"].astype(str).str.strip()
    df["DEPTNAME"]    = df["DEPTNAME"].astype(str).str.strip()
    df = df[df["Badgenumber"].notna() & (df["Badgenumber"] != "") & (df["Badgenumber"] != "nan") &
            (df["Name"] != "") & (df["Name"] != "nan")].copy()
    excl = [d.upper() for d in EXCLUDE_DEPARTMENTS]
    df = df[~df["DEPTNAME"].str.upper().isin(excl)].copy()
    print("  Employees loaded: {0} from {1}".format(len(df), os.path.basename(path)))
    return df.reset_index(drop=True)


def _mdb_uid_to_badge(conn):
    import pandas as pd
    df = pd.read_sql("SELECT [USERID],[Badgenumber] FROM [USERINFO]", conn)
    df["USERID"]      = df["USERID"].astype(str).str.strip()
    df["Badgenumber"] = df["Badgenumber"].astype(str).str.strip()
    return dict(zip(df["USERID"], df["Badgenumber"]))


def _write_absent_excel(absent_rows, output_path, title_suffix=""):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import datetime as _dt
    wb = Workbook(); ws = wb.active; ws.title = "Absent Report"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:D1")
    c = ws.cell(row=1, column=1, value="Absent Report -- {0}".format(title_suffix))
    c.font      = Font(name="Arial", bold=True, size=13, color=_CLR_HDR_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.cell(row=2, column=1,
        value="Generated: {0}  |  Total: {1} absent records".format(
            _dt.now().strftime('%d %b %Y %H:%M'), len(absent_rows))
    ).font = Font(name="Arial", size=9, italic=True, color="888888")
    ws.row_dimensions[2].height = 16

    for col, hdr in enumerate(["Emp Code", "Name", "Department", "Date"], 1):
        _rpt_header_cell(ws, 4, col, hdr)
    ws.row_dimensions[4].height = 18

    sorted_rows = sorted(absent_rows,
        key=lambda r: (_dept_sort_key(r["Department"]), r["Department"], r["Name"], r["Date"]))

    row = 5; prev_dept = None
    for i, rec in enumerate(sorted_rows):
        dept = rec["Department"]
        if dept != prev_dept:
            for col in range(1, 5):
                c = ws.cell(row=row, column=col, value=dept if col == 2 else "")
                from openpyxl.styles import PatternFill, Font, Border, Side
                c.fill   = PatternFill("solid", start_color=_CLR_ALT)
                c.font   = Font(name="Arial", bold=True, size=10, color=_CLR_HDR_BG)
                c.border = _rpt_thin()
            ws.row_dimensions[row].height = 17
            row += 1; prev_dept = dept
        bg = "EBF3FB" if i % 2 == 0 else "FFFFFF"
        _rpt_data_cell(ws, row, 1, rec["Emp Code"],   bg=bg, align="center")
        _rpt_data_cell(ws, row, 2, rec["Name"],       bg=bg)
        _rpt_data_cell(ws, row, 3, rec["Department"], bg=bg)
        _rpt_data_cell(ws, row, 4, rec["Date"],       bg=bg, align="center")
        ws.row_dimensions[row].height = 16
        row += 1

    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 14
    wb.save(output_path)


def _pull_punches_from_devices():
    from datetime import date as _date
    try: from zk import ZK
    except ImportError:
        print("[ERROR] pyzk not installed: pip install pyzk"); sys.exit(1)
    today = _date.today(); present_badges = set()
    for ip in DEVICE_IPS:
        from zk import ZK
        zk = ZK(ip, port=DEVICE_PORT, timeout=DEVICE_TIMEOUT, verbose=False); conn = None
        try:
            print("  Connecting to {0} ...".format(ip), end=" ")
            sys.stdout.flush()
            conn = zk.connect(); conn.disable_device()
            records = conn.get_attendance(); hits = 0
            for r in records:
                if r.timestamp.date() == today:
                    present_badges.add(str(r.user_id).strip()); hits += 1
            conn.enable_device()
            print("OK  ({0} punches today)".format(hits))
        except Exception as e:
            print("FAILED -- {0}".format(e))
        finally:
            if conn:
                try: conn.disconnect()
                except: pass
    return present_badges


def cmd_today():
    """python mdb_tools.py today"""
    from datetime import date as _date
    today = _date.today()
    print("\n{0}\n  TODAY'S ABSENT REPORT  --  {1}\n{0}".format(
        "=" * 55, today.strftime('%d %B %Y')))
    print("\nLoading employees...")
    emp_df = _load_employees()
    print("\nPulling punches from {0} device(s)...\n".format(len(DEVICE_IPS)))
    present_badges = _pull_punches_from_devices()
    absent_rows = []
    for _, emp in emp_df.iterrows():
        if emp["Badgenumber"] not in present_badges:
            absent_rows.append({
                "Emp Code":   emp["Badgenumber"],
                "Name":       emp["Name"],
                "Department": emp["DEPTNAME"],
                "Date":       today.strftime("%d/%m/%Y"),
            })
    print("\nPresent: {0}  |  Absent: {1}".format(len(present_badges), len(absent_rows)))
    _write_absent_excel(absent_rows, OUTPUT_TODAY, title_suffix=today.strftime("%d %B %Y"))
    print("[OK] Report saved -> {0}".format(OUTPUT_TODAY))


def cmd_history(argv):
    """python mdb_tools.py history backup.mdb DD/MM/YYYY DD/MM/YYYY"""
    import pandas as pd
    from datetime import datetime as _dt, date as _date
    if len(argv) < 3:
        print("[ERROR] Usage: python mdb_tools.py history backup.mdb DD/MM/YYYY DD/MM/YYYY")
        sys.exit(1)
    mdb_path = argv[0]
    try:
        date_from = _dt.strptime(argv[1], "%d/%m/%Y").date()
        date_to   = _dt.strptime(argv[2], "%d/%m/%Y").date()
    except ValueError:
        print("[ERROR] Invalid date format. Use DD/MM/YYYY"); sys.exit(1)

    print("\n{0}\n  ABSENT HISTORY REPORT\n  {1} -> {2}\n{0}\n".format(
        "=" * 55, date_from.strftime('%d %b %Y'), date_to.strftime('%d %b %Y')))
    print("Loading employees...")
    emp_df = _load_employees()
    print("Loading punch records from MDB...")
    conn     = connect_mdb(mdb_path)
    uid_map  = _mdb_uid_to_badge(conn)
    fs = date_from.strftime("%Y-%m-%d"); ts = date_to.strftime("%Y-%m-%d")
    sql = (
        "SELECT [USERID],[CHECKTIME] FROM [CHECKINOUT] "
        "WHERE [CHECKTIME] >= #{0}# AND [CHECKTIME] <= #{1}#".format(
            fs + " 00:00:00", ts + " 23:59:59")
    )
    try:
        punch_df = pd.read_sql(sql, conn)
    except Exception as e:
        print("[ERROR] CHECKINOUT query failed: {0}".format(e)); conn.close(); sys.exit(1)
    conn.close()

    punch_df["CHECKTIME"] = pd.to_datetime(punch_df["CHECKTIME"], errors="coerce")
    punch_df = punch_df.dropna(subset=["CHECKTIME"])
    punch_df["_date"]  = punch_df["CHECKTIME"].dt.date
    punch_df["_badge"] = punch_df["USERID"].astype(str).str.strip().map(uid_map).fillna("")
    print("Punch records in range: {0}".format(len(punch_df)))

    absent_rows = []
    for d in pd.date_range(date_from, date_to).date:
        day_df  = punch_df[punch_df["_date"] == d].copy()
        # Deduplicate: per badge keep only punches >= 60s apart, then check presence
        present = set()
        for badge, grp in day_df.groupby("_badge"):
            times = sorted(grp["CHECKTIME"].dt.strftime("%H:%M:%S").tolist())
            if _dedupe_times(times):  # if any real punches remain after dedup
                present.add(badge)
        for _, emp in emp_df.iterrows():
            if emp["Badgenumber"] not in present:
                absent_rows.append({
                    "Emp Code":   emp["Badgenumber"],
                    "Name":       emp["Name"],
                    "Department": emp["DEPTNAME"],
                    "Date":       d.strftime("%d/%m/%Y"),
                })

    print("Total absent records: {0}".format(len(absent_rows)))
    label = "{0} - {1}".format(date_from.strftime('%d %b'), date_to.strftime('%d %b %Y'))
    _write_absent_excel(absent_rows, OUTPUT_HISTORY, title_suffix=label)
    print("[OK] Report saved -> {0}".format(OUTPUT_HISTORY))


# ==============================================================================
#  ENTRY POINT
# ==============================================================================

def _usage():
    print("""
ZKTeco MDB Tools  --  Usage
===========================

  python mdb_tools.py scan                                        # scan MDB layout
  python mdb_tools.py scan path\\backup.mdb                       # scan specific file

  python mdb_tools.py export                                      # export employees
  python mdb_tools.py export path\\backup.mdb                     # from specific MDB

  python mdb_tools.py today                                       # today absent report
  python mdb_tools.py history backup.mdb 01/03/2026 12/03/2026   # date range report
""")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        _usage(); sys.exit(0)

    mode = sys.argv[1].lower()
    rest = sys.argv[2:]

    if mode == "scan":
        cmd_scan(rest)
    elif mode == "export":
        cmd_export(rest)
    elif mode == "today":
        cmd_today()
    elif mode == "history":
        cmd_history(rest)
    else:
        print("[ERROR] Unknown command: {0}".format(mode))
        _usage(); sys.exit(1)
